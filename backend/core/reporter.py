import datetime
import html
import json
import os
import re
import shutil
import threading
import time
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font
except ImportError:
    load_workbook = None


class Reporter:
    """Render HTML/Excel/annotated TXT outputs."""

    _template_bytes_cache = {}
    _template_cache_lock = threading.Lock()

    CLIENT_TEMPLATE = "(코드리뷰결과서-Client) 코드 리뷰 결과서 양식_v2.0_20251201.xlsx"
    SERVER_TEMPLATE = "(코드리뷰결과서-Server) 코드 리뷰 결과서 양식_v2.0_20251104.xlsx"

    UNMATCHED_SHEET_NAME = "미분류_위반사항"
    DETAIL_SHEET_NAME = "상세결과"
    VERIFY_SHEET_NAME = "검증 결과"
    CTRLPP_SHEET_NAME = "CtrlppCheck_결과"
    VERIFY_META_SHEET_NAME = "검증메타"

    MANUAL_REVIEW_MESSAGE = "자동 체크 불가 (수동 확인 권장)"
    RULE_ID_ITEM_MAP = {
        "PERF-EV-01": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
        "PERF-DPSET-CHAIN": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
        "PERF-DPSET-BATCH-01": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
        "STYLE-NAME-01": "명명 규칙 및 코딩 스타일 준수 확인",
        "STYLE-INDENT-01": "명명 규칙 및 코딩 스타일 준수 확인",
        "STYLE-HEADER-01": "명명 규칙 및 코딩 스타일 준수 확인",
        "STD-01": "명명 규칙 및 코딩 스타일 준수 확인",
        "UNUSED-01": "불필요한 코드 지양",
        "CLEAN-DEAD-01": "불필요한 코드 지양",
        "CLEAN-DUP-01": "불필요한 코드 지양",
        "COMP-01": "불필요한 코드 지양",
        "COMP-02": "불필요한 코드 지양",
        "HARD-01": "하드코딩 지양",
        "HARD-02": "하드코딩 지양",
        "CFG-01": "config 항목 정합성 확인",
        "CFG-ERR-01": "config 항목 정합성 확인",
        "SAFE-DIV-01": "config 항목 정합성 확인",
        "EXC-TRY-01": "불필요한 코드 지양",
        "PERF-DPGET-BATCH-01": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
        "PERF-SETVALUE-BATCH-01": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
        "PERF-SETMULTIVALUE-ADOPT-01": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
        "PERF-GETVALUE-BATCH-01": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
        "PERF-AGG-01": "불필요한 코드 지양",
        "STYLE-IDX-01": "명명 규칙 및 코딩 스타일 준수 확인",
        "HARD-03": "하드코딩 지양",
    }

    def __init__(self, config_dir: str = None):
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        self.config_dir = config_dir or os.path.join(base_dir, "..", "Config")

        self.output_base_dir = os.path.join(base_dir, "..", "CodeReview_Report")
        self.timestamp = ""
        self.output_dir = ""
        self.start_session()

        self.rules_path = os.path.join(self.config_dir, "parsed_rules.json")
        self.parsed_rules = self._load_json(self.rules_path, default=[])

        self.applicability_path = os.path.join(self.config_dir, "review_applicability.json")
        self.applicability = self._load_json(self.applicability_path, default={})
        self._item_policy_map = self._build_item_policy_map()
        self.last_excel_metrics = {}

    def _ensure_output_dir(self):
        if not self.output_dir:
            self.start_session()
        os.makedirs(self.output_dir, exist_ok=True)

    def start_session(self):
        # Start a new report session only when an analysis run begins.
        self.timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        self.output_dir = os.path.join(self.output_base_dir, self.timestamp)

    @staticmethod
    def _load_json(path: str, default):
        try:
            if not os.path.exists(path):
                return default
            with open(path, "r", encoding="utf-8") as f:
                payload = json.load(f)
            return payload
        except Exception:
            return default

    @staticmethod
    def _normalize_text(text: str) -> str:
        return re.sub(r"[\s\n\r]+", "", str(text or "").lower())

    @staticmethod
    def _escape(value) -> str:
        return html.escape(str(value or ""))

    @staticmethod
    def _to_int(value, fallback: int = 0) -> int:
        try:
            return int(value)
        except (TypeError, ValueError):
            return fallback

    @staticmethod
    def is_excel_support_available() -> bool:
        return load_workbook is not None

    @staticmethod
    def _perf_now() -> float:
        return time.perf_counter()

    @classmethod
    def _elapsed_ms(cls, started: float) -> int:
        return max(0, int((cls._perf_now() - started) * 1000))

    @staticmethod
    def _file_signature(path: str) -> Tuple[int, int]:
        st = os.stat(path)
        return (int(getattr(st, "st_mtime_ns", int(st.st_mtime * 1_000_000_000))), int(st.st_size))

    @classmethod
    def _read_template_bytes_cached(cls, template_path: str):
        if not template_path or not os.path.exists(template_path):
            return None, False
        key = os.path.normpath(template_path)
        sig = cls._file_signature(template_path)
        with cls._template_cache_lock:
            cached = cls._template_bytes_cache.get(key)
            if isinstance(cached, dict) and tuple(cached.get("sig", ())) == sig and isinstance(cached.get("bytes"), bytes):
                return cached["bytes"], True
        with open(template_path, "rb") as f:
            data = f.read()
        with cls._template_cache_lock:
            cls._template_bytes_cache[key] = {"sig": sig, "bytes": data}
        return data, False

    @staticmethod
    def _severity_rank(value: str) -> int:
        order = {
            "critical": 0,
            "high": 1,
            "warning": 2,
            "medium": 3,
            "low": 4,
            "info": 5,
            "information": 5,
        }
        return order.get(str(value or "").strip().lower(), 6)

    @staticmethod
    def _normalize_comment_message(message: str) -> str:
        text = str(message or "").strip()
        text = re.sub(r"(미사용 변수 감지:\s*)['\"]([^'\"]+)['\"]", r"\1\2", text)
        return text

    def _build_suggestion_text(self, violation: Dict) -> str:
        raw = str(violation.get("suggestion") or "").strip()
        if raw:
            return raw

        rule_id = str(violation.get("rule_id") or "").upper()
        message = self._normalize_comment_message(violation.get("message", ""))

        if rule_id == "UNUSED-01":
            match = re.search(r"미사용 변수 감지:\s*([a-zA-Z_][a-zA-Z0-9_]*)", message)
            var_name = match.group(1) if match else "변수"
            return f"// 사용되지 않는 변수 {var_name} 삭제 권장"
        if rule_id == "EXC-DP-01":
            return "// try/catch 또는 getLastError 기반 예외 처리 추가 권장"
        if rule_id == "PERF-DPSET-CHAIN":
            return "// 연속 dpSet 대신 조건 비교/배치(dpSetWait, dpSetTimed) 권장"
        if rule_id == "HARD-01":
            return "// 설정 파일 또는 상수로 대체 권장"
        return "// 규칙에 맞는 수정 로직 반영 권장"

    def _render_todo_review_block(self, violation: Dict, indent: str) -> List[str]:
        severity = str(violation.get("severity", "Info"))
        raw_message = str(violation.get("message", "") or "")
        message = self._normalize_comment_message(raw_message)
        suggestion = self._build_suggestion_text(violation)
        suggestion_text = self._normalize_comment_message(suggestion)
        if suggestion_text.startswith("//"):
            suggestion_text = suggestion_text[2:].strip()
        issue_id = str(violation.get("issue_id", "") or "").replace(";", ",")
        rule_id = str(violation.get("rule_id", "") or "").replace(";", ",")
        line_no = self._to_int(violation.get("line", 0), 0)
        file_name = str(violation.get("file", "") or violation.get("object", "") or "").replace(";", ",")
        review_suffix = f" - {suggestion_text}" if suggestion_text and suggestion_text != message else ""
        lines = [
            f"{indent}// >>TODO",
            f"{indent}// {message}",
            f"{indent}// [REVIEW] {severity}{review_suffix}",
            f"{indent}// [META] issue_id={issue_id}; rule_id={rule_id}; line={line_no}; file={file_name}",
        ]
        for ai_review in violation.get("_accepted_ai_reviews", []) or []:
            lines.extend(self._render_ai_code_comment_lines(ai_review, indent))
        return lines

    @staticmethod
    def _extract_ai_code_lines(review_text: str, max_lines: int = 8) -> List[str]:
        text = str(review_text or "")
        blocks = re.findall(r"```(?:[A-Za-z0-9_+-]+)?\s*\n(.*?)```", text, flags=re.DOTALL)
        code = ""
        if blocks:
            code = blocks[0]
        else:
            # Fallback: try the section after "코드:" label.
            m = re.search(r"코드\s*:\s*(.*)$", text, flags=re.DOTALL)
            if m:
                code = m.group(1)
        lines = []
        for line in str(code).splitlines():
            stripped = line.rstrip()
            if not stripped.strip():
                continue
            lines.append(stripped)
            if len(lines) >= max_lines:
                break
        return lines

    def _render_ai_code_comment_lines(self, ai_review: Dict, indent: str) -> List[str]:
        code_lines = self._extract_ai_code_lines(ai_review.get("review", ""))
        return [f"{indent}// [AI CODE] {line}" for line in code_lines]

    def _flatten_internal_violations(self, report_data: dict) -> List[Dict]:
        flattened = []
        for item in report_data.get("internal_violations", []):
            obj = item.get("object", "Unknown")
            event = item.get("event", "Unknown")
            for violation in item.get("violations", []):
                merged = dict(violation)
                merged.setdefault("object", obj)
                merged.setdefault("event", event)
                merged.setdefault("file", report_data.get("file", obj))
                flattened.append(merged)
        return flattened

    def _flatten_global_violations(self, report_data: dict) -> List[Dict]:
        flattened = []
        default_file = report_data.get("file", "")
        for violation in report_data.get("global_violations", []):
            merged = dict(violation)
            merged.setdefault("file", default_file)
            merged.setdefault("source", "CtrlppCheck")
            flattened.append(merged)
        return flattened

    @staticmethod
    def _replace_sheet(wb, sheet_name: str, headers: List[str], rows: List[List]):
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        for row in rows:
            ws.append(row)

    def _build_item_policy_map(self) -> Dict[str, Dict]:
        raw_items = self.applicability.get("items", {})
        if not isinstance(raw_items, dict):
            return {}

        normalized = {}
        for item_name, cfg in raw_items.items():
            if not isinstance(cfg, dict):
                continue
            normalized[self._normalize_text(item_name)] = cfg
        return normalized

    def _collect_rule_ids(self, violations: List[Dict]) -> set:
        return {str(v.get("rule_id") or "").strip() for v in violations if str(v.get("rule_id") or "").strip()}

    def _is_manual_na_item(self, item_text: str, condition_text: str) -> bool:
        norm_item = self._normalize_text(item_text)
        norm_condition = self._normalize_text(condition_text)

        manual_items = self.applicability.get("manual_only_items", [])
        for item in manual_items if isinstance(manual_items, list) else []:
            if self._normalize_text(item) == norm_item and norm_item:
                return True

        manual_keywords = self.applicability.get("manual_condition_keywords", [])
        for keyword in manual_keywords if isinstance(manual_keywords, list) else []:
            norm_keyword = self._normalize_text(keyword)
            if norm_keyword and norm_keyword in norm_condition:
                return True
        return False

    def _missing_required_signals(self, item_text: str, all_rule_ids: set) -> bool:
        cfg = self._item_policy_map.get(self._normalize_text(item_text))
        if not cfg:
            return False

        required = cfg.get("required_rule_ids", [])
        if not isinstance(required, list) or not required:
            return False

        required_set = {str(v) for v in required if str(v).strip()}
        return not bool(all_rule_ids & required_set)

    def _match_row_violations(self, condition_text: str, item_text: str, violations: List[Dict]) -> List[Dict]:
        norm_condition = self._normalize_text(condition_text)
        norm_item = self._normalize_text(item_text)
        matched = []
        seen = set()

        for violation in violations:
            rule_item = str(violation.get("rule_item", "") or "")
            rule_id = str(violation.get("rule_id", "") or "")
            norm_rule_item = self._normalize_text(rule_item)
            mapped_item = self.RULE_ID_ITEM_MAP.get(rule_id, "")
            norm_mapped_item = self._normalize_text(mapped_item)

            is_match = False
            if norm_mapped_item and norm_mapped_item == norm_item:
                is_match = True
            elif rule_item and ((rule_item in condition_text) or (rule_item in item_text)):
                is_match = True
            elif norm_rule_item and (
                (norm_rule_item == norm_item)
                or (norm_rule_item in norm_condition)
                or (norm_rule_item in norm_item)
            ):
                is_match = True
            elif rule_id and (rule_id in condition_text or rule_id in item_text):
                is_match = True

            if not is_match:
                continue

            key = violation.get("issue_id") or f"{rule_id}:{violation.get('line', 0)}:{violation.get('message', '')}"
            if key in seen:
                continue
            seen.add(key)
            matched.append(violation)

        return matched

    def _evaluate_status(
        self,
        item_text: str,
        condition_text: str,
        matched_p1: List[Dict],
        all_p1_rule_ids: set,
        file_name: str = "",
    ) -> Tuple[str, str]:
        # Config checklist is considered applicable only for config-like targets.
        if self._normalize_text(item_text) == self._normalize_text("config 항목 정합성 확인"):
            if not re.search(r"(config|cfg|ini|json)", str(file_name or ""), re.IGNORECASE):
                return "N/A", self.MANUAL_REVIEW_MESSAGE

        if matched_p1:
            remark = "\n".join(
                f"[{v.get('rule_id', '')}] {v.get('message', '')}" for v in matched_p1
            )
            return "NG", remark

        if self._is_manual_na_item(item_text, condition_text):
            return "N/A", self.MANUAL_REVIEW_MESSAGE

        if self._missing_required_signals(item_text, all_p1_rule_ids):
            return "N/A", self.MANUAL_REVIEW_MESSAGE

        return "OK", ""

    @staticmethod
    def _apply_status_style(cell, status: str):
        if status == "NG":
            cell.font = Font(color="FF0000", bold=True)
        elif status == "OK":
            cell.font = Font(color="0000FF")
        else:
            cell.font = Font(color="333333")

    def generate_html_report(self, report_data: dict, filename: str = "analysis_report.html", report_meta: dict = None):
        self._ensure_output_dir()
        output_path = os.path.join(self.output_dir, filename)
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        severity_ko = {
            "critical": "치명",
            "high": "높음",
            "warning": "경고",
            "medium": "보통",
            "low": "낮음",
            "info": "정보",
            "error": "오류",
            "style": "스타일",
            "performance": "성능",
            "portability": "이식성",
            "information": "정보",
        }
        severity_css_map = {
            "critical": "critical",
            "high": "high",
            "warning": "warning",
            "medium": "medium",
            "low": "low",
            "info": "info",
            "error": "critical",
            "style": "warning",
            "performance": "medium",
            "portability": "low",
            "information": "info",
        }

        title = "WinCC OA 코드 분석 결과"
        target = self._escape(report_data.get("file", ""))
        meta = report_meta if isinstance(report_meta, dict) else {}
        verification_level = str(meta.get("verification_level", "UNKNOWN") or "UNKNOWN")
        optional_deps = meta.get("optional_dependencies", {}) if isinstance(meta.get("optional_dependencies", {}), dict) else {}
        openpyxl_status = "available" if bool((optional_deps.get("openpyxl") or {}).get("available", False)) else "missing"

        html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="UTF-8">
  <meta http-equiv="X-UA-Compatible" content="IE=edge">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>{title}</title>
  <style>
    body {{ font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif; padding: 20px; }}
    table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; vertical-align: top; }}
    th {{ background-color: #f2f2f2; }}
    .critical {{ color: #fff; background-color: #d32f2f; padding: 2px 6px; border-radius: 3px; font-weight: bold; }}
    .high {{ color: #fff; background-color: #e65100; padding: 2px 6px; border-radius: 3px; font-weight: bold; }}
    .warning {{ color: #111; background-color: #ffb74d; padding: 2px 6px; border-radius: 3px; }}
    .medium {{ color: #fff; background-color: #1976d2; padding: 2px 6px; border-radius: 3px; }}
    .low {{ color: #fff; background-color: #388e3c; padding: 2px 6px; border-radius: 3px; }}
    .info {{ color: #555; background-color: #e0e0e0; padding: 2px 6px; border-radius: 3px; }}
  </style>
</head>
<body>
  <h1>{title}</h1>
  <p><strong>생성 일시:</strong> {now}</p>
  <p><strong>대상:</strong> {target}</p>
  <p><strong>검증 레벨:</strong> {self._escape(verification_level)} (openpyxl: {self._escape(openpyxl_status)})</p>
  <h2>1. 정적분석 결과 (P1)</h2>
  <table>
    <tr><th>객체</th><th>이벤트</th><th>심각도</th><th>규칙 항목</th><th>상세 메시지</th></tr>
"""

        for item in report_data.get("internal_violations", []):
            obj = self._escape(item.get("object", ""))
            event = self._escape(item.get("event", ""))
            for violation in item.get("violations", []):
                sev_raw = str(violation.get("severity", "Info"))
                sev_class = sev_raw.lower()
                sev_text = severity_ko.get(sev_class, sev_raw)
                rule_item = self._escape(violation.get("rule_item", "N/A"))
                message = self._escape(violation.get("message", ""))
                html_content += (
                    "<tr>"
                    f"<td>{obj}</td>"
                    f"<td>{event}</td>"
                    f"<td><span class=\"{sev_class}\">{self._escape(sev_text)}</span></td>"
                    f"<td>{rule_item}</td>"
                    f"<td>{message}</td>"
                    "</tr>"
                )

        html_content += "</table>"

        global_violations = report_data.get("global_violations", [])
        if global_violations:
            html_content += """
  <h2>2. CtrlppCheck 결과 (P2)</h2>
  <table>
    <tr><th>라인</th><th>유형</th><th>규칙 ID</th><th>메시지</th></tr>
"""
            for violation in global_violations:
                v_type = str(violation.get("type", "Info"))
                v_type_class = v_type.lower()
                v_type_text = severity_ko.get(v_type_class, v_type)
                v_type_css = severity_css_map.get(v_type_class, "info")
                line = self._escape(violation.get("line", "-"))
                rule_id = self._escape(violation.get("rule_id", ""))
                msg = self._escape(violation.get("message", ""))
                html_content += (
                    "<tr>"
                    f"<td>{line}</td>"
                    f"<td><span class=\"{v_type_css}\">{self._escape(v_type_text)}</span></td>"
                    f"<td>{rule_id}</td>"
                    f"<td>{msg}</td>"
                    "</tr>"
                )
            html_content += "</table>"

        if report_data.get("ai_reviews"):
            html_content += "<h2>3. AI 리뷰 코멘트 (P3)</h2>"
            for review in report_data.get("ai_reviews", []):
                obj = self._escape(review.get("object", ""))
                event = self._escape(review.get("event", ""))
                text = self._escape(review.get("review", ""))
                html_content += (
                    "<div style=\"border:1px solid #1976D2;border-radius:5px;margin-bottom:15px;"
                    "padding:10px;background-color:#E3F2FD;\">"
                    f"<h3 style=\"margin-top:0;color:#1565C0;\">[{obj}] - {event}</h3>"
                    f"<p style=\"white-space:pre-wrap;font-style:italic;\">{text}</p>"
                    "</div>"
                )

        html_content += "</body></html>"
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_content)

    def generate_annotated_txt(self, original_content: str, report_data: dict, filename: str):
        self._ensure_output_dir()
        output_path = os.path.join(self.output_dir, filename)
        lines = original_content.splitlines()
        violation_map = {}
        top_violations = []
        accepted_ai_map = {}

        for ai_item in report_data.get("ai_reviews", []):
            if str(ai_item.get("status", "")).lower() != "accepted":
                continue
            key = (str(ai_item.get("object", "")), str(ai_item.get("event", "Global")))
            accepted_ai_map.setdefault(key, []).append(ai_item)

        for item in report_data.get("internal_violations", []):
            item_object = str(item.get("object", "Unknown"))
            item_event = str(item.get("event", "Unknown"))
            accepted_ai = accepted_ai_map.get((item_object, item_event), [])
            for violation in item.get("violations", []):
                violation = dict(violation)
                if accepted_ai:
                    violation["_accepted_ai_reviews"] = accepted_ai
                line_no = self._to_int(violation.get("line", 1), 1)
                if 1 <= line_no <= len(lines):
                    violation_map.setdefault(line_no, []).append(violation)
                else:
                    top_violations.append(violation)

        for line_no in list(violation_map.keys()):
            violation_map[line_no] = sorted(
                violation_map[line_no],
                key=lambda item: self._severity_rank(item.get("severity", "")),
            )
        top_violations = sorted(top_violations, key=lambda item: self._severity_rank(item.get("severity", "")))

        annotated_lines = []
        for violation in top_violations:
            annotated_lines.extend(self._render_todo_review_block(violation, ""))
            annotated_lines.append("")

        for i, line in enumerate(lines, 1):
            if i in violation_map:
                for violation in violation_map[i]:
                    indent_match = re.match(r"^(\s*)", line)
                    indent = indent_match.group(1) if indent_match else ""
                    annotated_lines.extend(self._render_todo_review_block(violation, indent))
                    annotated_lines.append("")
            annotated_lines.append(line)

        with open(output_path, "w", encoding="utf-8") as f:
            f.write("\n".join(annotated_lines))

    def fill_excel_checklist(self, report_data: dict, file_type: str = "Client", output_filename: str = None, report_meta: dict = None):
        metrics = {
            "generated": False,
            "template_cache_hit": False,
            "timings_ms": {"total": 0, "copy": 0, "load": 0, "save": 0},
            "output_path": "",
        }
        total_started = self._perf_now()
        if load_workbook is None:
            self.last_excel_metrics = dict(metrics)
            return metrics
        self._ensure_output_dir()

        template_name = self.CLIENT_TEMPLATE if file_type == "Client" else self.SERVER_TEMPLATE
        template_path = os.path.join(self.config_dir, template_name)
        if not os.path.exists(template_path):
            self.last_excel_metrics = dict(metrics)
            return metrics

        if not output_filename:
            output_filename = f"CodeReview_Submission_{file_type}_{self.timestamp}.xlsx"

        output_path = os.path.join(self.output_dir, output_filename)
        metrics["output_path"] = output_path
        copy_started = self._perf_now()
        template_bytes, cache_hit = self._read_template_bytes_cached(template_path)
        if isinstance(template_bytes, bytes):
            with open(output_path, "wb") as out:
                out.write(template_bytes)
            metrics["template_cache_hit"] = bool(cache_hit)
        else:
            shutil.copy2(template_path, output_path)
        metrics["timings_ms"]["copy"] = self._elapsed_ms(copy_started)

        load_started = self._perf_now()
        wb = load_workbook(output_path)
        metrics["timings_ms"]["load"] = self._elapsed_ms(load_started)
        ws = wb.active

        meta = report_meta if isinstance(report_meta, dict) else {}
        verification_level = str(meta.get("verification_level", "CORE_ONLY") or "CORE_ONLY")
        optional_deps = meta.get("optional_dependencies", {}) if isinstance(meta.get("optional_dependencies", {}), dict) else {}
        openpyxl_status = "available" if bool((optional_deps.get("openpyxl") or {}).get("available", False)) else "missing"

        p1_violations = self._flatten_internal_violations(report_data)
        p2_violations = self._flatten_global_violations(report_data)
        all_p1_rule_ids = self._collect_rule_ids(p1_violations)

        start_row = 16
        for row in range(1, 40):
            value = ws.cell(row=row, column=2).value
            if value and "대분류" in str(value):
                start_row = row + 1
                break

        header_row = max(1, start_row - 1)
        result_col = 6
        remarks_col = 7
        for col in range(1, min(ws.max_column, 20) + 1):
            header_value = str(ws.cell(row=header_row, column=col).value or "").strip()
            if "검증 결과" in header_value:
                result_col = col
                remarks_col = col + 1
                break

        matched_issue_ids = set()
        verification_rows = []
        verify_no = 0

        for row in range(start_row, ws.max_row + 1):
            cat_large = ws.cell(row=row, column=2).value
            cat_mid = ws.cell(row=row, column=3).value
            item = ws.cell(row=row, column=4).value
            condition = ws.cell(row=row, column=5).value

            if not condition and not cat_mid:
                continue

            verify_no += 1
            item_text = str(item or "").strip()
            condition_text = str(condition or "").strip()

            matched = self._match_row_violations(condition_text, item_text, p1_violations)
            for violation in matched:
                issue_id = violation.get("issue_id")
                if issue_id:
                    matched_issue_ids.add(issue_id)

            status, remarks = self._evaluate_status(
                item_text,
                condition_text,
                matched,
                all_p1_rule_ids,
                file_name=str(report_data.get("file", "") or ""),
            )

            result_cell = ws.cell(row=row, column=result_col)
            remarks_cell = ws.cell(row=row, column=remarks_col)
            result_cell.value = status
            remarks_cell.value = remarks
            remarks_cell.alignment = Alignment(wrapText=True)
            self._apply_status_style(result_cell, status)

            verification_rows.append(
                {
                    "No": verify_no,
                    "대분류": cat_large,
                    "중분류": cat_mid,
                    "소분류": item,
                    "검증 조건": condition_text,
                    "중요도": "필수",
                    "검증 결과": status,
                    "비고": remarks or "",
                }
            )

        unmatched = []
        for violation in p1_violations:
            issue_id = violation.get("issue_id")
            if issue_id and issue_id in matched_issue_ids:
                continue
            unmatched.append(violation)

        if unmatched:
            if self.UNMATCHED_SHEET_NAME not in wb.sheetnames:
                ws_extra = wb.create_sheet(self.UNMATCHED_SHEET_NAME)
                ws_extra.append(["Object", "Event", "Rule ID", "Rule Item", "Message", "Line"])
            else:
                ws_extra = wb[self.UNMATCHED_SHEET_NAME]

            for violation in unmatched:
                ws_extra.append(
                    [
                        violation.get("object", "Unknown"),
                        violation.get("event", "Unknown"),
                        violation.get("rule_id", ""),
                        violation.get("rule_item", ""),
                        violation.get("message", ""),
                        violation.get("line", ""),
                    ]
                )

        detail_headers = ["파일명", "라인", "규칙ID", "중요도", "위반항목", "문제코드", "해결방법"]
        detail_rows = []
        for violation in p1_violations:
            detail_rows.append(
                [
                    os.path.basename(str(violation.get("file", report_data.get("file", "Unknown")))),
                    self._to_int(violation.get("line", 0), 0),
                    violation.get("rule_id", ""),
                    violation.get("severity", ""),
                    violation.get("message", ""),
                    violation.get("code", "N/A"),
                    self._build_suggestion_text(violation),
                ]
            )
        self._replace_sheet(wb, self.DETAIL_SHEET_NAME, detail_headers, detail_rows)

        verify_headers = ["No", "대분류", "중분류", "소분류", "검증 조건", "중요도", "검증 결과", "비고"]
        verify_rows = [
            [
                row.get("No"),
                row.get("대분류"),
                row.get("중분류"),
                row.get("소분류"),
                row.get("검증 조건"),
                row.get("중요도"),
                row.get("검증 결과"),
                row.get("비고"),
            ]
            for row in verification_rows
        ]
        self._replace_sheet(wb, self.VERIFY_SHEET_NAME, verify_headers, verify_rows)

        ctrlpp_headers = ["파일명", "라인", "규칙ID", "심각도", "메시지", "source"]
        ctrlpp_rows = [
            [
                os.path.basename(str(v.get("file", report_data.get("file", "Unknown")))),
                self._to_int(v.get("line", 0), 0),
                v.get("rule_id", ""),
                v.get("severity", v.get("type", "")),
                v.get("message", ""),
                v.get("source", "CtrlppCheck"),
            ]
            for v in p2_violations
        ]
        self._replace_sheet(wb, self.CTRLPP_SHEET_NAME, ctrlpp_headers, ctrlpp_rows)

        meta_headers = ["항목", "값"]
        meta_rows = [
            ["verification_level", verification_level],
            ["openpyxl", openpyxl_status],
            ["generated_at", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["file", str(report_data.get("file", ""))],
        ]
        self._replace_sheet(wb, self.VERIFY_META_SHEET_NAME, meta_headers, meta_rows)

        save_started = self._perf_now()
        wb.save(output_path)
        metrics["timings_ms"]["save"] = self._elapsed_ms(save_started)
        metrics["generated"] = True
        metrics["timings_ms"]["total"] = self._elapsed_ms(total_started)
        self.last_excel_metrics = dict(metrics)
        print(f"[*] Excel Submission report generated: {output_path}")
        return metrics
