import json
import os
import subprocess
import sys
import tempfile
import unittest
import html
import shutil
import xml.etree.ElementTree as ET
import zipfile
import glob

# Add backend directory to path to import core modules
sys.path.append(os.path.join(os.path.dirname(os.path.abspath(__file__))))

from core.heuristic_checker import HeuristicChecker
from core.llm_reviewer import LLMReviewer
from core.pnl_parser import PnlParser
from core.xml_parser import XmlParser
from core.ctrl_wrapper import CtrlppWrapper
from core.mcp_context import MCPContextClient
from core.reporter import Reporter
from main import CodeInspectorApp, DEFAULT_MODE


class SystemVerification(unittest.TestCase):
    def setUp(self):
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.project_root = os.path.dirname(self.base_dir)
        self.config_dir = os.path.join(self.project_root, "Config")
        self.data_dir = os.path.join(self.project_root, "CodeReview_Data")

    def test_config_existence(self):
        rules_path = os.path.join(self.config_dir, "parsed_rules.json")
        self.assertTrue(os.path.exists(rules_path), f"parsed_rules.json not found at {rules_path}")

        with open(rules_path, "r", encoding="utf-8-sig") as f:
            rules = json.load(f)
            self.assertIsInstance(rules, list, "Rules should be a list")
            self.assertGreater(len(rules), 0, "Rules should not be empty")

    def test_pnl_parser_complex(self):
        complex_pnl = '''
6 13
"btnStart"
""
1 10 10 E E E 1 E 1 E N "_Transparent" E N "_Transparent" E E
"Clicked" 1
"main()\n{\n    if(1) {\n        dpSet(\\"a\\", 1);\n    }\n}" 0

6 14
"rect"
""
1 20 20 E E E 1 E 1 E N "_Transparent" E N "_Transparent" E E
"Initialize" 1
"main()\n{\n    // Nested\n}" 0
'''
        parser = PnlParser()
        result = parser.normalize_pnl(complex_pnl)
        self.assertEqual(len(result), 2)
        self.assertEqual(result[0]["name"], "btnStart")
        self.assertEqual(result[0]["events"][0]["event"], "Clicked")
        self.assertEqual(result[1]["name"], "rect")

    def test_heuristic_checker_sqli(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = 'sprintf(sql, "SELECT * FROM users WHERE name = \'%s\'", input);'
        violations = checker.check_sql_injection(code)
        self.assertTrue(any(v["rule_id"] == "SEC-01" for v in violations), "SQL Injection not detected")

    def test_configured_composite_patterns_detect_with_standard_escape(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        rows = {str(r.get("rule_id") or r.get("id")): r for r in checker.p1_rule_defs if isinstance(r, dict)}
        samples = {
            "SEC-01": 'main(){ sprintf(sql, "SELECT * FROM t WHERE a = %s", input); }',
            "DB-ERR-01": 'main(){ dpQuery("SELECT * FROM T", result); }',
            "EXC-DP-01": 'main(){ dpSet("A.B.C", 1); }',
        }
        for rid, code in samples.items():
            rule = rows[rid]
            analysis = checker._remove_comments(code)
            findings = checker._run_composite_rule(
                rule,
                code=code,
                analysis_code=analysis,
                event_name="Global",
                base_line=1,
                anchor_line=checker._first_function_line(analysis),
            )
            found_ids = {str(x.get("rule_id", "")) for x in findings}
            self.assertIn(rid, found_ids, f"configured composite did not detect {rid}")

    def test_configured_composite_patterns_detect_with_double_escape(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        rows = {str(r.get("rule_id") or r.get("id")): r for r in checker.p1_rule_defs if isinstance(r, dict)}
        samples = {
            "SEC-01": {
                "code": 'main(){ sprintf(sql, "SELECT * FROM t WHERE a = %s", input); }',
                "patch_key": "sql_keywords_pattern",
            },
            "DB-ERR-01": {
                "code": 'main(){ dpQuery("SELECT * FROM T", result); }',
                "patch_key": "query_call_pattern",
            },
            "EXC-DP-01": {
                "code": 'main(){ dpSet("A.B.C", 1); }',
                "patch_key": "dp_call_pattern",
            },
        }
        for rid, meta in samples.items():
            base_rule = dict(rows[rid])
            detector = dict(base_rule.get("detector") or {})
            key = str(meta["patch_key"])
            detector[key] = str(detector.get(key, "")).replace("\\", "\\\\")
            base_rule["detector"] = detector
            code = str(meta["code"])
            analysis = checker._remove_comments(code)
            findings = checker._run_composite_rule(
                base_rule,
                code=code,
                analysis_code=analysis,
                event_name="Global",
                base_line=1,
                anchor_line=checker._first_function_line(analysis),
            )
            found_ids = {str(x.get("rule_id", "")) for x in findings}
            self.assertIn(rid, found_ids, f"double-escaped detector did not detect {rid}")

    def test_heuristic_checker_unused_var(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            int used = 1;
            int unused = 2;
            DebugN(used);
        }
        """
        violations = checker.check_unused_variables(code)
        self.assertTrue(any("unused" in v["message"] for v in violations), "Unused variable not detected")

    def test_reporter_excel_template_existence(self):
        templates = [name for name in os.listdir(self.config_dir) if name.lower().endswith(".xlsx")]
        self.assertGreaterEqual(len(templates), 2, "Excel templates not found")

    def test_exc_dp_rule_no_false_positive_with_try_catch(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            try {
                dpSet("A.B.C", 1);
            } catch {
                DebugTN(getLastError());
            }
        }
        """
        violations = checker.check_dp_function_exception(code)
        self.assertFalse(any(v["rule_id"] == "EXC-DP-01" for v in violations))

    def test_exc_dp_rule_detects_missing_error_handling(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            dpSet("A.B.C", 1);
        }
        """
        violations = checker.check_dp_function_exception(code)
        self.assertTrue(any(v["rule_id"] == "EXC-DP-01" for v in violations))

    def test_perf_dpset_chain_detected(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            dpSet("A.B.C", 1);
            dpSet("A.B.D", 2);
        }
        """
        violations = checker.check_consecutive_dpset(code)
        self.assertTrue(any(v["rule_id"] == "PERF-DPSET-CHAIN" for v in violations))

    def test_perf_dpset_chain_detects_multiple_clusters_in_one_function(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            dpSet("A.B.C", 1);
            dpSet("A.B.D", 2);
            dpSet("A.B.E", 3);

            int x = 0;
            int y = 0;
            int z = 0;

            dpSet("A.B.F", 4);
            dpSet("A.B.G", 5);
        }
        """
        groups = checker.analyze_raw_code("sample.ctl", code, file_type="Server")
        hits = [v for g in groups for v in g.get("violations", []) if v.get("rule_id") == "PERF-DPSET-CHAIN"]
        self.assertEqual(len(hits), 2, "연속 dpSet 클러스터는 함수 내에서 클러스터당 1건만 검출되어야 함")

    def test_perf_dpset_chain_not_detected_with_dpsetwait_or_delta_guard(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            if (oldVal != newVal) {
                dpSetWait("A.B.C", newVal);
            }
            dpSet("A.B.D", 2);
        }
        """
        violations = checker.check_consecutive_dpset(code)
        self.assertFalse(any(v["rule_id"] == "PERF-DPSET-CHAIN" for v in violations))

    def test_perf03_not_detected_when_loop_has_delay(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            while (1) {
                dpGet("A.B.C", x);
                delay(1);
            }
        }
        """
        violations = checker.check_while_delay_policy(code)
        self.assertFalse(any(v["rule_id"] == "PERF-03" for v in violations))

    def test_perf03_active_delay_rule_detects_delay_only_inside_active(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code_ng = """
        main() {
          while (1) {
            if (isActive) {
              dpSet("A.B.C", 1);
              delay(1);
            }
          }
        }
        """
        code_ok = """
        main() {
          while (1) {
            if (isActive) {
              dpSet("A.B.C", 1);
            }
            delay(1);
          }
        }
        """
        ng_groups = checker.analyze_raw_code("sample.ctl", code_ng, file_type="Server")
        ok_groups = checker.analyze_raw_code("sample.ctl", code_ok, file_type="Server")
        ng_ids = [v["rule_id"] for g in ng_groups for v in g.get("violations", [])]
        ok_ids = [v["rule_id"] for g in ok_groups for v in g.get("violations", [])]
        self.assertIn("PERF-03-ACTIVE-DELAY-01", ng_ids)
        self.assertNotIn("PERF-03-ACTIVE-DELAY-01", ok_ids)

    def test_perf02_where_dpt_in_rule_detected(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          dpQuery("SELECT '_online.._value' FROM '*.**' WHERE _DPT IN ('AI','DI')", result);
        }
        """
        groups = checker.analyze_raw_code("sample.ctl", code, file_type="Server")
        rule_ids = [v["rule_id"] for g in groups for v in g.get("violations", [])]
        self.assertIn("PERF-02-WHERE-DPT-IN-01", rule_ids)

    def test_log_level_rule_detects_missing_debug_level_and_skips_when_present(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code_ng = """
        main() {
          catch {
            writeLog("Script", "failed", LV_WARN);
          }
        }
        """
        code_ok = """
        main() {
          catch {
            writeLog("Script", "failed", LV_WARN);
            DebugTN("dbg2");
          }
        }
        """
        ng_groups = checker.analyze_raw_code("sample.ctl", code_ng, file_type="Server")
        ok_groups = checker.analyze_raw_code("sample.ctl", code_ok, file_type="Server")
        ng_ids = [v["rule_id"] for g in ng_groups for v in g.get("violations", [])]
        ok_ids = [v["rule_id"] for g in ok_groups for v in g.get("violations", [])]
        self.assertIn("LOG-LEVEL-01", ng_ids)
        self.assertNotIn("LOG-LEVEL-01", ok_ids)

    def test_perf05_context_only_alarm_like_dpset(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            dpSet("A.B.PVLAST", val);
        }
        """
        violations = checker.check_dpset_timed_context(code)
        self.assertFalse(any(v["rule_id"] == "PERF-05" for v in violations))

    def test_mem01_not_detected_when_dynremove_present(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            dyn_float buffer;
            while (1) {
                dynAppend(buffer, 1.0);
                dynRemove(buffer, 1);
            }
        }
        """
        violations = checker.check_memory_leaks_advanced(code)
        self.assertFalse(any(v["rule_id"] == "MEM-01" for v in violations))

    def test_std01_not_detected_for_multiline_declaration(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            string logMsg = "A"
                            + "B"
                            + "C";
        }
        """
        violations = checker.check_coding_standards_advanced(code)
        self.assertFalse(any(v["rule_id"] == "STD-01" for v in violations))

    def test_event_exchange_ng_detected_in_looped_dpset_without_guard(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          while(1) {
            dpSet("SYS.A.B", 1);
            dpSet("SYS.A.C", 2);
          }
        }
        """
        violations = checker.check_event_exchange_minimization(code)
        self.assertTrue(any(v["rule_id"] == "PERF-EV-01" for v in violations))

    def test_event_exchange_not_detected_with_batch_or_delta_guard(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          while(1) {
            if (oldVal != newVal) {
              dpSetWait("SYS.A.B", newVal);
            }
            dpSet("SYS.A.C", newVal);
          }
        }
        """
        violations = checker.check_event_exchange_minimization(code)
        self.assertFalse(any(v["rule_id"] == "PERF-EV-01" for v in violations))

    def test_style_name_rule_detects_prefix_violations(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        int globalValue = 0;
        const int maxCount = 10;
        main() {
          int configValue = 1;
        }
        """
        violations = checker.check_style_name_rules(code)
        self.assertTrue(any(v["rule_id"] == "STYLE-NAME-01" for v in violations))

    def test_style_indent_rule_detects_mixed_indent(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = "main() {\n\tint a = 0;\n  int b = 1;\n}\n"
        violations = checker.check_style_indent_rules(code)
        self.assertTrue(any(v["rule_id"] == "STYLE-INDENT-01" for v in violations))

    def test_hard02_detects_repeated_literal_dp_paths(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          dpSet("SITE.AREA.TAG.VALUE", 1);
          dpGet("SITE.AREA.TAG.VALUE", x);
        }
        """
        violations = checker.check_hardcoding_extended(code)
        self.assertTrue(any(v["rule_id"] == "HARD-02" for v in violations))

    def test_clean_dead_code_detected(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          if (false) {
            DebugN("dead");
          }
        }
        """
        violations = checker.check_dead_code(code)
        self.assertTrue(any(v["rule_id"] == "CLEAN-DEAD-01" for v in violations))

    def test_clean_duplicate_block_detected(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          dpSet("A.B.C", 1);
          dpSet("A.B.C", 1);
          dpSet("A.B.C", 1);
        }
        """
        violations = checker.check_duplicate_blocks(code)
        self.assertTrue(any(v["rule_id"] == "CLEAN-DUP-01" for v in violations))

    def test_dup_act_reports_one_per_cluster_and_multiple_clusters(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          setValue("OBJ_A","visible",true);
          setValue("OBJ_A","visible",false);
          setValue("OBJ_A","visible",true);
          setValue("OBJ_A","visible",false);

          int sep1 = 0;
          int sep2 = 0;
          int sep3 = 0;
          int sep4 = 0;
          int sep5 = 0;
          int sep6 = 0;
          int sep7 = 0;
          int sep8 = 0;
          int sep9 = 0;
          int sep10 = 0;
          int sep11 = 0;

          setValue("OBJ_A","visible",true);
          setValue("OBJ_A","visible",false);

          setValue("OBJ_B","enabled",true);
          setValue("OBJ_B","enabled",false);
          setValue("OBJ_B","enabled",true);
        }
        """
        groups = checker.analyze_raw_code("sample.txt", code, file_type="Client")
        hits = [v for g in groups for v in g.get("violations", []) if v.get("rule_id") == "DUP-ACT-01"]
        self.assertEqual(len(hits), 2, "DUP-ACT-01은 같은 연속 구간에서는 1건만 검출되어야 함")
        self.assertTrue(any("OBJ_A.visible" in str(v.get("message", "")) for v in hits), "DUP-ACT-01 메시지에 대상명+속성명이 포함되어야 함")

    def test_dup_act_skips_guarded_duplicate_pattern(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          if (prev != value) {
            setValue("OBJ_A","visible",value);
            setValue("OBJ_A","visible",value);
          }
        }
        """
        groups = checker.analyze_raw_code("sample.txt", code, file_type="Client")
        hits = [v for g in groups for v in g.get("violations", []) if v.get("rule_id") == "DUP-ACT-01"]
        self.assertEqual(len(hits), 0)

    def test_cfg01_detects_inconsistent_config_parsing(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        bool load_config() {
          string config_line = lineA;
          dyn_string a = strsplit(lineA, ",");
          dyn_string b = strsplit(lineB, ";");
          return true;
        }
        """
        violations = checker.check_config_format_consistency(code)
        self.assertTrue(any(v["rule_id"] == "CFG-01" for v in violations))

    def test_cfg_err_detects_missing_fail_return(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        bool load_config() {
          string cfg = raw;
          dyn_string parts = strsplit(raw, ",");
          if (dynlen(parts) < 7) {
            continue;
          }
          return true;
        }
        """
        violations = checker.check_config_error_contract(code)
        self.assertTrue(any(v["rule_id"] == "CFG-ERR-01" for v in violations))

    def test_perf_dpget_batch_detected_in_loop(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          for (int i=1; i<=10; i++) {
            dpGet("A.B.C1", v1);
            dpGet("A.B.C2", v2);
          }
        }
        """
        violations = checker.check_dpget_batch_optimization(code)
        self.assertTrue(any(v["rule_id"] == "PERF-DPGET-BATCH-01" for v in violations))

    def test_perf_dpget_batch_detects_multiple_loops_in_one_function(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          for (int i=1; i<=10; i++) {
            dpGet("A.B.C1", v1);
            dpGet("A.B.C2", v2);
          }

          for (int j=1; j<=10; j++) {
            dpGet("X.Y.C1", w1);
            dpGet("X.Y.C2", w2);
          }
        }
        """
        groups = checker.analyze_raw_code("sample.ctl", code, file_type="Server")
        hits = [v for g in groups for v in g.get("violations", []) if v.get("rule_id") == "PERF-DPGET-BATCH-01"]
        self.assertGreaterEqual(len(hits), 2)

    def test_perf_dpget_batch_outside_loop_reports_one_per_cluster(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          dpGet("A.B.C1", v1);
          dpGet("A.B.C2", v2);
          dpGet("A.B.C3", v3);
          dpGet("A.B.C4", v4);

          int sep = 0;
          int sep2 = 0;
          int sep3 = 0;
          int sep4 = 0;
          int sep5 = 0;
          int sep6 = 0;

          dpGet("X.Y.C1", w1);
          dpGet("X.Y.C2", w2);
          dpGet("X.Y.C3", w3);
          dpGet("X.Y.C4", w4);
        }
        """
        groups = checker.analyze_raw_code("sample.ctl", code, file_type="Server")
        hits = [v for g in groups for v in g.get("violations", []) if v.get("rule_id") == "PERF-DPGET-BATCH-01"]
        self.assertEqual(len(hits), 2, "연속 dpGet 클러스터는 함수 내에서 클러스터당 1건만 검출되어야 함")

    def test_perf_dpget_batch_not_detected_with_cache(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          for (int i=1; i<=10; i++) {
            if (mappingHasKey(cacheMap, key)) {
              continue;
            }
            dpGet("A.B.C1", v1);
            dpGet("A.B.C2", v2);
          }
        }
        """
        violations = checker.check_dpget_batch_optimization(code)
        self.assertFalse(any(v["rule_id"] == "PERF-DPGET-BATCH-01" for v in violations))

    def test_perf_dpset_batch_detected_in_loop(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          for (int i=1; i<=10; i++) {
            dpSet("A.B.C1", v1);
            dpSet("A.B.C2", v2);
          }
        }
        """
        violations = checker.check_dpset_batch_optimization(code)
        self.assertTrue(any(v["rule_id"] == "PERF-DPSET-BATCH-01" for v in violations))

    def test_perf_dpset_batch_not_detected_with_wait(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          for (int i=1; i<=10; i++) {
            dpSet("A.B.C1", v1);
            dpSet("A.B.C2", v2);
            dpSetWait("A.B.C1", v1);
          }
        }
        """
        violations = checker.check_dpset_batch_optimization(code)
        self.assertFalse(any(v["rule_id"] == "PERF-DPSET-BATCH-01" for v in violations))

    def test_exc_try_rule_detects_missing_try_catch(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          dpQuery("SELECT * FROM A.B.C");
          dpSet("A.B.C", 1);
        }
        """
        violations = checker.check_try_catch_for_risky_ops(code)
        self.assertTrue(any(v["rule_id"] == "EXC-TRY-01" for v in violations))

    def test_exc_try_rule_not_detected_with_try_catch(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          try {
            dpQuery("SELECT * FROM A.B.C");
          } catch (...) {
            DebugN("err");
          }
        }
        """
        violations = checker.check_try_catch_for_risky_ops(code)
        self.assertFalse(any(v["rule_id"] == "EXC-TRY-01" for v in violations))

    def test_exc_try_rule_not_detected_for_single_low_risk_call(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          dpGet("A.B.C", v);
        }
        """
        violations = checker.check_try_catch_for_risky_ops(code)
        self.assertFalse(any(v["rule_id"] == "EXC-TRY-01" for v in violations))

    def test_safe_div_detects_missing_guard(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void load_config() {
          string config_value = "x";
          ratio = total / count;
        }
        """
        violations = checker.check_division_zero_guard(code)
        self.assertTrue(any(v["rule_id"] == "SAFE-DIV-01" for v in violations))

    def test_safe_div_not_detected_with_guard(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void load_config() {
          string cfg = "x";
          if (count != 0) {
            ratio = total / count;
          }
        }
        """
        violations = checker.check_division_zero_guard(code)
        self.assertFalse(any(v["rule_id"] == "SAFE-DIV-01" for v in violations))

    def test_safe_div_not_detected_with_if_return_guard(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void load_config() {
          string cfg = "x";
          if (count == 0) return;
          ratio = total / count;
        }
        """
        violations = checker.check_division_zero_guard(code)
        self.assertFalse(any(v["rule_id"] == "SAFE-DIV-01" for v in violations))

    def test_dpset_chain_skips_loop_context(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          while (run) {
            dpSet("A.B.C1", 1);
            dpSet("A.B.C2", 2);
          }
        }
        """
        violations = checker.check_consecutive_dpset(code)
        self.assertFalse(any(v["rule_id"] == "PERF-DPSET-CHAIN" for v in violations))

    def test_perf_setvalue_batch_detected(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          for (int i=1; i<=10; i++) {
            setValue("A.B.C1", v1);
            setValue("A.B.C2", v2);
          }
        }
        """
        violations = checker.check_setvalue_batch_optimization(code)
        self.assertTrue(any(v["rule_id"] == "PERF-SETVALUE-BATCH-01" for v in violations))

    def test_perf_setvalue_batch_not_detected_when_setmultivalue_exists(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          for (int i=1; i<=10; i++) {
            setValue("A.B.C1", v1);
            setValue("A.B.C2", v2);
            setMultiValue("A.B.C1", v1, "A.B.C2", v2);
          }
        }
        """
        violations = checker.check_setvalue_batch_optimization(code)
        self.assertFalse(any(v["rule_id"] == "PERF-SETVALUE-BATCH-01" for v in violations))

    def test_perf_setvalue_batch_detects_multiple_loops_in_one_function(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          for (int i=1; i<=10; i++) {
            setValue("A.B.C1", v1);
            setValue("A.B.C2", v2);
          }

          while (run) {
            setValue("X.Y.C1", w1);
            setValue("X.Y.C2", w2);
          }
        }
        """
        groups = checker.analyze_raw_code("sample.ctl", code, file_type="Server")
        hits = [v for g in groups for v in g.get("violations", []) if v.get("rule_id") == "PERF-SETVALUE-BATCH-01"]
        self.assertGreaterEqual(len(hits), 2)

    def test_perf_setmultivalue_adopt_detected(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          setValue("A.B.C1", v1);
          setValue("A.B.C2", v2);
          setValue("A.B.C3", v3);
        }
        """
        violations = checker.check_setmultivalue_adoption(code)
        self.assertTrue(any(v["rule_id"] == "PERF-SETMULTIVALUE-ADOPT-01" for v in violations))

    def test_perf_setmultivalue_adopt_reports_one_per_cluster(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          setValue("A.B.C1", v1);
          setValue("A.B.C2", v2);
          setValue("A.B.C3", v3);
          setValue("A.B.C4", v4);

          int sep = 0;
          int sep2 = 0;
          int sep3 = 0;
          int sep4 = 0;
          int sep5 = 0;
          int sep6 = 0;

          setValue("X.Y.C1", w1);
          setValue("X.Y.C2", w2);
          setValue("X.Y.C3", w3);
          setValue("X.Y.C4", w4);
        }
        """
        groups = checker.analyze_raw_code("sample.ctl", code, file_type="Server")
        hits = [v for g in groups for v in g.get("violations", []) if v.get("rule_id") == "PERF-SETMULTIVALUE-ADOPT-01"]
        self.assertEqual(len(hits), 2)

    def test_perf_getvalue_batch_detected(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          while (run) {
            getValue("A.B.C1", v1);
            getValue("A.B.C2", v2);
          }
        }
        """
        violations = checker.check_getvalue_batch_optimization(code)
        self.assertTrue(any(v["rule_id"] == "PERF-GETVALUE-BATCH-01" for v in violations))

    def test_perf_getvalue_batch_detects_multiple_loops_in_one_function(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          while (run) {
            getValue("A.B.C1", v1);
            getValue("A.B.C2", v2);
          }

          for (int i=0; i<10; i++) {
            getValue("X.Y.C1", w1);
            getValue("X.Y.C2", w2);
          }
        }
        """
        groups = checker.analyze_raw_code("sample.ctl", code, file_type="Server")
        hits = [v for g in groups for v in g.get("violations", []) if v.get("rule_id") == "PERF-GETVALUE-BATCH-01"]
        self.assertGreaterEqual(len(hits), 2)

    def test_perf_getvalue_batch_not_detected_with_cache_guard(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          while (run) {
            if (mappingHasKey(cacheMap, key)) continue;
            getValue("A.B.C1", v1);
            getValue("A.B.C2", v2);
          }
        }
        """
        violations = checker.check_getvalue_batch_optimization(code)
        self.assertFalse(any(v["rule_id"] == "PERF-GETVALUE-BATCH-01" for v in violations))

    def test_perf_getmultivalue_adopt_reports_one_per_cluster(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          getValue("A.B.C1", v1);
          getValue("A.B.C2", v2);
          getValue("A.B.C3", v3);
          getValue("A.B.C4", v4);

          int sep = 0;
          int sep2 = 0;
          int sep3 = 0;
          int sep4 = 0;
          int sep5 = 0;
          int sep6 = 0;

          getValue("X.Y.C1", w1);
          getValue("X.Y.C2", w2);
          getValue("X.Y.C3", w3);
          getValue("X.Y.C4", w4);
        }
        """
        groups = checker.analyze_raw_code("sample.ctl", code, file_type="Server")
        hits = [v for g in groups for v in g.get("violations", []) if v.get("rule_id") == "PERF-GETMULTIVALUE-ADOPT-01"]
        self.assertEqual(len(hits), 2)

    def test_new_perf_rules_ignore_comment_only_tokens(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          // setValue("A.B.C1", v1);
          // setValue("A.B.C2", v2);
          int a = 1; // getValue("A.B.C1", v1);
          DebugN(a);
        }
        """
        event = {"event": "Global", "code": code, "line_start": 1}
        violations = checker.check_event(event, file_type="Server")
        blocked = {"PERF-SETVALUE-BATCH-01", "PERF-SETMULTIVALUE-ADOPT-01", "PERF-GETVALUE-BATCH-01"}
        self.assertFalse(any(v["rule_id"] in blocked for v in violations))

    def test_perf_agg_detected_manual_sum_avg_pattern(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          for (int i=1; i<=10; i++) {
            sum += value;
            count++;
          }
          avg = sum / count;
        }
        """
        violations = checker.check_manual_aggregation_pattern(code)
        self.assertTrue(any(v["rule_id"] == "PERF-AGG-01" for v in violations))

    def test_style_idx_detects_magic_indices(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          parsed[1] = parts[2];
          parsed[2] = parts[3];
          parsed[3] = parts[4];
        }
        """
        violations = checker.check_magic_index_usage(code)
        self.assertTrue(any(v["rule_id"] == "STYLE-IDX-01" for v in violations))

    def test_style_idx_not_detected_with_idx_constants(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        const int IDX_LEVEL_DP = 2;
        const int IDX_USAGE_DP = 3;
        void f() {
          parsed[1] = parts[IDX_LEVEL_DP];
          parsed[2] = parts[IDX_USAGE_DP];
          parsed[3] = parts[IDX_USAGE_DP];
        }
        """
        violations = checker.check_magic_index_usage(code)
        self.assertFalse(any(v["rule_id"] == "STYLE-IDX-01" for v in violations))

    def test_style_header_not_flagged_when_comment_block_exists(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        //******************************************************************************
        // name: main
        // argument:
        // return: void
        //******************************************************************************
        void main()
        {
          DebugN("ok");
        }
        """
        event = {"event": "Global", "code": code, "line_start": 1}
        violations = checker.check_event(event, file_type="Server")
        self.assertFalse(any(v["rule_id"] == "STYLE-HEADER-01" for v in violations))

    def test_hard03_detects_repeated_float_literal(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        void f() {
          fallback = 0.001;
          if (x < 0.001) return;
        }
        """
        violations = checker.check_float_literal_hardcoding(code)
        self.assertTrue(any(v["rule_id"] == "HARD-03" for v in violations))

    def test_hard03_not_detected_with_const(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        const float MIN_USAGE = 0.001;
        void f() {
          fallback = MIN_USAGE;
          if (x < MIN_USAGE) return;
        }
        """
        violations = checker.check_float_literal_hardcoding(code)
        self.assertFalse(any(v["rule_id"] == "HARD-03" for v in violations))

    def test_comment_line_does_not_trigger_dp_rules(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          // dpSet("A.B.C", 1);
          int a = 1; // test
          DebugN(a);
        }
        """
        event = {"event": "Global", "code": code, "line_start": 1}
        violations = checker.check_event(event, file_type="Server")
        self.assertFalse(any(v["rule_id"] == "EXC-DP-01" for v in violations))
        self.assertFalse(any(v["rule_id"] == "PERF-DPSET-CHAIN" for v in violations))

    def test_inline_comment_tail_is_excluded_from_analysis(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
          int a = 1; // dpSet("A.B.C", 1);
          DebugN(a);
        }
        """
        event = {"event": "Global", "code": code, "line_start": 1}
        violations = checker.check_event(event, file_type="Server")
        self.assertFalse(any(v["rule_id"] == "EXC-DP-01" for v in violations))

    def test_issue_id_deterministic(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        event = {"event": "Global", "code": 'sprintf(sql, "SELECT * FROM x WHERE y=%s", input);', "line_start": 1}
        first = checker.check_event(event)
        second = checker.check_event(event)
        self.assertEqual([v["issue_id"] for v in first], [v["issue_id"] for v in second])

    def test_target_collection_excludes_stale_generated_txt(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ctl_path = os.path.join(temp_dir, "sample.ctl")
            stale_txt_path = os.path.join(temp_dir, "stale_pnl.txt")
            with open(ctl_path, "w", encoding="utf-8") as f:
                f.write("main() { DebugN(\"ok\"); }")
            with open(stale_txt_path, "w", encoding="utf-8") as f:
                f.write("stale")

            app = CodeInspectorApp()
            app.data_dir = temp_dir
            targets = app.collect_targets()
            basenames = {os.path.basename(path) for path in targets}

            self.assertIn("sample.ctl", basenames)
            self.assertNotIn("stale_pnl.txt", basenames)

    def test_default_mode_is_ai_assist(self):
        self.assertEqual(DEFAULT_MODE, "AI 보조")

    def test_live_ai_toggle_off_uses_mock_review(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            target = os.path.join(temp_dir, "sample.ctl")
            with open(target, "w", encoding="utf-8") as f:
                f.write('main() { dpSet("A.B.C", 1); }')

            app = CodeInspectorApp()
            app.data_dir = temp_dir

            app.checker.analyze_raw_code = lambda *_args, **_kwargs: [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-R1-1",
                            "rule_id": "R1",
                            "rule_item": "test-item",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 1,
                            "message": "test violation",
                        }
                    ],
                }
            ]

            calls = {"live": 0, "mock": 0}

            def fake_live_review(*_args, **_kwargs):
                calls["live"] += 1
                return "LIVE REVIEW"

            def fake_mock_review(*_args, **_kwargs):
                calls["mock"] += 1
                return "MOCK REVIEW"

            app.ai_tool.generate_review = fake_live_review
            app.ai_tool.get_mock_review = fake_mock_review

            report = app.analyze_file(target, mode="AI 보조", enable_live_ai=False)
            self.assertEqual(calls["live"], 0)
            self.assertEqual(calls["mock"], 1)
            self.assertEqual(report["ai_reviews"][0]["source"], "mock")
            self.assertEqual(report["ai_reviews"][0]["review"], "MOCK REVIEW")

    def test_live_ai_toggle_on_uses_live_review(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            target = os.path.join(temp_dir, "sample.ctl")
            with open(target, "w", encoding="utf-8") as f:
                f.write('main() { dpSet("A.B.C", 1); }')

            app = CodeInspectorApp()
            app.data_dir = temp_dir

            app.checker.analyze_raw_code = lambda *_args, **_kwargs: [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-R1-1",
                            "rule_id": "R1",
                            "rule_item": "test-item",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 1,
                            "message": "test violation",
                        }
                    ],
                }
            ]

            calls = {"live": 0, "mock": 0}

            def fake_live_review(*_args, **_kwargs):
                calls["live"] += 1
                return "LIVE REVIEW"

            def fake_mock_review(*_args, **_kwargs):
                calls["mock"] += 1
                return "MOCK REVIEW"

            app.ai_tool.generate_review = fake_live_review
            app.ai_tool.get_mock_review = fake_mock_review

            report = app.analyze_file(target, mode="AI 보조", enable_live_ai=True)
            self.assertEqual(calls["live"], 1)
            self.assertEqual(calls["mock"], 0)
            self.assertEqual(report["ai_reviews"][0]["source"], "live")
            self.assertEqual(report["ai_reviews"][0]["review"], "LIVE REVIEW")

    def test_live_ai_fail_soft_message(self):
        reviewer = LLMReviewer(
            ai_config={
                "provider": "unsupported-provider",
                "fail_soft": True,
                "enabled_default": False,
                "timeout_sec": 5,
            }
        )
        text = reviewer.generate_review("main() {}", [])
        self.assertIn("AI live review failed:", text)

    def test_mcp_context_fetch_success(self):
        client = MCPContextClient(
            mcp_config={
                "url": "http://localhost:3000",
                "timeout_sec": 2,
                "max_drivers_in_prompt": 5,
            }
        )

        def fake_get_json(endpoint):
            if endpoint == "/context":
                return {"projectName": "TEST_PROJ", "site": "CHEONAN", "environment": "prod"}
            if endpoint == "/drivers":
                return {"drivers": [{"name": "CTRL"}, {"name": "DB"}, {"name": "ALARM"}]}
            return {}

        client._get_json = fake_get_json
        payload = client.fetch_context()
        self.assertTrue(payload.get("enabled"))
        self.assertEqual(payload.get("project", {}).get("projectName"), "TEST_PROJ")
        self.assertEqual(len(payload.get("drivers", [])), 3)

    def test_mcp_context_fetch_fail_soft(self):
        client = MCPContextClient(mcp_config={"url": "http://localhost:3000", "timeout_sec": 1})

        def failing_get_json(_endpoint):
            raise RuntimeError("mcp offline")

        client._get_json = failing_get_json
        payload = client.fetch_context()
        self.assertFalse(payload.get("enabled"))
        self.assertIn("mcp offline", str(payload.get("error", "")))

    def test_llm_prompt_contains_mcp_context_when_enabled(self):
        reviewer = LLMReviewer(ai_config={"provider": "ollama", "fail_soft": True})
        prompt = reviewer._build_prompt(
            code="main() {}",
            violations=[],
            use_context=True,
            context_payload={
                "enabled": True,
                "project": {"projectName": "PROJ_A", "site": "CHEONAN", "environment": "prod"},
                "drivers": [{"name": "CTRL"}, {"name": "DB"}],
            },
        )
        self.assertIn("project=PROJ_A", prompt)
        self.assertIn("drivers=CTRL, DB", prompt)

    def test_llm_prompt_context_fallback_when_missing(self):
        reviewer = LLMReviewer(ai_config={"provider": "ollama", "fail_soft": True})
        prompt = reviewer._build_prompt(
            code="main() {}",
            violations=[],
            use_context=True,
            context_payload={"enabled": False, "error": "offline"},
        )
        self.assertIn("Context: N/A", prompt)

    def test_collect_targets_raw_txt_off_by_default(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ctl_path = os.path.join(temp_dir, "sample.ctl")
            raw_txt_path = os.path.join(temp_dir, "raw_input.txt")
            with open(ctl_path, "w", encoding="utf-8") as f:
                f.write("main() { DebugN(\"ok\"); }")
            with open(raw_txt_path, "w", encoding="utf-8") as f:
                f.write("raw")

            app = CodeInspectorApp()
            app.data_dir = temp_dir
            targets = app.collect_targets()
            basenames = {os.path.basename(path) for path in targets}

            self.assertIn("sample.ctl", basenames)
            self.assertNotIn("raw_input.txt", basenames)

    def test_collect_targets_raw_txt_on(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            raw_txt_path = os.path.join(temp_dir, "raw_input.txt")
            with open(raw_txt_path, "w", encoding="utf-8") as f:
                f.write("raw")

            app = CodeInspectorApp()
            app.data_dir = temp_dir
            targets = app.collect_targets(allow_raw_txt=True)
            basenames = {os.path.basename(path) for path in targets}

            self.assertIn("raw_input.txt", basenames)

    def test_collect_targets_selected_raw_txt_requires_toggle(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            raw_txt_path = os.path.join(temp_dir, "raw_input.txt")
            with open(raw_txt_path, "w", encoding="utf-8") as f:
                f.write("raw")

            app = CodeInspectorApp()
            app.data_dir = temp_dir
            targets = app.collect_targets(selected_files=["raw_input.txt"], allow_raw_txt=False)
            basenames = {os.path.basename(path) for path in targets}

            self.assertNotIn("raw_input.txt", basenames)

    def test_ctl_uses_server_rules_only(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            while(1) {
                int x;
                dpGet("A.B.C", x);
            }
        }
        """
        findings = checker.analyze_raw_code("sample.ctl", code, file_type="Server")
        rule_items = [
            checker._normalize_rule_item(v["rule_item"])
            for group in findings
            for v in group.get("violations", [])
            if v.get("rule_item")
        ]
        self.assertIn("loop문내에처리조건", rule_items)

    def test_txt_uses_client_rules_only(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        code = """
        main() {
            while(1) {
                int x;
                dpGet("A.B.C", x);
            }
        }
        """
        findings = checker.analyze_raw_code("sample.txt", code, file_type="Client")
        rule_items = [
            checker._normalize_rule_item(v["rule_item"])
            for group in findings
            for v in group.get("violations", [])
            if v.get("rule_item")
        ]
        self.assertNotIn("loop문내에처리조건", rule_items)

    def test_pnl_xml_converted_txt_flows_to_client(self):
        app = CodeInspectorApp()
        self.assertEqual(app.infer_file_type("A_panel_pnl.txt"), "Client")
        self.assertEqual(app.infer_file_type("A_panel_xml.txt"), "Client")

    def test_rule_item_alias_normalization(self):
        checker = HeuristicChecker(os.path.join(self.config_dir, "parsed_rules.json"))
        self.assertEqual(
            checker._normalize_rule_item("Loop문 내의 처리 조건"),
            checker._normalize_rule_item("Loop문 내에 처리 조건"),
        )
        self.assertEqual(
            checker._normalize_rule_item("명명 규칙 및 코딩 스타일 준수 여부 확인"),
            checker._normalize_rule_item("명명 규칙 및 코딩 스타일 준수 확인"),
        )

    def test_ctrlpp_disabled_by_default(self):
        wrapper = CtrlppWrapper()
        self.assertFalse(wrapper.default_enabled)
        with tempfile.TemporaryDirectory() as temp_dir:
            target = os.path.join(temp_dir, "sample.ctl")
            with open(target, "w", encoding="utf-8") as f:
                f.write("main() { int a = 0; }")
            result = wrapper.run_check(target)
            self.assertEqual(result, [])

    def test_ctrlpp_enabled_ctl_only(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ctl_path = os.path.join(temp_dir, "sample.ctl")
            pnl_path = os.path.join(temp_dir, "sample.pnl")
            with open(ctl_path, "w", encoding="utf-8") as f:
                f.write("main() { int a = 0; }")
            with open(pnl_path, "w", encoding="utf-8") as f:
                f.write(
                    '6 13\n"btn"\n""\n1 10 10 E E E 1 E 1 E N "_Transparent" E N "_Transparent" E E\n'
                    '"Clicked" 1\n"main()\\n{\\n  int b = 1;\\n}" 0\n'
                )

            app = CodeInspectorApp()
            app.data_dir = temp_dir

            calls = []

            def fake_run_check(file_path, code_content=None, enabled=None, binary_path=None):
                calls.append(os.path.basename(file_path))
                return []

            app.ctrl_tool.run_check = fake_run_check
            app.run_directory_analysis(mode="Static", enable_ctrlppcheck=True)

            self.assertIn("sample.ctl", calls)
            self.assertFalse(any(name.endswith("_pnl.txt") for name in calls))

    def test_ctrlpp_missing_binary_fail_soft(self):
        wrapper = CtrlppWrapper()
        wrapper.config_binary_path = ""
        wrapper.auto_install_on_missing = False
        with tempfile.TemporaryDirectory() as temp_dir:
            target = os.path.join(temp_dir, "sample.ctl")
            with open(target, "w", encoding="utf-8") as f:
                f.write("main() { int a = 0; }")
            result = wrapper.run_check(
                target,
                enabled=True,
                binary_path=os.path.join(temp_dir, "does_not_exist", "ctrlppcheck.exe"),
            )
            self.assertGreaterEqual(len(result), 1)
            self.assertEqual(result[0].get("source"), "CtrlppCheck")
            self.assertIn("not found", result[0].get("message", ""))

    def test_ctrlpp_auto_install_attempt_on_missing_binary(self):
        wrapper = CtrlppWrapper()
        wrapper.config_binary_path = ""
        wrapper.auto_install_on_missing = True
        calls = {"count": 0}

        def fake_ensure():
            calls["count"] += 1
            return "", "CtrlppCheck install failed: mocked"

        wrapper.ensure_installed = fake_ensure
        with tempfile.TemporaryDirectory() as temp_dir:
            target = os.path.join(temp_dir, "sample.ctl")
            with open(target, "w", encoding="utf-8") as f:
                f.write("main() { int a = 0; }")
            result = wrapper.run_check(
                target,
                enabled=True,
                binary_path=os.path.join(temp_dir, "does_not_exist", "ctrlppcheck.exe"),
            )
        self.assertEqual(calls["count"], 1)
        self.assertGreaterEqual(len(result), 1)
        self.assertIn("install failed", result[0].get("message", "").lower())

    def test_ctrlpp_auto_install_skipped_when_disabled(self):
        wrapper = CtrlppWrapper()
        wrapper.config_binary_path = ""
        wrapper.auto_install_on_missing = True
        calls = {"count": 0}

        def fake_ensure():
            calls["count"] += 1
            return "", "CtrlppCheck install failed: mocked"

        wrapper.ensure_installed = fake_ensure
        with tempfile.TemporaryDirectory() as temp_dir:
            target = os.path.join(temp_dir, "sample.ctl")
            with open(target, "w", encoding="utf-8") as f:
                f.write("main() { int a = 0; }")
            result = wrapper.run_check(target, enabled=False)
        self.assertEqual(result, [])
        self.assertEqual(calls["count"], 0)

    def test_ctrlpp_auto_install_persists_binary_path_to_state_file_without_mutating_config(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            install_dir = os.path.join(temp_dir, "tools")
            config_path = os.path.join(temp_dir, "config.json")
            config_payload = {
                "ctrlppcheck": {
                    "enabled_default": False,
                    "binary_path": "",
                    "timeout_sec": 30,
                    "enable_levels": "style,information",
                    "version": "v1.0.2",
                    "auto_install_on_missing": True,
                    "install_dir": install_dir,
                    "github_repo": "siemens/CtrlppCheck",
                    "asset_pattern": "WinCCOA_QualityChecks_*.zip",
                }
            }
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(config_payload, f, ensure_ascii=False, indent=2)

            zip_fixture = os.path.join(temp_dir, "asset.zip")
            with zipfile.ZipFile(zip_fixture, "w") as zf:
                zf.writestr("WinCCOA_QualityChecks/bin/ctrlppcheck.exe", "fake-binary")

            wrapper = CtrlppWrapper(config_path=config_path)
            wrapper._fetch_release_payload = lambda: {
                "assets": [
                    {
                        "name": "WinCCOA_QualityChecks_v1.0.2.zip",
                        "browser_download_url": "mock://asset",
                        "digest": "",
                    }
                ]
            }
            wrapper._download_asset = lambda _url, dst: shutil.copy2(zip_fixture, dst)

            binary, error = wrapper.ensure_installed()
            self.assertEqual(error, "")
            self.assertTrue(binary.lower().endswith("ctrlppcheck.exe"))
            self.assertTrue(os.path.exists(binary))

            with open(config_path, "r", encoding="utf-8") as f:
                saved = json.load(f)
            self.assertEqual(saved["ctrlppcheck"]["binary_path"], "")

            self.assertTrue(os.path.exists(wrapper.install_state_path))
            with open(wrapper.install_state_path, "r", encoding="utf-8") as f:
                state_saved = json.load(f)
            self.assertEqual(os.path.normpath(state_saved["binary_path"]), os.path.normpath(binary))
            self.assertEqual(os.path.normpath(wrapper.state_binary_path), os.path.normpath(binary))

    def test_ctrlpp_auto_install_fail_soft_on_download_error(self):
        wrapper = CtrlppWrapper()
        wrapper.config_binary_path = ""
        wrapper.auto_install_on_missing = True

        def failing_ensure():
            raise RuntimeError("CtrlppCheck download failed: mocked offline")

        wrapper.ensure_installed = failing_ensure
        with tempfile.TemporaryDirectory() as temp_dir:
            target = os.path.join(temp_dir, "sample.ctl")
            with open(target, "w", encoding="utf-8") as f:
                f.write("main() { int a = 0; }")
            result = wrapper.run_check(target, enabled=True)

        self.assertGreaterEqual(len(result), 1)
        self.assertEqual(result[0].get("source"), "CtrlppCheck")
        self.assertIn("download failed", result[0].get("message", "").lower())

    def test_ctrlpp_prepare_for_analysis_skip_without_ctl(self):
        wrapper = CtrlppWrapper()
        wrapper.auto_install_on_missing = True
        calls = {"count": 0}

        def fake_ensure():
            calls["count"] += 1
            return "", "CtrlppCheck install failed: mocked"

        wrapper.ensure_installed = fake_ensure
        result = wrapper.prepare_for_analysis(True, ["sample_pnl.txt"])
        self.assertFalse(bool(result.get("attempted", False)))
        self.assertFalse(bool(result.get("ready", False)))
        self.assertEqual(calls["count"], 0)

    def test_ctrlpp_prepare_for_analysis_attempts_install_once(self):
        wrapper = CtrlppWrapper()
        wrapper.config_binary_path = ""
        wrapper.state_binary_path = ""
        wrapper.tool_path = ""
        wrapper.auto_install_on_missing = True
        calls = {"count": 0}

        def fake_ensure():
            calls["count"] += 1
            return "", "CtrlppCheck install failed: mocked"

        wrapper.ensure_installed = fake_ensure
        result = wrapper.prepare_for_analysis(True, ["sample.ctl"])
        self.assertTrue(bool(result.get("attempted", False)))
        self.assertFalse(bool(result.get("ready", False)))
        self.assertIn("install failed", str(result.get("message", "")).lower())
        self.assertEqual(calls["count"], 1)

    def test_ctrlpp_prepare_for_analysis_blocks_immediate_reinstall_retry(self):
        wrapper = CtrlppWrapper()
        wrapper.config_binary_path = ""
        wrapper.state_binary_path = ""
        wrapper.tool_path = ""
        wrapper.auto_install_on_missing = True
        calls = {"count": 0}

        def fake_ensure():
            calls["count"] += 1
            return "", "CtrlppCheck install failed: mocked"

        wrapper.ensure_installed = fake_ensure
        preflight = wrapper.prepare_for_analysis(True, ["sample.ctl"])
        self.assertTrue(bool(preflight.get("attempted", False)))
        self.assertFalse(bool(preflight.get("ready", False)))

        with tempfile.TemporaryDirectory() as temp_dir:
            target = os.path.join(temp_dir, "sample.ctl")
            with open(target, "w", encoding="utf-8") as f:
                f.write("main() { int a = 0; }")
            result = wrapper.run_check(target, enabled=True)
        self.assertEqual(calls["count"], 1)
        self.assertGreaterEqual(len(result), 1)
        self.assertIn("install failed", result[0].get("message", "").lower())

    def test_ctrlpp_checksum_validation(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            install_dir = os.path.join(temp_dir, "tools")
            config_path = os.path.join(temp_dir, "config.json")
            config_payload = {
                "ctrlppcheck": {
                    "enabled_default": False,
                    "binary_path": "",
                    "timeout_sec": 30,
                    "enable_levels": "style,information",
                    "version": "v1.0.2",
                    "auto_install_on_missing": True,
                    "install_dir": install_dir,
                    "github_repo": "siemens/CtrlppCheck",
                    "asset_pattern": "WinCCOA_QualityChecks_*.zip",
                }
            }
            with open(config_path, "w", encoding="utf-8") as f:
                json.dump(config_payload, f, ensure_ascii=False, indent=2)

            zip_fixture = os.path.join(temp_dir, "asset.zip")
            with zipfile.ZipFile(zip_fixture, "w") as zf:
                zf.writestr("WinCCOA_QualityChecks/bin/ctrlppcheck.exe", "fake-binary")

            wrapper = CtrlppWrapper(config_path=config_path)
            wrapper._fetch_release_payload = lambda: {
                "assets": [
                    {
                        "name": "WinCCOA_QualityChecks_v1.0.2.zip",
                        "browser_download_url": "mock://asset",
                        "digest": "sha256:0000000000000000000000000000000000000000000000000000000000000000",
                    }
                ]
            }
            wrapper._download_asset = lambda _url, dst: shutil.copy2(zip_fixture, dst)

            binary, error = wrapper.ensure_installed()
            self.assertEqual(binary, "")
            self.assertIn("checksum mismatch", error.lower())

    def test_ctrlpp_xml_parse_to_p2_schema(self):
        wrapper = CtrlppWrapper()
        sample_xml = """<?xml version="1.0" encoding="UTF-8"?>
<results version="2">
  <errors>
    <error id="unusedVariable" severity="warning" msg="unused variable" verbose="unused variable verbose">
      <location file="sample.ctl" line="12"/>
    </error>
  </errors>
</results>"""
        parsed = wrapper._parse_xml_report(sample_xml, default_file="sample.ctl")
        self.assertEqual(len(parsed), 1)
        finding = parsed[0]
        for key in ("type", "rule_id", "line", "message", "file", "source"):
            self.assertIn(key, finding)
        self.assertEqual(finding["rule_id"], "unusedVariable")
        self.assertEqual(finding["line"], 12)
        self.assertEqual(finding["source"], "CtrlppCheck")
    def test_ctrlpp_build_command_includes_project_name_option(self):
        wrapper = CtrlppWrapper()
        wrapper.winccoa_project_name = "TEST_PROJECT"
        cmd = wrapper._build_command("ctrlppcheck.exe", "sample.ctl")
        self.assertIn("--winccoa-projectName=TEST_PROJECT", cmd)

    def test_ctrlpp_no_xml_output_returns_warning(self):
        wrapper = CtrlppWrapper()
        wrapper.auto_install_on_missing = False

        original_find_binary = wrapper._find_binary
        original_run = subprocess.run
        wrapper._find_binary = lambda binary_path=None: "ctrlppcheck.exe"

        def fake_run(*_args, **_kwargs):
            return subprocess.CompletedProcess(
                args=["ctrlppcheck.exe", "sample.ctl"],
                returncode=0,
                stdout="Mandatory option missing: --winccoa-projectName",
                stderr="",
            )

        subprocess.run = fake_run
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                target = os.path.join(temp_dir, "sample.ctl")
                with open(target, "w", encoding="utf-8") as f:
                    f.write("main() { int a = 0; }")
                result = wrapper.run_check(target, enabled=True)
        finally:
            wrapper._find_binary = original_find_binary
            subprocess.run = original_run

        self.assertGreaterEqual(len(result), 1)
        self.assertEqual(result[0].get("source"), "CtrlppCheck")
        self.assertIn("no XML output", result[0].get("message", ""))
    @staticmethod
    def _normalize_text(text):
        return text.replace("\r\n", "\n").replace("\r", "\n").strip()

    def _copy_sources_to_temp(self, extensions):
        temp_dir = tempfile.TemporaryDirectory()
        copied = []
        for name in sorted(os.listdir(self.data_dir)):
            lower = name.lower()
            if any(lower.endswith(ext) for ext in extensions):
                src = os.path.join(self.data_dir, name)
                dst = os.path.join(temp_dir.name, name)
                shutil.copy2(src, dst)
                copied.append(name)
        return temp_dir, copied

    @staticmethod
    def _extract_xml_script_texts(content):
        root = ET.fromstring(content)
        scripts = []
        for script in root.iter("script"):
            if script.text and script.text.strip():
                scripts.append(html.unescape(script.text).strip())
        return scripts

    def test_convert_sources_generates_txt_for_all_pnl_xml(self):
        temp_ctx, copied = self._copy_sources_to_temp((".pnl", ".xml"))
        with temp_ctx as temp_dir:
            pnl_files = [name for name in copied if name.lower().endswith(".pnl")]
            xml_files = [name for name in copied if name.lower().endswith(".xml")]
            if not pnl_files and not xml_files:
                self.skipTest("No .pnl/.xml source files available for conversion test")

            app = CodeInspectorApp()
            app.data_dir = temp_dir
            generated = app.convert_sources()

            generated_names = sorted(os.path.basename(path) for path in generated)
            expected_names = sorted(
                [name.replace(".pnl", "_pnl.txt") for name in pnl_files]
                + [name.replace(".xml", "_xml.txt") for name in xml_files]
            )
            self.assertEqual(generated_names, expected_names)

            for expected in expected_names:
                self.assertTrue(os.path.exists(os.path.join(temp_dir, expected)), f"Missing output: {expected}")

    def test_pnl_conversion_no_script_loss(self):
        temp_ctx, copied = self._copy_sources_to_temp((".pnl",))
        parser = PnlParser()
        with temp_ctx as temp_dir:
            pnl_files = [name for name in copied if name.lower().endswith(".pnl")]
            if not pnl_files:
                self.skipTest("No .pnl source files available for no-loss test")

            app = CodeInspectorApp()
            app.data_dir = temp_dir
            app.convert_sources()

            for pnl_name in pnl_files:
                source_path = os.path.join(temp_dir, pnl_name)
                output_path = os.path.join(temp_dir, pnl_name.replace(".pnl", "_pnl.txt"))

                with open(source_path, "r", encoding="utf-8", errors="ignore") as f:
                    source_content = f.read()
                with open(output_path, "r", encoding="utf-8", errors="ignore") as f:
                    converted_content = f.read()

                converted_norm = self._normalize_text(converted_content)
                parsed = parser.normalize_pnl(source_content)
                expected_codes = [
                    self._normalize_text(event["code"])
                    for obj in parsed
                    for event in obj.get("events", [])
                    if event.get("code") and self._normalize_text(event["code"])
                ]

                self.assertGreater(len(expected_codes), 0, f"No executable script extracted from {pnl_name}")
                for script in expected_codes:
                    self.assertIn(script, converted_norm, f"Missing script block from converted txt: {pnl_name}")

    def test_xml_conversion_no_script_loss(self):
        temp_ctx, copied = self._copy_sources_to_temp((".xml",))
        with temp_ctx as temp_dir:
            xml_files = [name for name in copied if name.lower().endswith(".xml")]
            if not xml_files:
                self.skipTest("No .xml source files available for no-loss test")

            app = CodeInspectorApp()
            app.data_dir = temp_dir
            app.convert_sources()

            for xml_name in xml_files:
                source_path = os.path.join(temp_dir, xml_name)
                output_path = os.path.join(temp_dir, xml_name.replace(".xml", "_xml.txt"))

                with open(source_path, "r", encoding="utf-8", errors="ignore") as f:
                    source_content = f.read()
                with open(output_path, "r", encoding="utf-8", errors="ignore") as f:
                    converted_content = f.read()

                converted_norm = self._normalize_text(converted_content)
                expected_scripts = [
                    self._normalize_text(script)
                    for script in self._extract_xml_script_texts(source_content)
                    if self._normalize_text(script)
                ]

                self.assertGreater(len(expected_scripts), 0, f"No script tag extracted from {xml_name}")
                for script in expected_scripts:
                    self.assertIn(script, converted_norm, f"Missing XML script from converted txt: {xml_name}")

    def test_convert_sources_output_exact_match_with_parser(self):
        temp_ctx, copied = self._copy_sources_to_temp((".pnl", ".xml"))
        pnl_parser = PnlParser()
        xml_parser = XmlParser()
        with temp_ctx as temp_dir:
            if not copied:
                self.skipTest("No .pnl/.xml source files available for parser exact-match test")

            app = CodeInspectorApp()
            app.data_dir = temp_dir
            app.convert_sources()

            for source_name in copied:
                source_path = os.path.join(temp_dir, source_name)
                with open(source_path, "r", encoding="utf-8", errors="ignore") as f:
                    source_content = f.read()

                if source_name.lower().endswith(".pnl"):
                    output_name = source_name.replace(".pnl", "_pnl.txt")
                    expected_text = pnl_parser.convert_to_text(source_content)
                else:
                    output_name = source_name.replace(".xml", "_xml.txt")
                    expected_text = xml_parser.parse(source_content)

                output_path = os.path.join(temp_dir, output_name)
                with open(output_path, "r", encoding="utf-8", errors="ignore") as f:
                    actual_text = f.read()

                self.assertEqual(self._normalize_text(actual_text), self._normalize_text(expected_text))

    def test_run_directory_analysis_uses_request_scoped_reporter(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ctl_path = os.path.join(temp_dir, "sample.ctl")
            with open(ctl_path, "w", encoding="utf-8") as f:
                f.write('main() { dpSet("A.B.C", 1); }')

            app = CodeInspectorApp()
            app.data_dir = temp_dir

            result_a = app.run_directory_analysis(
                mode="Static",
                selected_files=["sample.ctl"],
                enable_ctrlppcheck=False,
                enable_live_ai=False,
            )
            result_b = app.run_directory_analysis(
                mode="Static",
                selected_files=["sample.ctl"],
                enable_ctrlppcheck=False,
                enable_live_ai=False,
            )

            output_a = os.path.normpath(result_a.get("output_dir", ""))
            output_b = os.path.normpath(result_b.get("output_dir", ""))
            self.assertTrue(output_a)
            self.assertTrue(output_b)
            self.assertNotEqual(output_a, output_b)
            self.assertTrue(os.path.isdir(output_a))
            self.assertTrue(os.path.isdir(output_b))

    def test_analyze_file_with_custom_reporter_writes_to_given_output_dir(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ctl_path = os.path.join(temp_dir, "sample.ctl")
            with open(ctl_path, "w", encoding="utf-8") as f:
                f.write('main() { dpSet("A.B.C", 1); }')

            app = CodeInspectorApp()
            app.data_dir = temp_dir

            custom_reporter = Reporter(config_dir=app.reporter.config_dir)
            custom_reporter.output_base_dir = os.path.join(temp_dir, "custom_reports")
            custom_reporter.start_session()

            app.analyze_file(
                ctl_path,
                mode="Static",
                enable_ctrlppcheck=False,
                enable_live_ai=False,
                reporter=custom_reporter,
            )

            reviewed_path = os.path.join(custom_reporter.output_dir, "sample_REVIEWED.txt")
            excel_matches = glob.glob(
                os.path.join(custom_reporter.output_dir, f"CodeReview_Submission_sample_{custom_reporter.timestamp}.xlsx")
            )
            self.assertTrue(os.path.exists(reviewed_path))
            self.assertEqual(len(excel_matches), 1)

    def test_submission_excel_and_reviewed_txt_format_after_real_analysis(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ctl_path = os.path.join(temp_dir, "sample.ctl")
            with open(ctl_path, "w", encoding="utf-8") as f:
                f.write('main() { while(1) { int x; dpGet("A.B.C", x); } }')

            app = CodeInspectorApp()
            app.data_dir = temp_dir
            result = app.run_directory_analysis(
                mode="Static",
                selected_files=["sample.ctl"],
                enable_ctrlppcheck=False,
                enable_live_ai=False,
            )

            output_dir = os.path.normpath(result.get("output_dir", ""))
            self.assertTrue(os.path.isdir(output_dir))
            excel_files = [
                name for name in os.listdir(output_dir) if name.startswith("CodeReview_Submission_sample_") and name.endswith(".xlsx")
            ]
            self.assertGreaterEqual(len(excel_files), 1)

            from openpyxl import load_workbook

            excel_path = os.path.join(output_dir, excel_files[0])
            wb = load_workbook(excel_path)
            self.assertIn("상세결과", wb.sheetnames)
            self.assertIn("검증 결과", wb.sheetnames)

            reviewed_path = os.path.join(output_dir, "sample_REVIEWED.txt")
            self.assertTrue(os.path.exists(reviewed_path))
            with open(reviewed_path, "r", encoding="utf-8") as f:
                content = f.read()
            self.assertIn("// >>TODO", content)
            self.assertIn("// [REVIEW]", content)

if __name__ == "__main__":
    unittest.main()

