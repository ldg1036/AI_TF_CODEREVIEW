"""Conversion cases for system verification."""

try:
    from ._system_verification_base import *  # noqa: F403
except ImportError:
    from _system_verification_base import *  # noqa: F403


class SystemVerificationConversionMixin:
    def test_convert_sources_generates_txt_for_all_pnl_xml(self):
        temp_ctx, copied = self._copy_sources_to_temp((".pnl", ".xml"))
        with temp_ctx as temp_dir:
            pnl_files = [name for name in copied if name.lower().endswith(".pnl")]
            xml_files = [name for name in copied if name.lower().endswith(".xml")]
            if not pnl_files and not xml_files:
                self.skipTest("No .pnl/.xml source files available for conversion test")
    
            app = CodeInspectorApp()
            app.data_dir = temp_dir
            generated = app.convert_sources()
    
            generated_names = sorted(os.path.basename(path) for path in generated)
            expected_names = sorted(
                [name.replace(".pnl", "_pnl.txt") for name in pnl_files]
                + [name.replace(".xml", "_xml.txt") for name in xml_files]
            )
            self.assertEqual(generated_names, expected_names)
    
            for expected in expected_names:
                self.assertTrue(os.path.exists(os.path.join(temp_dir, expected)), f"Missing output: {expected}")

    def test_pnl_conversion_no_script_loss(self):
        temp_ctx, copied = self._copy_sources_to_temp((".pnl",))
        parser = PnlParser()
        with temp_ctx as temp_dir:
            pnl_files = [name for name in copied if name.lower().endswith(".pnl")]
            if not pnl_files:
                self.skipTest("No .pnl source files available for no-loss test")
    
            app = CodeInspectorApp()
            app.data_dir = temp_dir
            app.convert_sources()
    
            for pnl_name in pnl_files:
                source_path = os.path.join(temp_dir, pnl_name)
                output_path = os.path.join(temp_dir, pnl_name.replace(".pnl", "_pnl.txt"))
    
                with open(source_path, "r", encoding="utf-8", errors="ignore") as f:
                    source_content = f.read()
                with open(output_path, "r", encoding="utf-8", errors="ignore") as f:
                    converted_content = f.read()
    
                converted_norm = self._normalize_text(converted_content)
                parsed = parser.normalize_pnl(source_content)
                expected_codes = [
                    self._normalize_text(event["code"])
                    for obj in parsed
                    for event in obj.get("events", [])
                    if event.get("code") and self._normalize_text(event["code"])
                ]
    
                self.assertGreater(len(expected_codes), 0, f"No executable script extracted from {pnl_name}")
                for script in expected_codes:
                    self.assertIn(script, converted_norm, f"Missing script block from converted txt: {pnl_name}")

    def test_xml_conversion_no_script_loss(self):
        temp_ctx, copied = self._copy_sources_to_temp((".xml",))
        with temp_ctx as temp_dir:
            xml_files = [name for name in copied if name.lower().endswith(".xml")]
            if not xml_files:
                self.skipTest("No .xml source files available for no-loss test")
    
            app = CodeInspectorApp()
            app.data_dir = temp_dir
            app.convert_sources()
    
            for xml_name in xml_files:
                source_path = os.path.join(temp_dir, xml_name)
                output_path = os.path.join(temp_dir, xml_name.replace(".xml", "_xml.txt"))
    
                with open(source_path, "r", encoding="utf-8", errors="ignore") as f:
                    source_content = f.read()
                with open(output_path, "r", encoding="utf-8", errors="ignore") as f:
                    converted_content = f.read()
    
                converted_norm = self._normalize_text(converted_content)
                expected_scripts = [
                    self._normalize_text(script)
                    for script in self._extract_xml_script_texts(source_content)
                    if self._normalize_text(script)
                ]
    
                self.assertGreater(len(expected_scripts), 0, f"No script tag extracted from {xml_name}")
                for script in expected_scripts:
                    self.assertIn(script, converted_norm, f"Missing XML script from converted txt: {xml_name}")

    def test_convert_sources_output_exact_match_with_parser(self):
        temp_ctx, copied = self._copy_sources_to_temp((".pnl", ".xml"))
        pnl_parser = PnlParser()
        xml_parser = XmlParser()
        with temp_ctx as temp_dir:
            if not copied:
                self.skipTest("No .pnl/.xml source files available for parser exact-match test")
    
            app = CodeInspectorApp()
            app.data_dir = temp_dir
            app.convert_sources()
    
            for source_name in copied:
                source_path = os.path.join(temp_dir, source_name)
                with open(source_path, "r", encoding="utf-8", errors="ignore") as f:
                    source_content = f.read()
    
                if source_name.lower().endswith(".pnl"):
                    output_name = source_name.replace(".pnl", "_pnl.txt")
                    expected_text = pnl_parser.convert_to_text(source_content)
                else:
                    output_name = source_name.replace(".xml", "_xml.txt")
                    expected_text = xml_parser.parse(source_content)
    
                output_path = os.path.join(temp_dir, output_name)
                with open(output_path, "r", encoding="utf-8", errors="ignore") as f:
                    actual_text = f.read()
    
                self.assertEqual(self._normalize_text(actual_text), self._normalize_text(expected_text))

    def test_run_directory_analysis_uses_request_scoped_reporter(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ctl_path = os.path.join(temp_dir, "sample.ctl")
            with open(ctl_path, "w", encoding="utf-8") as f:
                f.write('main() { dpSet("A.B.C", 1); }')
    
            app = CodeInspectorApp()
            app.data_dir = temp_dir
    
            result_a = app.run_directory_analysis(
                mode="Static",
                selected_files=["sample.ctl"],
                enable_ctrlppcheck=False,
                enable_live_ai=False,
            )
            result_b = app.run_directory_analysis(
                mode="Static",
                selected_files=["sample.ctl"],
                enable_ctrlppcheck=False,
                enable_live_ai=False,
            )
    
            output_a = os.path.normpath(result_a.get("output_dir", ""))
            output_b = os.path.normpath(result_b.get("output_dir", ""))
            self.assertTrue(output_a)
            self.assertTrue(output_b)
            self.assertNotEqual(output_a, output_b)
            self.assertTrue(os.path.isdir(output_a))
            self.assertTrue(os.path.isdir(output_b))

    def test_analyze_file_with_custom_reporter_writes_to_given_output_dir(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ctl_path = os.path.join(temp_dir, "sample.ctl")
            with open(ctl_path, "w", encoding="utf-8") as f:
                f.write('main() { dpSet("A.B.C", 1); }')
    
            app = CodeInspectorApp()
            app.data_dir = temp_dir
    
            custom_reporter = Reporter(config_dir=app.reporter.config_dir)
            custom_reporter.output_base_dir = os.path.join(temp_dir, "custom_reports")
            custom_reporter.start_session()
    
            app.analyze_file(
                ctl_path,
                mode="Static",
                enable_ctrlppcheck=False,
                enable_live_ai=False,
                reporter=custom_reporter,
            )
    
            reviewed_path = os.path.join(custom_reporter.output_dir, "sample_REVIEWED.txt")
            excel_matches = glob.glob(
                os.path.join(custom_reporter.output_dir, f"CodeReview_Submission_sample_{custom_reporter.timestamp}.xlsx")
            )
            self.assertTrue(os.path.exists(reviewed_path))
            self.assertEqual(len(excel_matches), 1)

    def test_submission_excel_and_reviewed_txt_format_after_real_analysis(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            ctl_path = os.path.join(temp_dir, "sample.ctl")
            with open(ctl_path, "w", encoding="utf-8") as f:
                f.write('main() { while(1) { int x; dpGet("A.B.C", x); } }')
    
            app = CodeInspectorApp()
            app.data_dir = temp_dir
            result = app.run_directory_analysis(
                mode="Static",
                selected_files=["sample.ctl"],
                enable_ctrlppcheck=False,
                enable_live_ai=False,
            )
    
            output_dir = os.path.normpath(result.get("output_dir", ""))
            self.assertTrue(os.path.isdir(output_dir))
            excel_files = [
                name for name in os.listdir(output_dir) if name.startswith("CodeReview_Submission_sample_") and name.endswith(".xlsx")
            ]
            self.assertGreaterEqual(len(excel_files), 1)
    
            from openpyxl import load_workbook
    
            excel_path = os.path.join(output_dir, excel_files[0])
            wb = load_workbook(excel_path)
            self.assertIn("상세결과", wb.sheetnames)
            self.assertIn("검증 결과", wb.sheetnames)
    
            reviewed_path = os.path.join(output_dir, "sample_REVIEWED.txt")
            self.assertTrue(os.path.exists(reviewed_path))
            with open(reviewed_path, "r", encoding="utf-8") as f:
                content = f.read()
            self.assertIn("// >>TODO", content)
            self.assertNotIn("// [REVIEW]", content)
            self.assertNotIn("// [META]", content)
