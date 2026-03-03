import http.client
import json
import concurrent.futures
import os
import sys
import tempfile
import threading
import time
import urllib.parse
import unittest
from http.server import ThreadingHTTPServer

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
BACKEND_DIR = os.path.join(PROJECT_ROOT, "backend")
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

from main import CodeInspectorApp
from core.reporter import Reporter
from server import BASE_DIR, CodeInspectorHandler


class ApiIntegrationTests(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.TemporaryDirectory()
        self.data_dir = self.tmp_dir.name

        with open(os.path.join(self.data_dir, "sample.ctl"), "w", encoding="utf-8") as f:
            f.write('main() { dpSet("A.B.C", 1); }')
        with open(os.path.join(self.data_dir, "raw_input.txt"), "w", encoding="utf-8") as f:
            f.write('main() { DebugN("raw"); }')
        with open(os.path.join(self.data_dir, "server_loop.ctl"), "w", encoding="utf-8") as f:
            f.write('main() { while(1) { int x; dpGet("A.B.C", x); } }')
        with open(os.path.join(self.data_dir, "raw_loop.txt"), "w", encoding="utf-8") as f:
            f.write('main() { while(1) { int x; dpGet("A.B.C", x); } }')

        self.app = CodeInspectorApp()
        self.app.data_dir = self.data_dir
        self.app.ctrl_tool.auto_install_on_missing = False
        self.app.ctrl_tool.config_binary_path = ""
        self.app.ctrl_tool.state_binary_path = ""
        self.app.ctrl_tool.tool_path = os.path.join(self.data_dir, "missing", "ctrlppcheck.exe")

        frontend_dir = os.path.join(BASE_DIR, "frontend")

        def handler(*args, **kwargs):
            return CodeInspectorHandler(*args, app=self.app, frontend_dir=frontend_dir, **kwargs)

        self.httpd = ThreadingHTTPServer(("127.0.0.1", 0), handler)
        self.port = self.httpd.server_address[1]
        self.thread = threading.Thread(target=self.httpd.serve_forever, daemon=True)
        self.thread.start()

    def tearDown(self):
        self.httpd.shutdown()
        self.httpd.server_close()
        self.thread.join(timeout=3)
        self.tmp_dir.cleanup()

    def _request(self, method, path, body=None, headers=None):
        conn = http.client.HTTPConnection("127.0.0.1", self.port, timeout=15)
        request_headers = {"Content-Type": "application/json"}
        if isinstance(headers, dict):
            request_headers.update(headers)
        payload = json.dumps(body) if body is not None else None
        conn.request(method, path, payload, request_headers)
        resp = conn.getresponse()
        data = resp.read().decode("utf-8")
        conn.close()
        parsed = json.loads(data) if data else {}
        return resp.status, parsed

    @staticmethod
    def _extract_rule_items(payload):
        p1_groups = payload.get("violations", {}).get("P1", [])
        items = []
        for group in p1_groups:
            for violation in group.get("violations", []):
                item = violation.get("rule_item")
                if item:
                    items.append(item)
        return items

    def _force_single_internal_violation(self):
        self.app.checker.analyze_raw_code = lambda *_args, **_kwargs: [
            {
                "object": "sample.ctl",
                "event": "Global",
                "violations": [
                    {
                        "issue_id": "P1-R1-1",
                        "rule_id": "R1",
                        "rule_item": "test-item",
                        "priority_origin": "P1",
                        "severity": "Warning",
                        "line": 1,
                        "message": "test violation",
                    }
                ],
            }
        ]

    def test_get_api_files(self):
        status, payload = self._request("GET", "/api/files")
        self.assertEqual(status, 200)
        names = [item["name"] for item in payload["files"]]
        self.assertIn("sample.ctl", names)

    def test_post_api_analyze_selected_files(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"mode": "Static", "selected_files": ["sample.ctl"]},
        )
        self.assertEqual(status, 200)
        self.assertIn("summary", payload)
        self.assertIn("violations", payload)
        self.assertIn("report_paths", payload)
        self.assertIn("p1_total", payload["summary"])
        self.assertIn("p2_total", payload["summary"])
        self.assertIn("p3_total", payload["summary"])
        self.assertIn("metrics", payload)
        self.assertIn("timings_ms", payload["metrics"])
        self.assertIn("total", payload["metrics"]["timings_ms"])
        self.assertIn("convert", payload["metrics"]["timings_ms"])
        self.assertIn("llm_calls", payload["metrics"])
        self.assertIn("convert_cache", payload["metrics"])

    def test_post_api_analyze_start_returns_job(self):
        status, payload = self._request(
            "POST",
            "/api/analyze/start",
            {"mode": "Static", "selected_files": ["sample.ctl"]},
        )
        self.assertEqual(status, 202)
        self.assertTrue(str(payload.get("job_id", "")))
        self.assertEqual(str(payload.get("status", "")), "queued")
        self.assertIn("progress", payload)
        self.assertIn("poll_interval_ms", payload)

    def test_get_api_analyze_status_unknown_job_returns_404(self):
        status, payload = self._request("GET", "/api/analyze/status?job_id=missing-job-id")
        self.assertEqual(status, 404)
        self.assertIn("error", payload)

    def test_post_api_analyze_start_and_poll_until_completed(self):
        status, payload = self._request(
            "POST",
            "/api/analyze/start",
            {"mode": "Static", "selected_files": ["sample.ctl"]},
        )
        self.assertEqual(status, 202)
        job_id = str(payload.get("job_id", ""))
        self.assertTrue(job_id)

        final_payload = {}
        for _ in range(60):
            s, p = self._request("GET", f"/api/analyze/status?job_id={urllib.parse.quote(job_id)}")
            self.assertEqual(s, 200)
            final_payload = p
            if str(p.get("status", "")) in ("completed", "failed"):
                break
            time.sleep(0.05)

        self.assertEqual(str(final_payload.get("status", "")), "completed")
        self.assertIn("result", final_payload)
        result = final_payload.get("result", {})
        self.assertIn("summary", result)
        self.assertIn("violations", result)

    def test_post_api_analyze_start_invalid_selection_eventually_fails(self):
        status, payload = self._request(
            "POST",
            "/api/analyze/start",
            {"mode": "Static", "selected_files": ["missing.ctl"]},
        )
        self.assertEqual(status, 202)
        job_id = str(payload.get("job_id", ""))
        self.assertTrue(job_id)

        final_payload = {}
        for _ in range(60):
            s, p = self._request("GET", f"/api/analyze/status?job_id={urllib.parse.quote(job_id)}")
            self.assertEqual(s, 200)
            final_payload = p
            if str(p.get("status", "")) in ("completed", "failed"):
                break
            time.sleep(0.05)

        self.assertEqual(str(final_payload.get("status", "")), "failed")
        self.assertIn("error", final_payload)

    def test_post_api_analyze_p1_violations_include_file_field(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"mode": "Static", "selected_files": ["sample.ctl"]},
        )
        self.assertEqual(status, 200)
        p1_groups = payload.get("violations", {}).get("P1", [])
        self.assertGreaterEqual(len(p1_groups), 1)
        first_group = p1_groups[0]
        violations = first_group.get("violations", [])
        self.assertGreaterEqual(len(violations), 1)
        self.assertEqual(violations[0].get("file"), "sample.ctl")

    def test_post_api_analyze_invalid_file_returns_400(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"mode": "Static", "selected_files": ["missing.ctl"]},
        )
        self.assertEqual(status, 400)
        self.assertIn("error", payload)

    def test_post_api_analyze_default_mode_ai_assist(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"]},
        )
        self.assertEqual(status, 200)
        self.assertIn("summary", payload)
        self.assertIn("violations", payload)

    def test_post_api_analyze_live_ai_toggle_type_validation(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": "yes"},
        )
        self.assertEqual(status, 400)
        self.assertIn("error", payload)
        self.assertIn("enable_live_ai", payload["error"])

    def test_post_api_analyze_ai_with_context_type_validation(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "ai_with_context": "yes"},
        )
        self.assertEqual(status, 400)
        self.assertIn("error", payload)
        self.assertIn("ai_with_context", payload["error"])

    def test_post_api_analyze_selected_files_type_validation(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": 1},
        )
        self.assertEqual(status, 400)
        self.assertIn("error", payload)
        self.assertIn("selected_files", payload["error"])

    def test_post_api_analyze_allow_raw_txt_type_validation(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["raw_input.txt"], "allow_raw_txt": "false"},
        )
        self.assertEqual(status, 400)
        self.assertIn("error", payload)
        self.assertIn("allow_raw_txt", payload["error"])

    def test_post_api_analyze_live_ai_on_generates_p3(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: "LIVE REVIEW"

        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        p3 = payload.get("violations", {}).get("P3", [])
        self.assertGreaterEqual(len(p3), 1)
        self.assertEqual(p3[0].get("source"), "live")
        self.assertEqual(p3[0].get("review"), "LIVE REVIEW")

    def test_post_api_analyze_live_ai_passes_focus_snippet(self):
        self._force_single_internal_violation()
        seen = {"focus_snippet": ""}

        def fake_live_review(_code, _violations, **kwargs):
            seen["focus_snippet"] = str(kwargs.get("focus_snippet", "") or "")
            return "LIVE REVIEW"

        self.app.ai_tool.generate_review = fake_live_review
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        self.assertGreaterEqual(len(payload.get("violations", {}).get("P3", [])), 1)
        self.assertTrue(seen["focus_snippet"])
        self.assertIn("dpSet", seen["focus_snippet"])
        self.assertIn("// lines", seen["focus_snippet"])

    def test_post_api_analyze_live_ai_batch_groups_per_file_reduces_calls(self):
        self.app.checker.analyze_raw_code = lambda *_args, **_kwargs: [
            {
                "object": "sample.ctl",
                "event": "Global",
                "violations": [
                    {
                        "issue_id": "P1-A-1",
                        "rule_id": "R-A",
                        "rule_item": "a",
                        "priority_origin": "P1",
                        "severity": "Warning",
                        "line": 1,
                        "message": "v1",
                    }
                ],
            },
            {
                "object": "sample.ctl",
                "event": "EvtA",
                "violations": [
                    {
                        "issue_id": "P1-B-1",
                        "rule_id": "R-B",
                        "rule_item": "b",
                        "priority_origin": "P1",
                        "severity": "Warning",
                        "line": 1,
                        "message": "v2",
                    }
                ],
            },
        ]
        calls = {"count": 0}

        def fake_live_review(_code, violations, **kwargs):
            calls["count"] += 1
            return f"LIVE REVIEW BATCH size={len(violations)} snippet={bool(kwargs.get('focus_snippet'))}"

        original_batch = self.app.live_ai_batch_groups_per_file
        self.app.live_ai_batch_groups_per_file = 2
        self.app.ai_tool.generate_review = fake_live_review
        try:
            status, payload = self._request(
                "POST",
                "/api/analyze",
                {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
            )
        finally:
            self.app.live_ai_batch_groups_per_file = original_batch
        self.assertEqual(status, 200)
        self.assertEqual(calls["count"], 1)
        p3 = payload.get("violations", {}).get("P3", [])
        self.assertEqual(len(p3), 2)
        self.assertTrue(all(r.get("source") == "live-batch" for r in p3))
        self.assertEqual((payload.get("metrics") or {}).get("llm_calls"), 1)

    def test_post_api_analyze_defer_excel_reports_and_flush_endpoint(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {
                "selected_files": ["sample.ctl"],
                "mode": "Static",
                "enable_ctrlppcheck": False,
                "enable_live_ai": False,
                "defer_excel_reports": True,
            },
        )
        self.assertEqual(status, 200)
        self.assertIn("report_jobs", payload)
        excel_jobs = (payload.get("report_jobs") or {}).get("excel") or {}
        self.assertGreaterEqual(len(excel_jobs.get("jobs", [])), 1)

        flush_status, flush_payload = self._request(
            "POST",
            "/api/report/excel",
            {
                "session_id": payload.get("output_dir", ""),
                "wait": True,
                "timeout_sec": 30,
            },
        )
        self.assertEqual(flush_status, 200)
        self.assertIn("report_jobs", flush_payload)
        self.assertIn("excel", flush_payload.get("report_paths", {}))
        excel_summary = (flush_payload.get("report_jobs") or {}).get("excel") or {}
        self.assertEqual(excel_summary.get("pending_count"), 0)
        self.assertEqual(excel_summary.get("running_count"), 0)
        self.assertTrue(isinstance(flush_payload.get("all_completed"), bool))
        jobs = excel_summary.get("jobs") or []
        self.assertGreaterEqual(len(jobs), 1)
        self.assertIn("metrics", jobs[0])
        self.assertIn("timings_ms", jobs[0].get("metrics", {}))

    def test_autofix_prepare_and_apply_ctl_diff_flow(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )

        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        p3 = analyze_payload.get("violations", {}).get("P3", [])
        self.assertGreaterEqual(len(p3), 1)
        ai_review = p3[0]

        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)
        self.assertTrue(prepare_payload.get("proposal_id"))
        self.assertEqual(prepare_payload.get("file"), "sample.ctl")
        self.assertIn("--- sample.ctl", prepare_payload.get("unified_diff", ""))
        self.assertTrue(prepare_payload.get("base_hash"))

        diff_status, diff_payload = self._request(
            "GET",
            "/api/autofix/file-diff?" + urllib.parse.urlencode(
                {"file": "sample.ctl", "session_id": analyze_payload.get("output_dir", "")}
            ),
        )
        self.assertEqual(diff_status, 200)
        self.assertEqual(diff_payload.get("proposal_id"), prepare_payload.get("proposal_id"))

        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 200)
        self.assertTrue(apply_payload.get("applied"))
        self.assertEqual(apply_payload.get("file"), "sample.ctl")
        self.assertTrue(apply_payload.get("backup_path"))
        self.assertTrue(apply_payload.get("audit_log_path"))
        validation = apply_payload.get("validation", {})
        self.assertTrue(validation.get("hash_match"))
        self.assertTrue(validation.get("anchors_match"))
        quality = apply_payload.get("quality_metrics", {})
        self.assertEqual(quality.get("generator_type"), "llm")
        self.assertTrue(quality.get("applied"))

        with open(os.path.join(self.data_dir, "sample.ctl"), "r", encoding="utf-8") as f:
            patched = f.read()
        self.assertIn("[AI-AUTOFIX:", patched)
        self.assertIn("if (isValid) {", patched)

    def test_autofix_prepare_and_apply_raw_txt_llm_only(self):
        self.app.checker.analyze_raw_code = lambda *_args, **_kwargs: [
            {
                "object": "raw_input.txt",
                "event": "Global",
                "violations": [
                    {
                        "issue_id": "P1-RAW-1",
                        "rule_id": "R-RAW",
                        "rule_item": "raw-item",
                        "priority_origin": "P1",
                        "severity": "Warning",
                        "line": 1,
                        "message": "raw test violation",
                    }
                ],
            }
        ]
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 로그를 개선하세요.\n\n"
            "코드:\n```cpp\nDebugN(\"raw-fix\");\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["raw_input.txt"], "allow_raw_txt": True, "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]

        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "raw_input.txt",
                "object": ai_review.get("object", "raw_input.txt"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "rule",
                "prepare_mode": "compare",
            },
        )
        self.assertEqual(prepare_status, 200)
        self.assertEqual(prepare_payload.get("generator_type"), "llm")
        proposals = prepare_payload.get("proposals", [])
        self.assertEqual(len(proposals), 1)
        self.assertEqual((proposals[0] or {}).get("generator_type"), "llm")

        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "raw_input.txt",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
                "check_ctrlpp_regression": True,
            },
        )
        self.assertEqual(apply_status, 200)
        validation = apply_payload.get("validation", {})
        self.assertEqual(validation.get("syntax_check_skipped_reason"), "non_ctl_file")
        self.assertEqual(validation.get("ctrlpp_regression_skipped_reason"), "non_ctl_file")

        with open(os.path.join(self.data_dir, "raw_input.txt"), "r", encoding="utf-8") as f:
            patched = f.read()
        self.assertIn("[AI-AUTOFIX:", patched)

    def test_autofix_instruction_flag_off_keeps_legacy_hunk_apply(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        self.app.autofix_structured_instruction_enabled = False
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)
        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 200)
        validation = apply_payload.get("validation", {})
        self.assertEqual(str(validation.get("instruction_mode", "off")), "off")
        self.assertFalse(bool(validation.get("instruction_apply_success", True)))

    def test_autofix_instruction_flag_on_applies_instruction_path(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        self.app.autofix_structured_instruction_enabled = True
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)
        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 200)
        validation = apply_payload.get("validation", {})
        self.assertEqual(str(validation.get("instruction_mode", "")), "applied")
        self.assertTrue(bool(validation.get("instruction_apply_success", False)))
        self.assertIn(str(validation.get("instruction_operation", "")), ("insert", "replace"))
        stats_status, stats_payload = self._request(
            "GET",
            "/api/autofix/stats?" + urllib.parse.urlencode({"session_id": analyze_payload.get("output_dir", "")}),
        )
        self.assertEqual(stats_status, 200)
        self.assertGreaterEqual(int(stats_payload.get("instruction_attempt_count", 0) or 0), 1)
        self.assertGreaterEqual(int(stats_payload.get("instruction_apply_success_count", 0) or 0), 1)

    def test_autofix_instruction_invalid_falls_back_to_hunks(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        self.app.autofix_structured_instruction_enabled = True
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)
        session = self.app._get_review_session(analyze_payload.get("output_dir", ""))
        with session["lock"]:
            proposal = session.get("autofix", {}).get("proposals", {}).get(prepare_payload.get("proposal_id"))
            self.assertIsNotNone(proposal)
            proposal["_structured_instruction"] = {
                "target": {"file": "sample.ctl", "object": "sample.ctl", "event": "Global"},
                "operation": "delete",
                "locator": {"kind": "anchor_context", "start_line": 1},
                "payload": {"code": "x"},
                "safety": {"requires_hash_match": True},
            }

        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 200)
        validation = apply_payload.get("validation", {})
        self.assertEqual(str(validation.get("instruction_mode", "")), "fallback_hunks")
        self.assertFalse(bool(validation.get("instruction_apply_success", True)))
        self.assertTrue(len(validation.get("instruction_validation_errors", []) or []) >= 1)
        stats_status, stats_payload = self._request(
            "GET",
            "/api/autofix/stats?" + urllib.parse.urlencode({"session_id": analyze_payload.get("output_dir", "")}),
        )
        self.assertEqual(stats_status, 200)
        self.assertGreaterEqual(int(stats_payload.get("instruction_attempt_count", 0) or 0), 1)
        self.assertGreaterEqual(int(stats_payload.get("instruction_validation_fail_count", 0) or 0), 1)
        self.assertGreaterEqual(int(stats_payload.get("instruction_fallback_to_hunk_count", 0) or 0), 1)

    def test_autofix_compare_proposals_include_structured_instruction_envelope(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "auto",
                "prepare_mode": "compare",
            },
        )
        self.assertEqual(prepare_status, 200)
        self.assertGreaterEqual(len(prepare_payload.get("proposals", []) or []), 1)
        session = self.app._get_review_session(analyze_payload.get("output_dir", ""))
        with session["lock"]:
            stored = session.get("autofix", {}).get("proposals", {})
            for view in (prepare_payload.get("proposals") or []):
                pid = str((view or {}).get("proposal_id", ""))
                self.assertTrue(pid)
                proposal = stored.get(pid, {})
                self.assertIsInstance(proposal.get("_structured_instruction"), dict)
                self.assertIsInstance((view or {}).get("instruction_preview"), dict)

    def test_autofix_compare_selection_prefers_valid_instruction_candidate(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        original_rule_builder = self.app._build_autofix_proposal_from_rule_template

        def _patched_rule_builder(*args, **kwargs):
            proposal = original_rule_builder(*args, **kwargs)
            proposal["_structured_instruction"] = {
                "target": {"file": "sample.ctl", "object": "sample.ctl", "event": "Global"},
                "operation": "delete",
                "locator": {"kind": "anchor_context", "start_line": 1},
                "payload": {"code": "x"},
                "safety": {"requires_hash_match": True},
            }
            return proposal

        self.app._build_autofix_proposal_from_rule_template = _patched_rule_builder
        try:
            status, analyze_payload = self._request(
                "POST",
                "/api/analyze",
                {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
            )
            self.assertEqual(status, 200)
            ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
            prepare_status, prepare_payload = self._request(
                "POST",
                "/api/autofix/prepare",
                {
                    "file": "sample.ctl",
                    "object": ai_review.get("object", "sample.ctl"),
                    "event": ai_review.get("event", "Global"),
                    "review": ai_review.get("review", ""),
                    "session_id": analyze_payload.get("output_dir", ""),
                    "generator_preference": "auto",
                    "prepare_mode": "compare",
                },
            )
            self.assertEqual(prepare_status, 200)
            compare_meta = prepare_payload.get("compare_meta", {})
            self.assertEqual(compare_meta.get("selection_policy"), "instruction_validity_then_syntax_then_rule")
            self.assertIsInstance(compare_meta.get("selected_compare_score", {}), dict)
            self.assertGreaterEqual(int((compare_meta.get("selected_compare_score", {}) or {}).get("total", 0) or 0), 0)
            selected_pid = str(prepare_payload.get("selected_proposal_id", ""))
            selected = None
            for item in (prepare_payload.get("proposals") or []):
                if str((item or {}).get("proposal_id", "")) == selected_pid:
                    selected = item
                    break
            self.assertIsNotNone(selected)
            self.assertEqual(str((selected or {}).get("generator_type", "")), "llm")
        finally:
            self.app._build_autofix_proposal_from_rule_template = original_rule_builder

    def test_autofix_apply_rejects_hash_mismatch(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)

        self.assertIn("generator_type", prepare_payload)
        self.assertIn("generator_reason", prepare_payload)
        self.assertIn("quality_preview", prepare_payload)
        self.assertEqual(prepare_payload.get("generator_type"), "llm")

        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": "deadbeef",
                "apply_mode": "source_ctl",
            },
        )
        self.assertEqual(apply_status, 409)
        self.assertEqual(apply_payload.get("error_code"), "BASE_HASH_MISMATCH")
        self.assertIn("quality_metrics", apply_payload)
        self.assertEqual((apply_payload.get("quality_metrics") or {}).get("hash_match"), False)
        self.assertIn("hash mismatch", apply_payload.get("error", "").lower())

    def test_autofix_apply_benchmark_relaxed_requires_env_flag(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)

        previous = os.environ.get("AUTOFIX_BENCHMARK_OBSERVE")
        try:
            os.environ.pop("AUTOFIX_BENCHMARK_OBSERVE", None)
            apply_status, apply_payload = self._request(
                "POST",
                "/api/autofix/apply",
                {
                    "proposal_id": prepare_payload.get("proposal_id"),
                    "session_id": analyze_payload.get("output_dir", ""),
                    "file": "sample.ctl",
                    "expected_base_hash": "deadbeef",
                    "apply_mode": "source_ctl",
                    "block_on_regression": False,
                },
                headers={"X-Autofix-Benchmark-Observe-Mode": "benchmark_relaxed"},
            )
        finally:
            if previous is None:
                os.environ.pop("AUTOFIX_BENCHMARK_OBSERVE", None)
            else:
                os.environ["AUTOFIX_BENCHMARK_OBSERVE"] = previous

        self.assertEqual(apply_status, 409)
        self.assertEqual(apply_payload.get("error_code"), "BASE_HASH_MISMATCH")

    def test_autofix_apply_benchmark_relaxed_bypasses_hash_gate_with_env_flag(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)

        previous = os.environ.get("AUTOFIX_BENCHMARK_OBSERVE")
        try:
            os.environ["AUTOFIX_BENCHMARK_OBSERVE"] = "1"
            apply_status, apply_payload = self._request(
                "POST",
                "/api/autofix/apply",
                {
                    "proposal_id": prepare_payload.get("proposal_id"),
                    "session_id": analyze_payload.get("output_dir", ""),
                    "file": "sample.ctl",
                    "expected_base_hash": "deadbeef",
                    "apply_mode": "source_ctl",
                    "block_on_regression": False,
                },
                headers={"X-Autofix-Benchmark-Observe-Mode": "benchmark_relaxed"},
            )
        finally:
            if previous is None:
                os.environ.pop("AUTOFIX_BENCHMARK_OBSERVE", None)
            else:
                os.environ["AUTOFIX_BENCHMARK_OBSERVE"] = previous

        self.assertEqual(apply_status, 200)
        validation = apply_payload.get("validation", {})
        self.assertTrue(bool(validation.get("hash_gate_bypassed", False)))
        self.assertEqual(validation.get("benchmark_observe_mode"), "benchmark_relaxed")

    def test_autofix_apply_benchmark_strict_hash_with_env_still_blocks_hash_mismatch(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)

        previous = os.environ.get("AUTOFIX_BENCHMARK_OBSERVE")
        try:
            os.environ["AUTOFIX_BENCHMARK_OBSERVE"] = "1"
            apply_status, apply_payload = self._request(
                "POST",
                "/api/autofix/apply",
                {
                    "proposal_id": prepare_payload.get("proposal_id"),
                    "session_id": analyze_payload.get("output_dir", ""),
                    "file": "sample.ctl",
                    "expected_base_hash": "deadbeef",
                    "apply_mode": "source_ctl",
                    "block_on_regression": False,
                },
                headers={"X-Autofix-Benchmark-Observe-Mode": "strict_hash"},
            )
        finally:
            if previous is None:
                os.environ.pop("AUTOFIX_BENCHMARK_OBSERVE", None)
            else:
                os.environ["AUTOFIX_BENCHMARK_OBSERVE"] = previous

        self.assertEqual(apply_status, 409)
        self.assertEqual(apply_payload.get("error_code"), "BASE_HASH_MISMATCH")

    def test_autofix_apply_benchmark_tuning_headers_ignored_without_observe_gate(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)
        previous = os.environ.get("AUTOFIX_BENCHMARK_OBSERVE")
        try:
            os.environ.pop("AUTOFIX_BENCHMARK_OBSERVE", None)
            apply_status, apply_payload = self._request(
                "POST",
                "/api/autofix/apply",
                {
                    "proposal_id": prepare_payload.get("proposal_id"),
                    "session_id": analyze_payload.get("output_dir", ""),
                    "file": "sample.ctl",
                    "expected_base_hash": prepare_payload.get("base_hash"),
                    "apply_mode": "source_ctl",
                    "block_on_regression": False,
                },
                headers={
                    "X-Autofix-Benchmark-Observe-Mode": "benchmark_relaxed",
                    "X-Autofix-Benchmark-Tuning-Min-Confidence": "0.55",
                    "X-Autofix-Benchmark-Tuning-Min-Gap": "0.05",
                    "X-Autofix-Benchmark-Tuning-Max-Line-Drift": "900",
                },
            )
        finally:
            if previous is None:
                os.environ.pop("AUTOFIX_BENCHMARK_OBSERVE", None)
            else:
                os.environ["AUTOFIX_BENCHMARK_OBSERVE"] = previous
        self.assertEqual(apply_status, 200)
        validation = apply_payload.get("validation", {})
        self.assertFalse(bool(validation.get("benchmark_tuning_applied", True)))
        self.assertEqual(float(validation.get("token_min_confidence_used", 0.0)), 0.8)
        self.assertEqual(float(validation.get("token_min_gap_used", 0.0)), 0.15)

    def test_autofix_apply_benchmark_tuning_headers_applied_with_observe_gate(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)
        previous = os.environ.get("AUTOFIX_BENCHMARK_OBSERVE")
        try:
            os.environ["AUTOFIX_BENCHMARK_OBSERVE"] = "1"
            apply_status, apply_payload = self._request(
                "POST",
                "/api/autofix/apply",
                {
                    "proposal_id": prepare_payload.get("proposal_id"),
                    "session_id": analyze_payload.get("output_dir", ""),
                    "file": "sample.ctl",
                    "expected_base_hash": "deadbeef",
                    "apply_mode": "source_ctl",
                    "block_on_regression": False,
                },
                headers={
                    "X-Autofix-Benchmark-Observe-Mode": "benchmark_relaxed",
                    "X-Autofix-Benchmark-Tuning-Min-Confidence": "0.55",
                    "X-Autofix-Benchmark-Tuning-Min-Gap": "0.05",
                    "X-Autofix-Benchmark-Tuning-Max-Line-Drift": "900",
                },
            )
        finally:
            if previous is None:
                os.environ.pop("AUTOFIX_BENCHMARK_OBSERVE", None)
            else:
                os.environ["AUTOFIX_BENCHMARK_OBSERVE"] = previous
        self.assertEqual(apply_status, 200)
        validation = apply_payload.get("validation", {})
        self.assertTrue(bool(validation.get("benchmark_tuning_applied", False)))
        self.assertAlmostEqual(float(validation.get("token_min_confidence_used", 0.0)), 0.55, places=3)
        self.assertAlmostEqual(float(validation.get("token_min_gap_used", 0.0)), 0.05, places=3)
        self.assertEqual(int(validation.get("token_max_line_drift_used", 0)), 900)

    def test_autofix_apply_rejects_invalid_benchmark_tuning_headers(self):
        self._force_single_internal_violation()
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": False, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        p1_group = (analyze_payload.get("violations", {}) or {}).get("P1", [])[0]
        violation = (p1_group.get("violations") or [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": p1_group.get("object", "sample.ctl"),
                "event": p1_group.get("event", "Global"),
                "review": "",
                "issue_id": violation.get("issue_id", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "rule",
            },
        )
        self.assertEqual(prepare_status, 200)
        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
            headers={
                "X-Autofix-Benchmark-Observe-Mode": "benchmark_relaxed",
                "X-Autofix-Benchmark-Tuning-Min-Confidence": "1.5",
            },
        )
        self.assertEqual(apply_status, 400)
        self.assertIn("Min-Confidence", str(apply_payload.get("error", "")))

    def test_autofix_prepare_rule_generator_without_review(self):
        self._force_single_internal_violation()
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": False, "mode": "Static"},
        )
        self.assertEqual(status, 200)

        p1_group = (analyze_payload.get("violations", {}) or {}).get("P1", [])[0]
        violation = (p1_group.get("violations") or [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": p1_group.get("object", "sample.ctl"),
                "event": p1_group.get("event", "Global"),
                "review": "",
                "issue_id": violation.get("issue_id", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "rule",
            },
        )
        self.assertEqual(prepare_status, 200)
        self.assertEqual(prepare_payload.get("generator_type"), "rule")
        self.assertEqual(prepare_payload.get("source"), "rule-template")
        self.assertIn("quality_preview", prepare_payload)
        self.assertIn("generator_reason", prepare_payload)
        self.assertTrue(prepare_payload.get("unified_diff"))

    def test_autofix_stats_endpoint_reports_generator_counts(self):
        self._force_single_internal_violation()
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": False, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        p1_group = (analyze_payload.get("violations", {}) or {}).get("P1", [])[0]
        violation = (p1_group.get("violations") or [])[0]
        prepare_status, _prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": p1_group.get("object", "sample.ctl"),
                "event": p1_group.get("event", "Global"),
                "review": "",
                "issue_id": violation.get("issue_id", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "rule",
            },
        )
        self.assertEqual(prepare_status, 200)

        stats_status, stats_payload = self._request(
            "GET",
            "/api/autofix/stats?" + urllib.parse.urlencode({"session_id": analyze_payload.get("output_dir", "")}),
        )
        self.assertEqual(stats_status, 200)
        self.assertGreaterEqual((stats_payload.get("by_generator") or {}).get("rule", 0), 1)
        self.assertGreaterEqual(stats_payload.get("proposal_count", 0), 1)

    def test_autofix_prepare_compare_mode_returns_candidates(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "auto",
                "prepare_mode": "compare",
            },
        )
        self.assertEqual(prepare_status, 200)
        proposals = prepare_payload.get("proposals", [])
        self.assertGreaterEqual(len(proposals), 1)
        self.assertLessEqual(len(proposals), 2)
        proposal_ids = {str(p.get("proposal_id", "")) for p in proposals if isinstance(p, dict)}
        self.assertIn(str(prepare_payload.get("selected_proposal_id", "")), proposal_ids)
        compare_meta = prepare_payload.get("compare_meta", {})
        self.assertEqual(compare_meta.get("mode"), "compare")
        self.assertGreaterEqual(int(compare_meta.get("generated_count", 0) or 0), 1)

    def test_autofix_prepare_compare_mode_validation(self):
        status, payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": "sample.ctl",
                "event": "Global",
                "review": "",
                "session_id": "dummy",
                "prepare_mode": "invalid-mode",
            },
        )
        self.assertEqual(status, 400)
        self.assertIn("prepare_mode", str(payload.get("error", "")))

    def test_autofix_stats_compare_mode_selected_generator_counts(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "prepare_mode": "compare",
                "generator_preference": "auto",
            },
        )
        self.assertEqual(prepare_status, 200)
        proposals = prepare_payload.get("proposals", [])
        self.assertGreaterEqual(len(proposals), 1)
        selected = next(
            (p for p in proposals if str(p.get("proposal_id", "")) == str(prepare_payload.get("selected_proposal_id", ""))),
            proposals[0],
        )
        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": selected.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": selected.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 200)
        stats_status, stats_payload = self._request(
            "GET",
            "/api/autofix/stats?" + urllib.parse.urlencode({"session_id": analyze_payload.get("output_dir", "")}),
        )
        self.assertEqual(stats_status, 200)
        self.assertGreaterEqual(int(stats_payload.get("prepare_compare_count", 0) or 0), 1)
        self.assertGreaterEqual(int(stats_payload.get("compare_apply_count", 0) or 0), 1)
        self.assertIn("apply_engine_structure_success_count", stats_payload)
        self.assertIn("apply_engine_text_fallback_count", stats_payload)
        self.assertIn("selected_apply_engine_mode", stats_payload)
        selected_counts = stats_payload.get("selected_generator_counts", {}) or {}
        selected_gen = str(selected.get("generator_type", "")).lower()
        if selected_gen in ("rule", "llm"):
            self.assertGreaterEqual(int(selected_counts.get(selected_gen, 0) or 0), 1)

    def test_autofix_apply_token_fallback_success(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "llm",
            },
        )
        self.assertEqual(prepare_status, 200)
        session = self.app._get_review_session(analyze_payload.get("output_dir", ""))
        self.assertIsNotNone(session)
        with session["lock"]:
            proposal = session.get("autofix", {}).get("proposals", {}).get(prepare_payload.get("proposal_id"))
            self.assertIsNotNone(proposal)
            hunks = proposal.get("hunks", [])
            self.assertTrue(hunks and isinstance(hunks[0], dict))
            hunks[0]["context_after"] = 'main(){dpSet("A.B.C",1);}'

        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 200)
        validation = apply_payload.get("validation", {})
        self.assertIn(validation.get("apply_engine_mode"), ("structure_apply", "text_fallback"))
        self.assertEqual(validation.get("locator_mode"), "token_fallback")
        self.assertTrue(validation.get("token_fallback_attempted"))
        self.assertGreaterEqual(float(validation.get("token_fallback_confidence", 0.0) or 0.0), 0.8)

    def test_autofix_apply_anchor_normalized_success(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "llm",
            },
        )
        self.assertEqual(prepare_status, 200)
        session = self.app._get_review_session(analyze_payload.get("output_dir", ""))
        with session["lock"]:
            proposal = session.get("autofix", {}).get("proposals", {}).get(prepare_payload.get("proposal_id"))
            hunks = proposal.get("hunks", [])
            self.assertTrue(hunks and isinstance(hunks[0], dict))
            # Same semantic line with whitespace differences should pass normalized anchor mode.
            hunks[0]["context_after"] = 'main()    {   dpSet("A.B.C", 1);    }'

        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 200)
        validation = apply_payload.get("validation", {})
        self.assertEqual(validation.get("locator_mode"), "anchor_normalized")
        self.assertFalse(validation.get("token_fallback_attempted"))

    def test_autofix_apply_token_fallback_ambiguous_fail_soft(self):
        with open(os.path.join(self.data_dir, "sample.ctl"), "w", encoding="utf-8") as f:
            f.write('main() { dpSet("A.B.C", 1); }\nmain() { dpSet("A.B.C", 1); }')
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "llm",
            },
        )
        self.assertEqual(prepare_status, 200)
        session = self.app._get_review_session(analyze_payload.get("output_dir", ""))
        with session["lock"]:
            proposal = session.get("autofix", {}).get("proposals", {}).get(prepare_payload.get("proposal_id"))
            hunks = proposal.get("hunks", [])
            hunks[0]["context_after"] = 'main(){dpSet("A.B.C",1);}'
        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
            },
        )
        self.assertEqual(apply_status, 409)
        self.assertEqual(apply_payload.get("error_code"), "ANCHOR_MISMATCH")
        quality = apply_payload.get("quality_metrics", {})
        self.assertTrue(quality.get("token_fallback_attempted"))
        self.assertGreaterEqual(int(quality.get("token_fallback_candidates", 0) or 0), 2)

        stats_status, stats_payload = self._request(
            "GET",
            "/api/autofix/stats?" + urllib.parse.urlencode({"session_id": analyze_payload.get("output_dir", "")}),
        )
        self.assertEqual(stats_status, 200)
        self.assertGreaterEqual(int(stats_payload.get("anchor_mismatch_count", 0) or 0), 1)
        self.assertGreaterEqual(int(stats_payload.get("token_fallback_attempt_count", 0) or 0), 1)
        self.assertGreaterEqual(int(stats_payload.get("token_fallback_ambiguous_count", 0) or 0), 1)

    def test_autofix_apply_semantic_guard_blocked_and_stats(self):
        self._force_single_internal_violation()
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": False, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        p1_group = (analyze_payload.get("violations", {}) or {}).get("P1", [])[0]
        violation = (p1_group.get("violations") or [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": p1_group.get("object", "sample.ctl"),
                "event": p1_group.get("event", "Global"),
                "review": "",
                "issue_id": violation.get("issue_id", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "rule",
            },
        )
        self.assertEqual(prepare_status, 200)
        session = self.app._get_review_session(analyze_payload.get("output_dir", ""))
        with session["lock"]:
            proposal = session.get("autofix", {}).get("proposals", {}).get(prepare_payload.get("proposal_id"))
            self.assertIsNotNone(proposal)
            hunks = proposal.get("hunks", [])
            self.assertTrue(hunks and isinstance(hunks[0], dict))
            original_candidate = str(proposal.get("_candidate_content", ""))
            hunks[0]["replacement_text"] = 'main() { dpSet("A.B.D", 1); }'
            proposal["_candidate_content"] = original_candidate

        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 409)
        self.assertEqual(apply_payload.get("error_code"), "SEMANTIC_GUARD_BLOCKED")
        quality = apply_payload.get("quality_metrics", {})
        self.assertFalse(bool(quality.get("semantic_check_passed", True)))
        self.assertGreaterEqual(int(quality.get("semantic_violation_count", 0) or 0), 1)
        self.assertIn("semantic guard blocked", str(quality.get("rejected_reason", "")).lower())

        stats_status, stats_payload = self._request(
            "GET",
            "/api/autofix/stats?" + urllib.parse.urlencode({"session_id": analyze_payload.get("output_dir", "")}),
        )
        self.assertEqual(stats_status, 200)
        self.assertGreaterEqual(int(stats_payload.get("semantic_guard_checked_count", 0) or 0), 1)
        self.assertGreaterEqual(int(stats_payload.get("semantic_guard_blocked_count", 0) or 0), 1)

    def test_autofix_apply_multi_hunk_success_and_stats(self):
        with open(os.path.join(self.data_dir, "sample.ctl"), "w", encoding="utf-8") as f:
            f.write("main()\n{\n  int a = 1;\n  int b = 2;\n  return;\n}\n")
        self._force_single_internal_violation()
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": False, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        p1_group = (analyze_payload.get("violations", {}) or {}).get("P1", [])[0]
        violation = (p1_group.get("violations") or [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": p1_group.get("object", "sample.ctl"),
                "event": p1_group.get("event", "Global"),
                "review": "",
                "issue_id": violation.get("issue_id", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "rule",
            },
        )
        self.assertEqual(prepare_status, 200)
        with open(os.path.join(self.data_dir, "sample.ctl"), "r", encoding="utf-8", errors="ignore") as f:
            current_lines = f.read().splitlines()
        session = self.app._get_review_session(analyze_payload.get("output_dir", ""))
        with session["lock"]:
            proposal = session.get("autofix", {}).get("proposals", {}).get(prepare_payload.get("proposal_id"))
            self.assertIsNotNone(proposal)
            proposal["hunks"] = [
                {
                    "start_line": 3,
                    "end_line": 3,
                    "context_before": current_lines[1],
                    "context_after": current_lines[2],
                    "replacement_text": "  int a = 10;",
                },
                {
                    "start_line": 4,
                    "end_line": 4,
                    "context_before": current_lines[2],
                    "context_after": current_lines[3],
                    "replacement_text": "  int b = 20;",
                },
            ]
            proposal["_candidate_content"] = "main()\n{\n  int a = 10;\n  int b = 20;\n  return;\n}\n"

        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 200)
        self.assertEqual((apply_payload.get("validation") or {}).get("apply_engine_mode"), "structure_apply")

        stats_status, stats_payload = self._request(
            "GET",
            "/api/autofix/stats?" + urllib.parse.urlencode({"session_id": analyze_payload.get("output_dir", "")}),
        )
        self.assertEqual(stats_status, 200)
        self.assertGreaterEqual(int(stats_payload.get("multi_hunk_attempt_count", 0) or 0), 1)
        self.assertGreaterEqual(int(stats_payload.get("multi_hunk_success_count", 0) or 0), 1)

    def test_autofix_apply_multi_hunk_overlap_blocked(self):
        with open(os.path.join(self.data_dir, "sample.ctl"), "w", encoding="utf-8") as f:
            f.write("main()\n{\n  int a = 1;\n  int b = 2;\n  return;\n}\n")
        self._force_single_internal_violation()
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": False, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        p1_group = (analyze_payload.get("violations", {}) or {}).get("P1", [])[0]
        violation = (p1_group.get("violations") or [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": p1_group.get("object", "sample.ctl"),
                "event": p1_group.get("event", "Global"),
                "review": "",
                "issue_id": violation.get("issue_id", ""),
                "session_id": analyze_payload.get("output_dir", ""),
                "generator_preference": "rule",
            },
        )
        self.assertEqual(prepare_status, 200)
        with open(os.path.join(self.data_dir, "sample.ctl"), "r", encoding="utf-8", errors="ignore") as f:
            current_lines = f.read().splitlines()
        session = self.app._get_review_session(analyze_payload.get("output_dir", ""))
        with session["lock"]:
            proposal = session.get("autofix", {}).get("proposals", {}).get(prepare_payload.get("proposal_id"))
            self.assertIsNotNone(proposal)
            proposal["hunks"] = [
                {
                    "start_line": 3,
                    "end_line": 4,
                    "context_before": current_lines[1],
                    "context_after": current_lines[2],
                    "replacement_text": "  int a = 10;\n  int b = 20;",
                },
                {
                    "start_line": 4,
                    "end_line": 4,
                    "context_before": current_lines[2],
                    "context_after": current_lines[3],
                    "replacement_text": "  int b = 200;",
                },
            ]
            proposal["_candidate_content"] = "main()\n{\n  int a = 10;\n  int b = 200;\n  return;\n}\n"

        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
            },
        )
        self.assertEqual(apply_status, 409)
        self.assertEqual(apply_payload.get("error_code"), "APPLY_ENGINE_FAILED")
        quality = apply_payload.get("quality_metrics", {}) or {}
        self.assertEqual(quality.get("apply_engine_mode"), "failed")
        self.assertEqual(quality.get("apply_engine_fallback_reason"), "overlapping_hunks")

        stats_status, stats_payload = self._request(
            "GET",
            "/api/autofix/stats?" + urllib.parse.urlencode({"session_id": analyze_payload.get("output_dir", "")}),
        )
        self.assertEqual(stats_status, 200)
        self.assertGreaterEqual(int(stats_payload.get("multi_hunk_attempt_count", 0) or 0), 1)
        self.assertGreaterEqual(int(stats_payload.get("multi_hunk_blocked_count", 0) or 0), 1)

    def test_autofix_apply_ctrlpp_regression_count_is_reported(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )

        original_run_check = self.app.ctrl_tool.run_check

        def fake_run_check(file_path, code_content=None, enabled=None, binary_path=None):
            try:
                with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
                    text = f.read()
            except Exception:
                text = str(code_content or "")
            if "[AI-AUTOFIX:" in text:
                return [
                    {
                        "type": "warning",
                        "severity": "warning",
                        "rule_id": "ctrlppcheck.test",
                        "line": 1,
                        "message": "mock ctrlpp regression",
                        "source": "CtrlppCheck",
                        "priority_origin": "P2",
                    }
                ]
            return []

        self.app.ctrl_tool.run_check = fake_run_check
        try:
            status, analyze_payload = self._request(
                "POST",
                "/api/analyze",
                {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
            )
            self.assertEqual(status, 200)
            ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
            prepare_status, prepare_payload = self._request(
                "POST",
                "/api/autofix/prepare",
                {
                    "file": "sample.ctl",
                    "object": ai_review.get("object", "sample.ctl"),
                    "event": ai_review.get("event", "Global"),
                    "review": ai_review.get("review", ""),
                    "session_id": analyze_payload.get("output_dir", ""),
                },
            )
            self.assertEqual(prepare_status, 200)

            apply_status, apply_payload = self._request(
                "POST",
                "/api/autofix/apply",
                {
                    "proposal_id": prepare_payload.get("proposal_id"),
                    "session_id": analyze_payload.get("output_dir", ""),
                    "file": "sample.ctl",
                    "expected_base_hash": prepare_payload.get("base_hash"),
                    "apply_mode": "source_ctl",
                    "block_on_regression": False,
                    "check_ctrlpp_regression": True,
                },
            )
            self.assertEqual(apply_status, 200)
            validation = apply_payload.get("validation", {})
            self.assertEqual(validation.get("ctrlpp_regression_count"), 1)
        finally:
            self.app.ctrl_tool.run_check = original_run_check

    def test_autofix_apply_ctrlpp_regression_check_real_binary_optional(self):
        if os.environ.get("RUN_CTRLPPCHECK_INTEGRATION", "").strip() != "1":
            self.skipTest("Set RUN_CTRLPPCHECK_INTEGRATION=1 to run CtrlppCheck integration test")
        binary = self.app.ctrl_tool._find_binary()
        if not binary:
            self.skipTest("CtrlppCheck binary not found")
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": analyze_payload.get("output_dir", ""),
            },
        )
        self.assertEqual(prepare_status, 200)
        apply_status, apply_payload = self._request(
            "POST",
            "/api/autofix/apply",
            {
                "proposal_id": prepare_payload.get("proposal_id"),
                "session_id": analyze_payload.get("output_dir", ""),
                "file": "sample.ctl",
                "expected_base_hash": prepare_payload.get("base_hash"),
                "apply_mode": "source_ctl",
                "block_on_regression": False,
                "check_ctrlpp_regression": True,
            },
        )
        self.assertIn(apply_status, (200, 409))
        if apply_status == 200:
            self.assertIn("validation", apply_payload)
            self.assertIn("ctrlpp_regression_count", apply_payload.get("validation", {}))

    def test_autofix_session_ttl_eviction_returns_409(self):
        self._force_single_internal_violation()
        self.app.ai_tool.generate_review = lambda *_args, **_kwargs: (
            "요약: 조건 검증을 추가하세요.\n\n"
            "코드:\n```cpp\nif (isValid) {\n  return;\n}\n```"
        )
        status, analyze_payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        session_id = analyze_payload.get("output_dir", "")
        ai_review = analyze_payload.get("violations", {}).get("P3", [])[0]
        prepare_status, _prepare_payload = self._request(
            "POST",
            "/api/autofix/prepare",
            {
                "file": "sample.ctl",
                "object": ai_review.get("object", "sample.ctl"),
                "event": ai_review.get("event", "Global"),
                "review": ai_review.get("review", ""),
                "session_id": session_id,
            },
        )
        self.assertEqual(prepare_status, 200)
        session = self.app._get_review_session(session_id)
        self.assertIsNotNone(session)
        self.app.review_session_ttl_sec = 60
        session["last_accessed_at"] = 0

        diff_status, diff_payload = self._request(
            "GET",
            "/api/autofix/file-diff?" + urllib.parse.urlencode({"file": "sample.ctl", "session_id": session_id}),
        )
        self.assertEqual(diff_status, 409)
        self.assertIn("session cache", diff_payload.get("error", "").lower())

    def test_post_api_analyze_live_ai_off_hides_p3(self):
        self._force_single_internal_violation()
        self.app.ai_tool.get_mock_review = lambda *_args, **_kwargs: "MOCK REVIEW SHOULD NOT APPEAR"

        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": False, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        p3 = payload.get("violations", {}).get("P3", [])
        self.assertEqual(p3, [])
        self.assertEqual(payload.get("summary", {}).get("p3_total"), 0)

    def test_post_api_analyze_live_ai_fail_soft_keeps_200(self):
        self._force_single_internal_violation()
        self.app.ai_tool.provider = "unsupported-provider"
        self.app.ai_tool.fail_soft = True

        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_live_ai": True, "mode": "AI 보조"},
        )
        self.assertEqual(status, 200)
        p3 = payload.get("violations", {}).get("P3", [])
        self.assertEqual(p3, [])
        self.assertEqual(payload.get("summary", {}).get("p3_total"), 0)

    def test_post_api_analyze_ai_with_context_live_on_uses_context(self):
        self._force_single_internal_violation()
        calls = {"count": 0}

        def fake_fetch_context():
            calls["count"] += 1
            return {
                "enabled": True,
                "project": {"projectName": "TEST_PROJ"},
                "drivers": [{"name": "CTRL"}],
                "error": None,
            }

        def fake_live_review(_code, _violations, use_context=False, context_payload=None, focus_snippet=""):
            if use_context and isinstance(context_payload, dict) and context_payload.get("enabled"):
                return "LIVE REVIEW WITH CONTEXT"
            return "LIVE REVIEW NO CONTEXT"

        self.app.mcp_tool.fetch_context = fake_fetch_context
        self.app.ai_tool.generate_review = fake_live_review

        status, payload = self._request(
            "POST",
            "/api/analyze",
            {
                "selected_files": ["sample.ctl"],
                "enable_live_ai": True,
                "ai_with_context": True,
                "mode": "AI 보조",
            },
        )
        self.assertEqual(status, 200)
        self.assertEqual(calls["count"], 1)
        p3 = payload.get("violations", {}).get("P3", [])
        self.assertGreaterEqual(len(p3), 1)
        self.assertEqual(p3[0].get("review"), "LIVE REVIEW WITH CONTEXT")

    def test_post_api_analyze_ai_with_context_fail_soft_keeps_200(self):
        self._force_single_internal_violation()

        def failing_fetch_context():
            return {"enabled": False, "project": {}, "drivers": [], "error": "mcp timeout"}

        def fake_live_review(_code, _violations, use_context=False, context_payload=None, focus_snippet=""):
            if use_context and isinstance(context_payload, dict) and not context_payload.get("enabled"):
                return "LIVE REVIEW CONTEXT_FALLBACK"
            return "LIVE REVIEW"

        self.app.mcp_tool.fetch_context = failing_fetch_context
        self.app.ai_tool.generate_review = fake_live_review

        status, payload = self._request(
            "POST",
            "/api/analyze",
            {
                "selected_files": ["sample.ctl"],
                "enable_live_ai": True,
                "ai_with_context": True,
                "mode": "AI 보조",
            },
        )
        self.assertEqual(status, 200)
        p3 = payload.get("violations", {}).get("P3", [])
        self.assertGreaterEqual(len(p3), 1)
        self.assertEqual(p3[0].get("review"), "LIVE REVIEW CONTEXT_FALLBACK")

    def test_post_api_analyze_raw_txt_rejected_without_toggle(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["raw_input.txt"]},
        )
        self.assertEqual(status, 400)
        self.assertIn("error", payload)
        self.assertIn("allow_raw_txt", payload["error"])

    def test_post_api_analyze_raw_txt_allowed_with_toggle(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["raw_input.txt"], "allow_raw_txt": True},
        )
        self.assertEqual(status, 200)
        self.assertIn("summary", payload)

    def test_post_api_analyze_ctrlpp_toggle_off(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_ctrlppcheck": False, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        self.assertEqual(payload.get("violations", {}).get("P2", []), [])
        summary = payload.get("summary", {})
        self.assertFalse(bool(summary.get("ctrlpp_preflight_attempted", False)))
        self.assertFalse(bool(summary.get("ctrlpp_preflight_ready", False)))

    def test_post_api_analyze_ctrlpp_toggle_on_missing_binary(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_ctrlppcheck": True, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        p2 = payload.get("violations", {}).get("P2", [])
        self.assertGreaterEqual(len(p2), 1)
        self.assertTrue(any(v.get("source") == "CtrlppCheck" for v in p2))
        summary = payload.get("summary", {})
        self.assertFalse(bool(summary.get("ctrlpp_preflight_ready", False)))
        self.assertTrue("ctrlpp" in str(summary.get("ctrlpp_preflight_message", "")).lower())

    def test_post_api_analyze_ctrlpp_toggle_type_validation(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_ctrlppcheck": "yes"},
        )
        self.assertEqual(status, 400)
        self.assertIn("error", payload)
        self.assertIn("enable_ctrlppcheck", payload["error"])

    def test_api_ctrlpp_toggle_on_triggers_autoinstall_path(self):
        self.app.ctrl_tool.auto_install_on_missing = True
        calls = {"count": 0}

        def fake_ensure():
            calls["count"] += 1
            return "", "CtrlppCheck install failed: mocked"

        self.app.ctrl_tool.ensure_installed = fake_ensure
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_ctrlppcheck": True, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        self.assertEqual(calls["count"], 1)
        p2 = payload.get("violations", {}).get("P2", [])
        self.assertTrue(any("install failed" in (v.get("message", "").lower()) for v in p2))
        summary = payload.get("summary", {})
        self.assertTrue(bool(summary.get("ctrlpp_preflight_attempted", False)))
        self.assertFalse(bool(summary.get("ctrlpp_preflight_ready", False)))

    def test_api_ctrlpp_toggle_on_install_failure_returns_200_with_p2_info(self):
        self.app.ctrl_tool.auto_install_on_missing = True

        def failing_ensure():
            raise RuntimeError("CtrlppCheck download failed: mocked offline")

        self.app.ctrl_tool.ensure_installed = failing_ensure
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "enable_ctrlppcheck": True, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        p2 = payload.get("violations", {}).get("P2", [])
        self.assertGreaterEqual(len(p2), 1)
        self.assertTrue(any("download failed" in (v.get("message", "").lower()) for v in p2))
        summary = payload.get("summary", {})
        self.assertTrue(bool(summary.get("ctrlpp_preflight_attempted", False)))
        self.assertFalse(bool(summary.get("ctrlpp_preflight_ready", False)))

    def test_api_ctrlpp_preflight_skips_without_ctl_target(self):
        self.app.ctrl_tool.auto_install_on_missing = True
        calls = {"count": 0}

        def fake_ensure():
            calls["count"] += 1
            return "", "CtrlppCheck install failed: mocked"

        self.app.ctrl_tool.ensure_installed = fake_ensure
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["raw_input.txt"], "allow_raw_txt": True, "enable_ctrlppcheck": True, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        self.assertEqual(calls["count"], 0)
        summary = payload.get("summary", {})
        self.assertFalse(bool(summary.get("ctrlpp_preflight_attempted", False)))

    def test_api_analyze_ctl_filters_to_server_rules(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["server_loop.ctl"], "mode": "Static"},
        )
        self.assertEqual(status, 200)
        compact_items = [item.replace(" ", "") for item in self._extract_rule_items(payload)]
        self.assertTrue(any("Loop문내" in item and "처리조건" in item for item in compact_items))

    def test_api_analyze_raw_txt_filters_to_client_rules(self):
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["raw_loop.txt"], "allow_raw_txt": True, "mode": "Static"},
        )
        self.assertEqual(status, 200)
        compact_items = [item.replace(" ", "") for item in self._extract_rule_items(payload)]
        self.assertFalse(any("Loop문내" in item and "처리조건" in item for item in compact_items))

    def test_post_api_analyze_file_failure_exposes_errors_in_payload(self):
        def failing_analyze_file(*_args, **_kwargs):
            raise RuntimeError("forced failure for test")

        self.app.analyze_file = failing_analyze_file
        status, payload = self._request(
            "POST",
            "/api/analyze",
            {"selected_files": ["sample.ctl"], "mode": "Static", "enable_ctrlppcheck": False, "enable_live_ai": False},
        )
        self.assertEqual(status, 500)
        self.assertIn("errors", payload)
        self.assertGreaterEqual(len(payload["errors"]), 1)
        self.assertEqual(payload["errors"][0]["file"], "sample.ctl")
        self.assertIn("forced failure for test", payload["errors"][0]["error"])
        self.assertEqual(payload.get("summary", {}).get("failed_file_count"), 1)
        self.assertEqual(payload.get("summary", {}).get("successful_file_count"), 0)

    def test_post_api_analyze_partial_file_failure_returns_207(self):
        original_analyze_file = self.app.analyze_file

        def partially_failing_analyze_file(target, *args, **kwargs):
            if os.path.basename(str(target)) == "server_loop.ctl":
                raise RuntimeError("forced partial failure")
            return original_analyze_file(target, *args, **kwargs)

        self.app.analyze_file = partially_failing_analyze_file
        try:
            status, payload = self._request(
                "POST",
                "/api/analyze",
                {
                    "selected_files": ["sample.ctl", "server_loop.ctl"],
                    "mode": "Static",
                    "enable_ctrlppcheck": False,
                    "enable_live_ai": False,
                },
            )
        finally:
            self.app.analyze_file = original_analyze_file

        self.assertEqual(status, 207)
        self.assertEqual(payload.get("summary", {}).get("requested_file_count"), 2)
        self.assertEqual(payload.get("summary", {}).get("successful_file_count"), 1)
        self.assertEqual(payload.get("summary", {}).get("failed_file_count"), 1)
        self.assertGreaterEqual(len(payload.get("errors", [])), 1)
        self.assertEqual(payload["errors"][0].get("file"), "server_loop.ctl")

    def test_concurrent_analyze_requests_are_isolated(self):
        payload_a = {
            "selected_files": ["sample.ctl"],
            "mode": "Static",
            "enable_ctrlppcheck": False,
            "enable_live_ai": False,
        }
        payload_b = {
            "selected_files": ["server_loop.ctl"],
            "mode": "Static",
            "enable_ctrlppcheck": False,
            "enable_live_ai": False,
        }

        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as pool:
            future_a = pool.submit(self._request, "POST", "/api/analyze", payload_a)
            future_b = pool.submit(self._request, "POST", "/api/analyze", payload_b)
            status_a, resp_a = future_a.result(timeout=60)
            status_b, resp_b = future_b.result(timeout=60)

        self.assertEqual(status_a, 200)
        self.assertEqual(status_b, 200)

        output_a = os.path.normpath(resp_a.get("output_dir", ""))
        output_b = os.path.normpath(resp_b.get("output_dir", ""))
        self.assertTrue(os.path.isdir(output_a))
        self.assertTrue(os.path.isdir(output_b))
        self.assertNotEqual(output_a, output_b)

        reviewed_a = set(resp_a.get("report_paths", {}).get("reviewed_txt", []))
        reviewed_b = set(resp_b.get("report_paths", {}).get("reviewed_txt", []))
        self.assertIn("sample_REVIEWED.txt", reviewed_a)
        self.assertIn("server_loop_REVIEWED.txt", reviewed_b)
        self.assertNotIn("server_loop_REVIEWED.txt", reviewed_a)
        self.assertNotIn("sample_REVIEWED.txt", reviewed_b)

    def test_concurrent_analyze_no_report_path_cross_over(self):
        payload_a = {
            "selected_files": ["sample.ctl"],
            "mode": "Static",
            "enable_ctrlppcheck": False,
            "enable_live_ai": False,
        }
        payload_b = {
            "selected_files": ["server_loop.ctl"],
            "mode": "Static",
            "enable_ctrlppcheck": False,
            "enable_live_ai": False,
        }

        with concurrent.futures.ThreadPoolExecutor(max_workers=2) as pool:
            future_a = pool.submit(self._request, "POST", "/api/analyze", payload_a)
            future_b = pool.submit(self._request, "POST", "/api/analyze", payload_b)
            status_a, resp_a = future_a.result(timeout=60)
            status_b, resp_b = future_b.result(timeout=60)

        self.assertEqual(status_a, 200)
        self.assertEqual(status_b, 200)

        output_a = os.path.normpath(resp_a.get("output_dir", ""))
        output_b = os.path.normpath(resp_b.get("output_dir", ""))
        excel_a = resp_a.get("report_paths", {}).get("excel", [])
        excel_b = resp_b.get("report_paths", {}).get("excel", [])
        reviewed_a = resp_a.get("report_paths", {}).get("reviewed_txt", [])
        reviewed_b = resp_b.get("report_paths", {}).get("reviewed_txt", [])

        for name in excel_a + reviewed_a:
            self.assertTrue(os.path.exists(os.path.join(output_a, name)))
            self.assertFalse(os.path.exists(os.path.join(output_b, name)))

        for name in excel_b + reviewed_b:
            self.assertTrue(os.path.exists(os.path.join(output_b, name)))
            self.assertFalse(os.path.exists(os.path.join(output_a, name)))

    def test_concurrent_analyze_requests_can_overlap_during_analysis(self):
        first_started = threading.Event()
        second_started = threading.Event()
        release_first = threading.Event()

        original_analyze_file = self.app.analyze_file

        def fake_analyze_file(*args, **kwargs):
            target = args[0] if args else kwargs.get("target", "")
            name = os.path.basename(str(target))
            if name == "sample.ctl":
                first_started.set()
                if not release_first.wait(timeout=5):
                    raise RuntimeError("timeout waiting for release_first")
            elif name == "server_loop.ctl":
                second_started.set()
            return {
                "file": name,
                "internal_violations": [],
                "global_violations": [],
                "ai_reviews": [],
            }

        self.app.analyze_file = fake_analyze_file
        try:
            payload_a = {
                "selected_files": ["sample.ctl"],
                "mode": "Static",
                "enable_ctrlppcheck": False,
                "enable_live_ai": False,
            }
            payload_b = {
                "selected_files": ["server_loop.ctl"],
                "mode": "Static",
                "enable_ctrlppcheck": False,
                "enable_live_ai": False,
            }

            with concurrent.futures.ThreadPoolExecutor(max_workers=2) as pool:
                future_a = pool.submit(self._request, "POST", "/api/analyze", payload_a)
                self.assertTrue(first_started.wait(timeout=5))
                future_b = pool.submit(self._request, "POST", "/api/analyze", payload_b)
                self.assertTrue(
                    second_started.wait(timeout=2),
                    "Second request should reach analyze_file while first request is still blocked",
                )
                release_first.set()
                status_a, _resp_a = future_a.result(timeout=30)
                status_b, _resp_b = future_b.result(timeout=30)

            self.assertEqual(status_a, 200)
            self.assertEqual(status_b, 200)
        finally:
            self.app.analyze_file = original_analyze_file
            release_first.set()


class ReportQualityTests(unittest.TestCase):
    def setUp(self):
        self.project_root = PROJECT_ROOT
        self.config_dir = os.path.join(self.project_root, "Config")
        self.tmp_dir = tempfile.TemporaryDirectory()

        self.reporter = Reporter(config_dir=self.config_dir)
        self.reporter.output_dir = self.tmp_dir.name
        os.makedirs(self.reporter.output_dir, exist_ok=True)

    def tearDown(self):
        self.tmp_dir.cleanup()

    @staticmethod
    def _find_status_col(ws):
        for r in range(1, min(ws.max_row, 40) + 1):
            for c in range(1, min(ws.max_column, 20) + 1):
                value = str(ws.cell(r, c).value or "").strip()
                if "검증 결과" in value:
                    return c
        return 6

    def _sample_report_data(self):
        return {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-PERF-02-test1",
                            "rule_id": "PERF-02",
                            "rule_item": "DP Query 최적화 구현",
                            "priority_origin": "P1",
                            "severity": "Critical",
                            "line": 1,
                            "message": "query scope too wide",
                        },
                        {
                            "issue_id": "P1-UNKNOWN-test2",
                            "rule_id": "UNKNOWN-99",
                            "rule_item": "non-matching-item",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 2,
                            "message": "unmatched violation",
                        },
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }

    def _sample_report_with_p2_only(self):
        return {
            "file": "sample.ctl",
            "internal_violations": [],
            "global_violations": [
                {
                    "rule_id": "PERF-02",
                    "line": 10,
                    "message": "p2 only finding",
                    "severity": "warning",
                    "source": "CtrlppCheck",
                }
            ],
            "ai_reviews": [],
        }

    @staticmethod
    def _find_item_row(ws, item_text):
        for r in range(1, ws.max_row + 1):
            if str(ws.cell(r, 4).value or "").strip() == item_text:
                return r
        return None

    def test_html_report_contains_rows_and_severity_class(self):
        data = self._sample_report_data()
        self.reporter.generate_html_report(data, "quality.html")
        report_path = os.path.join(self.reporter.output_dir, "quality.html")

        self.assertTrue(os.path.exists(report_path))
        with open(report_path, "r", encoding="utf-8") as f:
            html = f.read()

        self.assertIn("<table>", html)
        self.assertIn('class="critical"', html)
        self.assertIn("query scope too wide", html)

    def test_excel_report_creates_unmatched_sheet_and_marks_ng(self):
        data = self._sample_report_data()
        output_name = "quality.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)

        output_path = os.path.join(self.reporter.output_dir, output_name)
        self.assertTrue(os.path.exists(output_path))

        from openpyxl import load_workbook

        wb = load_workbook(output_path)
        active = wb.active
        status_col = self._find_status_col(active)
        col_values = [active.cell(row=r, column=status_col).value for r in range(1, active.max_row + 1)]
        self.assertIn("NG", col_values)
        self.assertIn("미분류_위반사항", wb.sheetnames)

    def test_excel_report_returns_timing_metrics_and_template_cache_hits(self):
        data = self._sample_report_data()
        first = self.reporter.fill_excel_checklist(data, file_type="Server", output_filename="quality_cache_1.xlsx")
        second = self.reporter.fill_excel_checklist(data, file_type="Server", output_filename="quality_cache_2.xlsx")
        self.assertIsInstance(first, dict)
        self.assertIsInstance(second, dict)
        self.assertTrue(first.get("generated"))
        self.assertTrue(second.get("generated"))
        self.assertIn("timings_ms", first)
        self.assertIn("copy", (first.get("timings_ms") or {}))
        self.assertIn("load", (first.get("timings_ms") or {}))
        self.assertIn("save", (first.get("timings_ms") or {}))
        self.assertTrue(second.get("template_cache_hit"))

    def test_excel_submission_contains_detail_and_verify_sheets(self):
        data = self._sample_report_data()
        output_name = "quality_sheets.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)

        from openpyxl import load_workbook

        output_path = os.path.join(self.reporter.output_dir, output_name)
        wb = load_workbook(output_path)
        self.assertIn("상세결과", wb.sheetnames)
        self.assertIn("검증 결과", wb.sheetnames)
        self.assertIn("CtrlppCheck_결과", wb.sheetnames)

    def test_ctrlpp_sheet_exists_and_checklist_ignores_p2_for_status(self):
        data = self._sample_report_with_p2_only()
        output_name = "quality_p2_only.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)

        from openpyxl import load_workbook

        output_path = os.path.join(self.reporter.output_dir, output_name)
        wb = load_workbook(output_path)
        self.assertIn("CtrlppCheck_결과", wb.sheetnames)
        ctrlpp_ws = wb["CtrlppCheck_결과"]
        self.assertGreaterEqual(ctrlpp_ws.max_row, 2)

        active = wb.active
        target_row = None
        for r in range(1, active.max_row + 1):
            if active.cell(r, 4).value == "DP Query 최적화 구현":
                target_row = r
                break
        self.assertIsNotNone(target_row)
        # P2 only finding must not flip checklist body result to NG.
        status_col = self._find_status_col(active)
        self.assertNotEqual(active.cell(target_row, status_col).value, "NG")

    def test_checklist_event_exchange_turns_ng_when_perf_ev_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-PERF-EV-01-1",
                            "rule_id": "PERF-EV-01",
                            "rule_item": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 10,
                            "message": "loop dpset burst",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_event_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "Event, Ctrl Manager 이벤트 교환 횟수 최소화")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_style_item_turns_ng_when_style_rules_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-STYLE-NAME-01-1",
                            "rule_id": "STYLE-NAME-01",
                            "rule_item": "명명 규칙 및 코딩 스타일 준수 확인",
                            "priority_origin": "P1",
                            "severity": "Low",
                            "line": 2,
                            "message": "name style",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_style_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "명명 규칙 및 코딩 스타일 준수 확인")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_unnecessary_code_turns_ng_when_clean_rules_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-CLEAN-DEAD-01-1",
                            "rule_id": "CLEAN-DEAD-01",
                            "rule_item": "불필요한 코드 지양",
                            "priority_origin": "P1",
                            "severity": "Medium",
                            "line": 20,
                            "message": "dead code",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_clean_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "불필요한 코드 지양")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_config_item_turns_ng_when_cfg_rules_found(self):
        data = {
            "file": "sample_config.ctl",
            "internal_violations": [
                {
                    "object": "sample_config.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-CFG-01-1",
                            "rule_id": "CFG-01",
                            "rule_item": "config 항목 정합성 확인",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 12,
                            "message": "cfg mismatch",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_cfg_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "config 항목 정합성 확인")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_event_item_turns_ng_when_dpget_batch_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-PERF-DPGET-BATCH-01-1",
                            "rule_id": "PERF-DPGET-BATCH-01",
                            "rule_item": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 30,
                            "message": "dpget batch",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_event_dpget_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "Event, Ctrl Manager 이벤트 교환 횟수 최소화")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_event_item_turns_ng_when_dpset_batch_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-PERF-DPSET-BATCH-01-1",
                            "rule_id": "PERF-DPSET-BATCH-01",
                            "rule_item": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 31,
                            "message": "dpset batch",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_event_dpset_batch_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "Event, Ctrl Manager 이벤트 교환 횟수 최소화")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_config_item_turns_ng_when_safe_div_found(self):
        data = {
            "file": "sample_config.ctl",
            "internal_violations": [
                {
                    "object": "sample_config.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-SAFE-DIV-01-1",
                            "rule_id": "SAFE-DIV-01",
                            "rule_item": "config 항목 정합성 확인",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 31,
                            "message": "safe div",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_cfg_safe_div_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "config 항목 정합성 확인")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_event_item_turns_ng_when_setvalue_batch_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-PERF-SETVALUE-BATCH-01-1",
                            "rule_id": "PERF-SETVALUE-BATCH-01",
                            "rule_item": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 31,
                            "message": "setvalue batch",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_event_setvalue_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "Event, Ctrl Manager 이벤트 교환 횟수 최소화")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_event_item_turns_ng_when_setmultivalue_adopt_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-PERF-SETMULTIVALUE-ADOPT-01-1",
                            "rule_id": "PERF-SETMULTIVALUE-ADOPT-01",
                            "rule_item": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 32,
                            "message": "setmultivalue adopt",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_event_setmultivalue_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "Event, Ctrl Manager 이벤트 교환 횟수 최소화")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_event_item_turns_ng_when_getvalue_batch_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-PERF-GETVALUE-BATCH-01-1",
                            "rule_id": "PERF-GETVALUE-BATCH-01",
                            "rule_item": "Event, Ctrl Manager 이벤트 교환 횟수 최소화",
                            "priority_origin": "P1",
                            "severity": "Warning",
                            "line": 33,
                            "message": "getvalue batch",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_event_getvalue_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "Event, Ctrl Manager 이벤트 교환 횟수 최소화")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_style_item_turns_ng_when_style_idx_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-STYLE-IDX-01-1",
                            "rule_id": "STYLE-IDX-01",
                            "rule_item": "명명 규칙 및 코딩 스타일 준수 확인",
                            "priority_origin": "P1",
                            "severity": "Low",
                            "line": 33,
                            "message": "index style",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_style_idx_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "명명 규칙 및 코딩 스타일 준수 확인")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_checklist_hardcoding_turns_ng_when_hard03_found(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-HARD-03-1",
                            "rule_id": "HARD-03",
                            "rule_item": "하드코딩 지양",
                            "priority_origin": "P1",
                            "severity": "Medium",
                            "line": 40,
                            "message": "float literal",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        output_name = "quality_hard03_ng.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)
        from openpyxl import load_workbook

        wb = load_workbook(os.path.join(self.reporter.output_dir, output_name))
        ws = wb.active
        row = self._find_item_row(ws, "하드코딩 지양")
        self.assertIsNotNone(row)
        status_col = self._find_status_col(ws)
        self.assertEqual(ws.cell(row, status_col).value, "NG")

    def test_detail_sheet_row_count_matches_findings(self):
        data = self._sample_report_data()
        output_name = "quality_detail_rows.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)

        from openpyxl import load_workbook

        output_path = os.path.join(self.reporter.output_dir, output_name)
        wb = load_workbook(output_path)
        ws = wb["상세결과"]
        detail_rows = ws.max_row - 1
        self.assertEqual(detail_rows, 2)

    def test_verify_sheet_columns_and_status_values(self):
        data = self._sample_report_data()
        output_name = "quality_verify_sheet.xlsx"
        self.reporter.fill_excel_checklist(data, file_type="Server", output_filename=output_name)

        from openpyxl import load_workbook

        output_path = os.path.join(self.reporter.output_dir, output_name)
        wb = load_workbook(output_path)
        ws = wb["검증 결과"]

        headers = [ws.cell(1, c).value for c in range(1, 9)]
        self.assertEqual(headers, ["No", "대분류", "중분류", "소분류", "검증 조건", "중요도", "검증 결과", "비고"])

        status_values = [
            ws.cell(r, 7).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(r, 7).value is not None
        ]
        self.assertGreater(len(status_values), 0)
        for status in status_values:
            self.assertIn(status, {"NG", "OK", "N/A"})

    def test_annotated_txt_inserts_review_comments(self):
        data = self._sample_report_data()
        source = "line1\nline2\nline3"
        self.reporter.generate_annotated_txt(source, data, "annotated.txt")

        output_path = os.path.join(self.reporter.output_dir, "annotated.txt")
        self.assertTrue(os.path.exists(output_path))

        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        self.assertIn("// [REVIEW]", content)
        self.assertIn("// >>TODO", content)

    def test_annotated_txt_inserts_todo_and_review_above_line(self):
        data = self._sample_report_data()
        source = "line1\nline2\nline3"
        self.reporter.generate_annotated_txt(source, data, "annotated_order.txt")

        output_path = os.path.join(self.reporter.output_dir, "annotated_order.txt")
        with open(output_path, "r", encoding="utf-8") as f:
            lines = f.read().splitlines()

        idx_todo = lines.index("// >>TODO")
        idx_msg = idx_todo + 1
        idx_review = idx_todo + 2
        idx_code = lines.index("line1")

        self.assertTrue(lines[idx_msg].startswith("// "))
        self.assertTrue(lines[idx_review].startswith("// [REVIEW]"))
        self.assertLess(idx_todo, idx_code)
        self.assertLess(idx_review, idx_code)

    def test_annotated_txt_review_line_avoids_duplicate_message_text(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-HARD-1",
                            "rule_id": "HARD-01",
                            "severity": "Medium",
                            "line": 1,
                            "message": "IP/URL/설정 경로 하드코딩 감지.",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        self.reporter.generate_annotated_txt("line1", data, "annotated_dedup.txt")
        output_path = os.path.join(self.reporter.output_dir, "annotated_dedup.txt")
        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        self.assertIn("// IP/URL/설정 경로 하드코딩 감지.", content)
        self.assertIn("// [REVIEW] Medium", content)
        self.assertNotIn("// [REVIEW] Medium - IP/URL/설정 경로 하드코딩 감지.", content)
        self.assertNotIn("// // 설정 파일 또는 상수로 대체 권장", content)

    def test_annotated_txt_inserts_only_ai_code_lines_for_accepted_ai_review(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "TankMgr",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-CFG-1",
                            "rule_id": "CFG-01",
                            "severity": "Warning",
                            "line": 1,
                            "message": "config 파싱 형식 불일치 가능성, delimiter/필드수 검증 권장.",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [
                {
                    "file": "sample.ctl",
                    "object": "TankMgr",
                    "event": "Global",
                    "status": "Accepted",
                    "review": "요약: config 파싱 검증을 추가하세요.\n\n코드:\n```cpp\nif (parts.size() != 6) {\n  return;\n}\n```",
                }
            ],
        }
        self.reporter.generate_annotated_txt("line1", data, "annotated_ai_code.txt")
        output_path = os.path.join(self.reporter.output_dir, "annotated_ai_code.txt")
        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        self.assertIn("// [AI CODE] if (parts.size() != 6) {", content)
        self.assertIn("// [AI CODE]   return;", content)
        self.assertNotIn("// [AI REVIEW]", content)
        self.assertNotIn("요약:", content)

    def test_unused_variable_comment_message_format(self):
        data = {
            "file": "sample.ctl",
            "internal_violations": [
                {
                    "object": "sample.ctl",
                    "event": "Global",
                    "violations": [
                        {
                            "issue_id": "P1-UNUSED-1",
                            "rule_id": "UNUSED-01",
                            "rule_item": "불필요한 코드 지양",
                            "priority_origin": "P1",
                            "severity": "Low",
                            "line": 2,
                            "message": "미사용 변수 감지: 'ip'",
                        }
                    ],
                }
            ],
            "global_violations": [],
            "ai_reviews": [],
        }
        source = "line1\n  string ip = \"192.168.0.1\";\nline3"
        self.reporter.generate_annotated_txt(source, data, "annotated_unused.txt")

        output_path = os.path.join(self.reporter.output_dir, "annotated_unused.txt")
        with open(output_path, "r", encoding="utf-8") as f:
            content = f.read()

        self.assertIn("// 미사용 변수 감지: ip", content)
        self.assertIn("// [REVIEW] Low - 사용되지 않는 변수 ip 삭제 권장", content)

    def test_duplicate_block_rule_message_matches_line_based_detection(self):
        from backend.core.heuristic_checker import HeuristicChecker

        checker = HeuristicChecker()
        code = "\n".join(
            [
                "main() {",
                "  totalValue = 1;",
                "  totalValue = 1;",
                "  totalValue = 1;",
                "}",
            ]
        )
        findings = checker.check_duplicate_blocks(code)
        self.assertGreaterEqual(len(findings), 1)
        self.assertEqual(findings[0].get("rule_id"), "CLEAN-DUP-01")
        self.assertEqual(findings[0].get("message"), "동일 코드 라인 반복(3회 이상) 감지.")


if __name__ == "__main__":
    unittest.main()
