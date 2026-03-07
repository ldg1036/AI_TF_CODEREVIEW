import os
import sys
import tempfile
import unittest
from unittest import mock

from openpyxl import Workbook

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
BACKEND_DIR = os.path.join(PROJECT_ROOT, "backend")
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

from tools import analyze_template_coverage as coverage_tool


class AnalyzeTemplateCoverageToolTests(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.TemporaryDirectory()
        self.project_root = self.tmp_dir.name

    def tearDown(self):
        self.tmp_dir.cleanup()

    def test_fail_soft_returns_skipped_payload_when_openpyxl_missing(self):
        with mock.patch.object(coverage_tool, "_resolve_load_workbook", side_effect=RuntimeError("openpyxl is required")):
            output_path, payload = coverage_tool.analyze_template_coverage(
                self.project_root,
                ensure_openpyxl=False,
                fail_soft=True,
            )

        self.assertTrue(os.path.isfile(output_path))
        self.assertEqual(payload.get("status"), "skipped_optional_missing")
        self.assertEqual(payload.get("missing_dependency"), "openpyxl")
        self.assertIn("openpyxl", str(payload.get("reason", "")))

    def test_without_fail_soft_raises_runtime_error(self):
        with mock.patch.object(coverage_tool, "_resolve_load_workbook", side_effect=RuntimeError("openpyxl is required")):
            with self.assertRaises(RuntimeError):
                coverage_tool.analyze_template_coverage(
                    self.project_root,
                    ensure_openpyxl=False,
                    fail_soft=False,
                )

    def test_extract_template_rows_inherits_merged_item_cell(self):
        workbook_path = os.path.join(self.project_root, "coverage.xlsx")
        wb = Workbook()
        ws = wb.active
        ws["D16"] = "Async callback bottleneck minimized"
        ws["E16"] = "1) prior condition"
        ws["E17"] = "2) async callback bottleneck minimized"
        ws.merge_cells("D16:D17")
        wb.save(workbook_path)

        rows = coverage_tool._extract_template_rows(workbook_path, coverage_tool._resolve_load_workbook(self.project_root))

        self.assertEqual(rows[1]["item"], ws["D16"].value)

    def test_analyze_one_matches_rule_when_merged_item_is_propagated(self):
        template_rows = [
            {
                "row": 20,
                "item": "Async callback bottleneck minimized",
                "condition": "2) async callback bottleneck minimized",
            }
        ]
        rule_items = [
            {
                "item": "Async callback bottleneck minimized",
                "norm": coverage_tool._normalize("Async callback bottleneck minimized"),
            }
        ]

        result = coverage_tool._analyze_one(rule_items, template_rows)

        self.assertEqual(result["matched_rule_count"], 1)
        self.assertEqual(result["unmatched_template_rows"], [])
        self.assertEqual(result["unmatched_rules"], [])


if __name__ == "__main__":
    unittest.main()
