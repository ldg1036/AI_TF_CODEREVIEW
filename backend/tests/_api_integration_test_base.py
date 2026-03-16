import http.client
import json
import concurrent.futures
import os
import re
import sys
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import unittest
from unittest import mock
from http.server import ThreadingHTTPServer

PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
BACKEND_DIR = os.path.join(PROJECT_ROOT, "backend")
if BACKEND_DIR not in sys.path:
    sys.path.insert(0, BACKEND_DIR)

from main import CodeInspectorApp, build_arg_parser
from core.analysis_pipeline import DirectoryAnalysisPipeline
from core.errors import CtrlppDownloadError, ReviewerResponseError, ReviewerTimeoutError, ReviewerTransportError
from core.heuristic_checker import HeuristicChecker
from core.llm_reviewer import LLMReviewer
from core.reporter import Reporter
from server import BASE_DIR, CodeInspectorHandler

try:
    from openpyxl import load_workbook as _openpyxl_load_workbook
except ImportError:
    _openpyxl_load_workbook = None


def _require_openpyxl(testcase):
    if _openpyxl_load_workbook is None:
        testcase.skipTest("openpyxl is not installed; skipping Excel report validation tests")
    return _openpyxl_load_workbook


class ApiIntegrationTestBase(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = tempfile.TemporaryDirectory()
        self.data_dir = self.tmp_dir.name

        with open(os.path.join(self.data_dir, "sample.ctl"), "w", encoding="utf-8") as f:
            f.write('main() { dpSet("A.B.C", 1); }')
        with open(os.path.join(self.data_dir, "raw_input.txt"), "w", encoding="utf-8") as f:
            f.write('main() { DebugN("raw"); }')
        with open(os.path.join(self.data_dir, "server_loop.ctl"), "w", encoding="utf-8") as f:
            f.write('main() { while(1) { int x; dpGet("A.B.C", x); } }')
        with open(os.path.join(self.data_dir, "raw_loop.txt"), "w", encoding="utf-8") as f:
            f.write('main() { while(1) { int x; dpGet("A.B.C", x); } }')

        self.app = CodeInspectorApp()
        self.app.data_dir = self.data_dir
        self.app.reporter.output_base_dir = self.data_dir
        self.app.ctrl_tool.auto_install_on_missing = False
        self.app.ctrl_tool.config_binary_path = ""
        self.app.ctrl_tool.state_binary_path = ""
        self.app.ctrl_tool.tool_path = os.path.join(self.data_dir, "missing", "ctrlppcheck.exe")

        frontend_dir = os.path.join(BASE_DIR, "frontend")

        def handler(*args, **kwargs):
            return CodeInspectorHandler(*args, app=self.app, frontend_dir=frontend_dir, **kwargs)

        self.httpd = ThreadingHTTPServer(("127.0.0.1", 0), handler)
        self.port = self.httpd.server_address[1]
        self.thread = threading.Thread(target=self.httpd.serve_forever, daemon=True)
        self.thread.start()

    def tearDown(self):
        self.httpd.shutdown()
        self.httpd.server_close()
        self.thread.join(timeout=3)
        try:
            self.app.flush_deferred_excel_reports(wait=True, timeout_sec=30)
        except Exception:
            pass
        try:
            self.app._excel_report_executor.shutdown(wait=True, cancel_futures=False)
        except Exception:
            pass
        self.tmp_dir.cleanup()

    def _request(self, method, path, body=None, headers=None):
        conn = http.client.HTTPConnection("127.0.0.1", self.port, timeout=15)
        request_headers = {"Content-Type": "application/json"}
        if isinstance(headers, dict):
            request_headers.update(headers)
        payload = json.dumps(body) if body is not None else None
        conn.request(method, path, payload, request_headers)
        resp = conn.getresponse()
        data = resp.read().decode("utf-8")
        conn.close()
        parsed = json.loads(data) if data else {}
        return resp.status, parsed

    def _request_raw(self, method, path, body=None, headers=None):
        conn = http.client.HTTPConnection("127.0.0.1", self.port, timeout=15)
        request_headers = {}
        if body is not None:
            request_headers["Content-Type"] = "application/json"
        if isinstance(headers, dict):
            request_headers.update(headers)
        payload = json.dumps(body) if body is not None else None
        conn.request(method, path, payload, request_headers)
        resp = conn.getresponse()
        data = resp.read()
        response_headers = dict(resp.getheaders())
        conn.close()
        return resp.status, data, response_headers

    @staticmethod
    def _extract_rule_items(payload):
        p1_groups = payload.get("violations", {}).get("P1", [])
        items = []
        for group in p1_groups:
            for violation in group.get("violations", []):
                item = violation.get("rule_item")
                if item:
                    items.append(item)
        return items

    @staticmethod
    def _collect_rule_ids(groups):
        found = []
        for group in groups or []:
            for violation in group.get("violations", []) or []:
                rule_id = violation.get("rule_id")
                if rule_id:
                    found.append(str(rule_id))
        return found

    def _analyze_rule_ids_for_code(self, code: str, *, file_type: str = "Server"):
        groups = self.app.checker.analyze_raw_code("sample.ctl", code, file_type=file_type)
        return self._collect_rule_ids(groups)

    def _force_single_internal_violation(self):
        self.app.checker.analyze_raw_code = lambda *_args, **_kwargs: [
            {
                "object": "sample.ctl",
                "event": "Global",
                "violations": [
                    {
                        "issue_id": "P1-R1-1",
                        "rule_id": "R1",
                        "rule_item": "test-item",
                        "priority_origin": "P1",
                        "severity": "Warning",
                        "line": 1,
                        "message": "test violation",
                    }
                ],
            }
        ]

    def _install_temp_rule_config(self, p1_rows, parsed_rows=None, review_applicability=None):
        config_dir = os.path.join(self.data_dir, "Config")
        os.makedirs(config_dir, exist_ok=True)
        with open(os.path.join(config_dir, "p1_rule_defs.json"), "w", encoding="utf-8") as f:
            json.dump(list(p1_rows), f, ensure_ascii=False, indent=2)
        with open(os.path.join(config_dir, "parsed_rules.json"), "w", encoding="utf-8") as f:
            json.dump(list(parsed_rows or []), f, ensure_ascii=False, indent=2)
        with open(os.path.join(config_dir, "review_applicability.json"), "w", encoding="utf-8") as f:
            json.dump(review_applicability or {"items": {}}, f, ensure_ascii=False, indent=2)

        self.app.checker = HeuristicChecker(os.path.join(config_dir, "parsed_rules.json"))
        reporter = Reporter(config_dir=config_dir)
        reporter.output_base_dir = self.data_dir
        self.app.reporter = reporter
        return config_dir
