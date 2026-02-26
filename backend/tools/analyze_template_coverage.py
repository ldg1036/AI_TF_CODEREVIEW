import argparse
import datetime
import json
import math
import os
import re
from typing import Dict, List, Tuple

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def _normalize(text: str) -> str:
    return re.sub(r"[\s\r\n]+", "", str(text or "").lower())


def _is_empty_item(value) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    text = str(value).strip()
    return text == "" or text.lower() in {"nan", "none"}


def _load_json(path: str):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def _find_project_root() -> str:
    return os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))


def _load_config(project_root: str) -> Dict:
    config_path = os.path.join(project_root, "Config", "config.json")
    if not os.path.exists(config_path):
        return {}
    payload = _load_json(config_path)
    return payload if isinstance(payload, dict) else {}


def _extract_template_rows(template_path: str) -> List[Dict]:
    wb = load_workbook(template_path, data_only=True)
    ws = wb.active

    start_row = 16
    for row in range(1, min(ws.max_row, 80) + 1):
        value = ws.cell(row=row, column=2).value
        if value and "대분류" in str(value):
            start_row = row + 1
            break

    rows: List[Dict] = []
    for row in range(start_row, ws.max_row + 1):
        item = ws.cell(row=row, column=4).value
        condition = ws.cell(row=row, column=5).value
        if _is_empty_item(item) and _is_empty_item(condition):
            continue
        rows.append(
            {
                "row": row,
                "item": "" if _is_empty_item(item) else str(item).strip(),
                "condition": "" if _is_empty_item(condition) else str(condition).strip(),
            }
        )
    return rows


def _load_rules_by_type(parsed_rules_path: str) -> Dict[str, List[Dict]]:
    payload = _load_json(parsed_rules_path)
    by_type: Dict[str, List[Dict]] = {"Client": [], "Server": []}
    for row in payload:
        if not isinstance(row, dict):
            continue
        rule_type = str(row.get("type", "")).strip()
        item = row.get("item")
        if rule_type not in by_type or _is_empty_item(item):
            continue
        text = str(item).strip()
        by_type[rule_type].append(
            {
                "item": text,
                "norm": _normalize(text),
            }
        )
    return by_type


def _analyze_one(rule_items: List[Dict], template_rows: List[Dict]) -> Dict:
    unmatched_rows = []
    matched_rule_indexes = set()

    for row in template_rows:
        row_text = f"{row['item']} {row['condition']}".strip()
        norm_row = _normalize(row_text)
        matched = False
        for idx, rule in enumerate(rule_items):
            norm_rule = rule["norm"]
            if not norm_rule:
                continue
            if norm_rule in norm_row or (norm_row and norm_row in norm_rule):
                matched_rule_indexes.add(idx)
                matched = True
        if not matched:
            unmatched_rows.append(row)

    unmatched_rules = [rule_items[idx]["item"] for idx in range(len(rule_items)) if idx not in matched_rule_indexes]
    return {
        "template_row_count": len(template_rows),
        "rule_count": len(rule_items),
        "matched_rule_count": len(matched_rule_indexes),
        "rule_coverage_pct": round((len(matched_rule_indexes) / len(rule_items) * 100.0), 1) if rule_items else 0.0,
        "unmatched_template_rows": unmatched_rows,
        "unmatched_rules": unmatched_rules,
    }


def analyze_template_coverage(project_root: str) -> Tuple[str, Dict]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl")

    config = _load_config(project_root)
    paths = config.get("paths", {}) if isinstance(config.get("paths"), dict) else {}

    parsed_rules_path = os.path.join(project_root, "Config", "parsed_rules.json")
    if not os.path.exists(parsed_rules_path):
        raise FileNotFoundError(f"parsed_rules.json not found: {parsed_rules_path}")

    client_template = os.path.join(project_root, paths.get("client_template", ""))
    server_template = os.path.join(project_root, paths.get("server_template", ""))
    if not os.path.exists(client_template):
        raise FileNotFoundError(f"Client template not found: {client_template}")
    if not os.path.exists(server_template):
        raise FileNotFoundError(f"Server template not found: {server_template}")

    rules_by_type = _load_rules_by_type(parsed_rules_path)
    client_rows = _extract_template_rows(client_template)
    server_rows = _extract_template_rows(server_template)

    result = {
        "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "inputs": {
            "parsed_rules": os.path.relpath(parsed_rules_path, project_root),
            "client_template": os.path.relpath(client_template, project_root),
            "server_template": os.path.relpath(server_template, project_root),
        },
        "coverage": {
            "Client": _analyze_one(rules_by_type.get("Client", []), client_rows),
            "Server": _analyze_one(rules_by_type.get("Server", []), server_rows),
        },
    }

    output_dir = os.path.join(project_root, "CodeReview_Report")
    os.makedirs(output_dir, exist_ok=True)
    stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = os.path.join(output_dir, f"template_coverage_{stamp}.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    return output_path, result


def main():
    parser = argparse.ArgumentParser(description="Analyze template coverage against parsed review rules")
    parser.add_argument("--project-root", default=None, help="Project root path (default: auto-detect)")
    args = parser.parse_args()

    project_root = os.path.abspath(args.project_root) if args.project_root else _find_project_root()
    output_path, result = analyze_template_coverage(project_root)

    client = result["coverage"]["Client"]
    server = result["coverage"]["Server"]
    print(f"[+] Coverage report saved: {output_path}")
    print(
        "[Client] rules={rule_count}, matched={matched_rule_count}, coverage={rule_coverage_pct}%".format(**client)
    )
    print(
        "[Server] rules={rule_count}, matched={matched_rule_count}, coverage={rule_coverage_pct}%".format(**server)
    )


if __name__ == "__main__":
    main()
