"""
Excel 데이터 파서 및 로더 (08_ADR_실제_Excel_양식_계약.md 기준).

Client 및 Server 체크리스트형 Excel 양식을 읽고
헤더 17행, B:H열 데이터, 병합셀 상속, 범위(Client: 15개, Server: 20개)를 파싱합니다.
"""

from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class RawExcelRow:
    """Excel 1개 데이터 행의 파싱 결과."""

    row_index: int
    category: str
    subcategory: str
    check_item: str
    condition: str
    first_review_status: str
    review_result: str
    remark: str
    source_key: str


def normalize_string(val: Any) -> str:
    """문자열의 공백 및 줄바꿈을 정규화합니다."""
    if val is None:
        return ""
    s = str(val).strip()
    s = re.sub(r"\s+", " ", s)
    return s


def make_source_key(category: str, subcategory: str, check_item: str) -> str:
    """B/C/D 열 정규화 조합으로 source_key를 생성합니다."""
    cat = normalize_string(category)
    sub = normalize_string(subcategory)
    item = normalize_string(check_item)
    return f"{cat}|{sub}|{item}"


def get_merged_cell_value(sheet: Worksheet, row: int, col: int) -> Any:
    """
    지정한 (row, col) 셀의 값을 읽어오되, 병합셀인 경우 좌상단 셀의 값을 반환합니다.
    """
    cell = sheet.cell(row=row, column=col)
    if type(cell).__name__ == "MergedCell" or cell.coordinate in sheet.merged_cells:
        for rng in sheet.merged_cells.ranges:
            if cell.coordinate in rng:
                top_left_cell = sheet.cell(row=rng.min_row, column=rng.min_col)
                return top_left_cell.value
    return cell.value


def calculate_file_sha256(file_path: Path) -> str:
    """파일의 SHA256 해시를 계산합니다."""
    hasher = hashlib.sha256()
    with open(file_path, "rb") as f:
        while chunk := f.read(8192):
            hasher.update(chunk)
    return hasher.hexdigest()


class ExcelRuleLoader:
    """Client / Server Excel 체크리스트 파서."""

    CLIENT_SHEET_NAME = "(클라이언트) 코드 리뷰 결과서"
    SERVER_SHEET_NAME = "(서버) 코드 리뷰 결과서"

    HEADER_ROW = 17
    DATA_START_ROW = 18

    CLIENT_END_ROW = 32  # 18~32 (15개 항목)
    SERVER_END_ROW = 37  # 18~37 (20개 항목)

    @classmethod
    def find_header_and_columns(cls, sheet: Worksheet) -> tuple[int, dict[str, int]]:
        """
        상단 1~30행을 스캔하여 헤더 행 및 각 컬럼별 열 인덱스를 동적으로 탐지합니다.
        탐지 실패 시 기본 좌표(HEADER_ROW=17, B~H열)를 반환합니다.
        """
        cols_map = {
            "category": 2,
            "subcategory": 3,
            "check_item": 4,
            "condition": 5,
            "first_review": 6,
            "review_result": 7,
            "remark": 8,
        }
        header_row = cls.HEADER_ROW

        for r in range(1, min(30, sheet.max_row + 1)):
            row_vals = [normalize_string(sheet.cell(row=r, column=c).value) for c in range(1, 15)]
            if any("소분류" in v or "점검항목" in v or "check item" in v.lower() for v in row_vals):
                header_row = r
                found_map = {}
                for c_idx, val in enumerate(row_vals, start=1):
                    v_lower = val.lower()
                    if "대분류" in val or "category" in v_lower:
                        found_map["category"] = c_idx
                    elif "중분류" in val or "subcategory" in v_lower:
                        found_map["subcategory"] = c_idx
                    elif "소분류" in val or "점검항목" in val or "check item" in v_lower:
                        found_map["check_item"] = c_idx
                    elif "검증 조건" in val or "condition" in v_lower or "검증조건" in val:
                        found_map["condition"] = c_idx
                    elif "1차 검증" in val or "first review" in v_lower or "1차검증" in val:
                        found_map["first_review"] = c_idx
                    elif "검증 결과" in val or "review result" in v_lower or "검증결과" in val:
                        found_map["review_result"] = c_idx
                    elif "비고" in val or "remark" in v_lower:
                        found_map["remark"] = c_idx

                cols_map.update(found_map)
                break

        return header_row, cols_map

    @classmethod
    def load_excel(cls, file_path: Path, sheet_name: str | None = None) -> tuple[list[RawExcelRow], str]:
        """
        Excel 파일을 읽고 파싱된 RawExcelRow 목록과 파일의 SHA256 해시를 반환합니다.

        Args:
            file_path: Excel 파일 경로
            sheet_name: 시트명 (None일 경우 자동 판별 또는 기본 시트 사용)

        Returns:
            (RawExcelRow 목록, sha256_hash)
        """
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundError(f"Excel 파일을 찾을 수 없습니다: {path}")

        file_sha256 = calculate_file_sha256(path)

        wb = openpyxl.load_workbook(path, data_only=True)

        target_sheet_name = sheet_name
        if not target_sheet_name:
            if cls.CLIENT_SHEET_NAME in wb.sheetnames:
                target_sheet_name = cls.CLIENT_SHEET_NAME
            elif cls.SERVER_SHEET_NAME in wb.sheetnames:
                target_sheet_name = cls.SERVER_SHEET_NAME
            else:
                target_sheet_name = wb.active.title

        sheet: Worksheet = wb[target_sheet_name]

        # 동적 헤더 및 컬럼 위치 탐지
        header_row, cols_map = cls.find_header_and_columns(sheet)
        start_row = header_row + 1

        is_client = target_sheet_name == cls.CLIENT_SHEET_NAME
        is_server = target_sheet_name == cls.SERVER_SHEET_NAME

        if is_client and header_row == cls.HEADER_ROW:
            end_row = cls.CLIENT_END_ROW
        elif is_server and header_row == cls.HEADER_ROW:
            end_row = cls.SERVER_END_ROW
        else:
            end_row = sheet.max_row

        rows: list[RawExcelRow] = []
        source_key_counts: dict[str, int] = {}

        for r in range(start_row, end_row + 1):
            category_val = get_merged_cell_value(sheet, r, cols_map["category"])
            subcategory_val = get_merged_cell_value(sheet, r, cols_map["subcategory"])
            check_item_val = get_merged_cell_value(sheet, r, cols_map["check_item"])
            condition_val = sheet.cell(row=r, column=cols_map["condition"]).value
            first_review_val = sheet.cell(row=r, column=cols_map["first_review"]).value
            review_result_val = sheet.cell(row=r, column=cols_map["review_result"]).value
            remark_val = sheet.cell(row=r, column=cols_map["remark"]).value

            # 빈 행 필터링 (D열 소분류가 비어있는 경우 제외)
            if check_item_val is None or str(check_item_val).strip() == "":
                continue

            category = normalize_string(category_val)
            subcategory = normalize_string(subcategory_val)
            check_item = normalize_string(check_item_val)
            condition = normalize_string(condition_val)
            first_review = normalize_string(first_review_val)
            review_result = normalize_string(review_result_val)
            remark = normalize_string(remark_val)

            base_key = make_source_key(category, subcategory, check_item)
            count = source_key_counts.get(base_key, 0) + 1
            source_key_counts[base_key] = count

            if count > 1:
                source_key = f"{base_key}_{count}"
            else:
                source_key = base_key

            rows.append(
                RawExcelRow(
                    row_index=r,
                    category=category,
                    subcategory=subcategory,
                    check_item=check_item,
                    condition=condition,
                    first_review_status=first_review,
                    review_result=review_result,
                    remark=remark,
                    source_key=source_key,
                )
            )

        wb.close()
        return rows, file_sha256

