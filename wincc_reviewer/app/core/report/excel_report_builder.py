"""
Excel 품질 검수 보고서 빌더 (TRD 및 Phase 11 납품 서식 준수).
openpyxl을 사용하여 납품용 Excel 종합 검수 명세서를 생성합니다.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.models import ReviewReport, SeverityLevel


class ExcelReportBuilder:
    """납품용 Excel 품질 검수 보고서 생성기."""

    @staticmethod
    def export_excel(report: ReviewReport, output_path: Path) -> Path:
        """
        ReviewReport 데이터를 바탕으로 납품용 Excel(.xlsx) 보고서를 생성합니다.

        Args:
            report: 파이프라인 검사 결과 리포트 객체
            output_path: 저장할 엑셀 파일 경로

        Returns:
            저장된 파일 경로 (Path)
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        wb = Workbook()

        # -------------------------------------------------------------
        # 스타일 정의
        # -------------------------------------------------------------
        title_font = Font(name="맑은 고딕", size=16, bold=True, color="1E3A8A")
        header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
        sub_header_font = Font(name="맑은 고딕", size=11, bold=True, color="1E293B")
        bold_font = Font(name="맑은 고딕", size=10, bold=True)
        normal_font = Font(name="맑은 고딕", size=10)

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        summary_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        # 심각도별 배경 색상
        severity_fills = {
            SeverityLevel.CRITICAL: PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            SeverityLevel.HIGH: PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid"),
            SeverityLevel.MEDIUM: PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
            SeverityLevel.LOW: PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid"),
            SeverityLevel.INFO: PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
        }

        thin_border_side = Side(border_style="thin", color="CBD5E1")
        border_thin = Border(
            left=thin_border_side,
            right=thin_border_side,
            top=thin_border_side,
            bottom=thin_border_side,
        )

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        # -------------------------------------------------------------
        # 시트 1: 검수 요약 보고서
        # -------------------------------------------------------------
        ws_summary = wb.active
        ws_summary.title = "검수 요약 보고서"

        ws_summary["A1"] = "WinCC OA 소스 코드 품질 검수 요약 보고서"
        ws_summary["A1"].font = title_font

        ws_summary["A3"] = "■ 기본 검사 정보"
        ws_summary["A3"].font = sub_header_font

        info_data = [
            ("검사 실행 ID", report.run_id),
            ("검사 룰셋 출처", report.rule_source),
            ("총 대상 파일 수", len(report.files)),
            ("위반 파일 수", len(report.errors) if hasattr(report, "errors") else 0),
            ("총 위반 항목 수", len(report.violations)),
        ]

        for idx, (label, val) in enumerate(info_data, start=4):
            ws_summary.cell(row=idx, column=1, value=label).font = bold_font
            ws_summary.cell(row=idx, column=1).fill = summary_fill
            ws_summary.cell(row=idx, column=1).border = border_thin

            val_cell = ws_summary.cell(row=idx, column=2, value=str(val))
            val_cell.font = normal_font
            val_cell.border = border_thin

        ws_summary["A10"] = "■ 심각도별 위반 통계"
        ws_summary["A10"].font = sub_header_font

        headers_stat = ["심각도 (Severity)", "위반 건수 (Violations)", "비율 (%)"]
        for col_num, h_text in enumerate(headers_stat, start=1):
            cell = ws_summary.cell(row=11, column=col_num, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_thin

        sev_counts: dict[SeverityLevel, int] = {
            SeverityLevel.CRITICAL: 0,
            SeverityLevel.HIGH: 0,
            SeverityLevel.MEDIUM: 0,
            SeverityLevel.LOW: 0,
            SeverityLevel.INFO: 0,
        }
        for v in report.violations:
            sev = v.severity if isinstance(v.severity, SeverityLevel) else SeverityLevel(v.severity)
            sev_counts[sev] = sev_counts.get(sev, 0) + 1

        total_v = len(report.violations)
        row_cur = 12
        for sev, count in sev_counts.items():
            ratio = (count / total_v * 100) if total_v > 0 else 0.0
            c1 = ws_summary.cell(row=row_cur, column=1, value=sev.value)
            c2 = ws_summary.cell(row=row_cur, column=2, value=count)
            c3 = ws_summary.cell(row=row_cur, column=3, value=f"{ratio:.1f}%")

            c1.font = bold_font
            c1.alignment = align_center
            c2.font = normal_font
            c2.alignment = align_right
            c3.font = normal_font
            c3.alignment = align_right

            c1.border = border_thin
            c2.border = border_thin
            c3.border = border_thin
            row_cur += 1

        # -------------------------------------------------------------
        # 시트 2: 위반 상세 명세서
        # -------------------------------------------------------------
        ws_detail = wb.create_sheet(title="위반 상세 명세서")

        headers_detail = ["No", "규칙 ID", "심각도", "대상 파일 경로", "라인 번호", "위반 상세 내용 및 가이드", "코드 스니펫"]
        for col_num, h_text in enumerate(headers_detail, start=1):
            cell = ws_detail.cell(row=1, column=col_num, value=h_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_thin

        for idx, v in enumerate(report.violations, start=1):
            row_num = idx + 1
            sev = v.severity if isinstance(v.severity, SeverityLevel) else SeverityLevel(v.severity)

            c_no = ws_detail.cell(row=row_num, column=1, value=idx)
            c_rule = ws_detail.cell(row=row_num, column=2, value=v.rule_id)
            c_sev = ws_detail.cell(row=row_num, column=3, value=sev.value)
            c_file = ws_detail.cell(row=row_num, column=4, value=str(v.file_id))
            c_line = ws_detail.cell(row=row_num, column=5, value=v.line_start)
            c_msg = ws_detail.cell(row=row_num, column=6, value=v.message)
            c_snip = ws_detail.cell(row=row_num, column=7, value=v.snippet or "")

            c_no.alignment = align_center
            c_rule.alignment = align_center
            c_sev.alignment = align_center
            c_sev.fill = severity_fills.get(sev, summary_fill)
            c_file.alignment = align_left
            c_line.alignment = align_center
            c_msg.alignment = align_left
            c_snip.alignment = align_left

            for cell in (c_no, c_rule, c_sev, c_file, c_line, c_msg, c_snip):
                cell.font = normal_font
                cell.border = border_thin

        # 열 폭 자동 조정
        for ws in (ws_summary, ws_detail):
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        wb.save(output_path)
        return output_path
