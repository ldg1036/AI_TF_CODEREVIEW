"""
납품용 Excel / PDF 품질 검수 보고서 내보내기 모듈 (Phase 11) 단위 테스트 스위트.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import pytest

from app.core.models import Metrics, ParseStatus, ParseStatusType, ReviewReport, SeverityLevel, Violation, ViolationStatus
from app.core.report.excel_report_builder import ExcelReportBuilder
from app.core.report.pdf_report_builder import PDFReportBuilder


class TestPDFExcelExport:
    """Excel 및 PDF 보고서 내보내기 모듈 검증."""

    @pytest.fixture
    def sample_report(self) -> ReviewReport:
        violation1 = Violation(
            violation_id="V-CTL_PRF_001-001",
            rule_id="CTL_PRF_001",
            file_id="scripts/pump_control.ctl",
            status=ViolationStatus.FAIL,
            severity=SeverityLevel.HIGH,
            message="[CTL_PRF_001] dpConnect 호출 후 dpDisconnect 누락",
            line_start=15,
            line_end=15,
            snippet="dpConnect(\"cb\", \"tag1\");",
        )
        violation2 = Violation(
            violation_id="V-MANUAL-002-010",
            rule_id="MANUAL-002",
            file_id="panels/main.pnl",
            status=ViolationStatus.FAIL,
            severity=SeverityLevel.CRITICAL,
            message="[MANUAL-002] while(1) 무한 루프 내 delay 누락",
            line_start=30,
            line_end=30,
            snippet="while(1) { i++; }",
        )

        return ReviewReport(
            schema_version="1.0.0",
            run_id="run-test-100",
            rule_source="Test Ruleset",
            files=["scripts/pump_control.ctl", "panels/main.pnl"],
            violations=[violation1, violation2],
            errors=[],
            metrics=Metrics(file_count=2, violation_count=2),

        )

    def test_excel_report_builder_export(self, sample_report: ReviewReport, tmp_path: Path):
        excel_path = tmp_path / "quality_report.xlsx"
        res_path = ExcelReportBuilder.export_excel(sample_report, excel_path)

        assert res_path.exists()
        assert res_path.stat().st_size > 0

        # openpyxl로 파일 로드하여 시트 구조 검증
        from openpyxl import load_workbook
        wb = load_workbook(res_path)
        sheet_names = wb.sheetnames
        assert "검수 요약 보고서" in sheet_names
        assert "위반 상세 명세서" in sheet_names

        ws_detail = wb["위반 상세 명세서"]
        assert ws_detail.max_row >= 3  # 헤더 1행 + 위반 2행

    def test_pdf_report_builder_export(self, sample_report: ReviewReport, tmp_path: Path):
        pdf_path = tmp_path / "quality_certificate.pdf"
        res_path = PDFReportBuilder.export_pdf(sample_report, pdf_path)

        assert res_path.exists()
        assert res_path.stat().st_size > 0
