"""
동적 엑셀 파서, API 키 환경변수 지원, AI 폴백 및 마스킹 유틸리티 검증 테스트.
"""

from pathlib import Path

import openpyxl
from app.core.rules.excel_rule_loader import ExcelRuleLoader
from app.utils.log_masker import mask_api_key, mask_code_snippet


def test_log_masker():
    """로그 마스킹 유틸리티 정상 작동 검증."""
    assert mask_api_key("sk_test_1234567890") == "sk_***7890"
    assert mask_api_key("short") == "***"
    assert mask_api_key(None) == ""

    code = "line1\nline2\nline3\nline4\nline5\nline6"
    masked = mask_code_snippet(code, max_lines=3)
    assert "[CODE_SNIPPET_MASKED" in masked


def test_find_header_and_columns_dynamic(tmp_path: Path):
    """헤더 행 위치와 컬럼 순서가 변경된 엑셀 파싱 동적 탐지 테스트."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "(클라이언트) 코드 리뷰 결과서"

    # 상단 5행을 공백 처리 후 6행에 헤더 작성 (행 위치 변경)
    ws.cell(row=6, column=1, value="No")
    ws.cell(row=6, column=2, value="대분류")
    ws.cell(row=6, column=3, value="중분류")
    ws.cell(row=6, column=4, value="소분류 (점검항목)")
    ws.cell(row=6, column=5, value="검증 조건")
    ws.cell(row=6, column=6, value="1차 검증")
    ws.cell(row=6, column=7, value="검증 결과")
    ws.cell(row=6, column=8, value="비고")

    # 7행에 데이터 작성
    ws.cell(row=7, column=2, value="클라이언트")
    ws.cell(row=7, column=3, value="스크립트")
    ws.cell(row=7, column=4, value="dpConnect 콜백 선언 검사")
    ws.cell(row=7, column=5, value="dpConnect 호출 시 콜백 함수 존재 필수")
    ws.cell(row=7, column=6, value="적용")
    ws.cell(row=7, column=7, value="PASS")
    ws.cell(row=7, column=8, value="정상")

    excel_path = tmp_path / "dynamic_test.xlsx"
    wb.save(excel_path)
    wb.close()

    header_row, cols_map = ExcelRuleLoader.find_header_and_columns(ws)
    assert header_row == 6
    assert cols_map["category"] == 2
    assert cols_map["check_item"] == 4

    rows, _ = ExcelRuleLoader.load_excel(excel_path)
    assert len(rows) == 1
    assert rows[0].check_item == "dpConnect 콜백 선언 검사"
